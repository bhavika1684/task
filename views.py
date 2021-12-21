from django.shortcuts import render
from django.http import HttpResponse,HttpResponseRedirect,JsonResponse
from .forms import DocumentForm,DeclarationForm,VendorForm,VendorUploadDocumentForm,hrDocumentForm
import os
import re
from django.db import connection
import asyncio
from .models import Document,Declaration_form,Vendor_form,Vendor_files,Resignation_data,hr_document_data,pf_gratuity_data,twl_log_data,twillio_email,health_ins_fix_premium,health_ins_details,health_ins_prem_price,greetings_data_details
from django.views import View
from contextlib import closing
from xlplatpy.XlplatMiddleware import XlplatAuthMiddleware
import base64
import shutil
import tkinter
import subprocess
import requests
from requests.auth import HTTPBasicAuth 
from urllib.request import urlopen, Request
from django.template.defaulttags import register
from datetime import datetime,timedelta
from dateutil.relativedelta import relativedelta
from datetime import date
import json
import glob
import pdfkit
from django.conf import settings
import threading
import tempfile
from django.core.files.storage import FileSystemStorage
from django.core.files.base import ContentFile
from django.views.decorators.http import require_http_methods


# from twilio.twiml.voice_response import Conference, Dial, VoiceResponse

from django.views.decorators.csrf import csrf_exempt
from twilio.rest import Client
from twilio.twiml.voice_response import Dial, VoiceResponse, Say, Dial

# from twilio.twiml.voice_response import Conference, Dial, VoiceResponse
from django.db.models import Q,CharField
from django.db.models.functions import Lower
from openpyxl import Workbook,load_workbook


class Upload_documents(View):
	def get(self,request):
		user_id_val = XlplatAuthMiddleware.user_id 
		if user_id_val != 0 :
			final_val='1'
		else:
			final_val=0 
		if user_id_val == 0 or final_val == '1' :
			user_id_value=check_for_permission_assigned(user_id_val)
			if user_id_value == '1':
				hr_view=1
			else:
				hr_view=0

			user_or_cand=request.GET['user_id']
			candidate_hr="0"
			if user_or_cand != "":
				candidate_id = base64.b64decode(request.GET['user_id']).decode('utf-8')
				cand_employee=1
				requisition_id=""
				if str(user_id_val)!= str(candidate_id):
					candidate_hr=check_hr_permission(candidate_id,user_id_val)
			else:
				candidate_id = base64.b64decode(request.GET['c_id']).decode('utf-8')
				requisition_id = base64.b64decode(request.GET['req_id']).decode('utf-8')
				cand_employee=0
				if str(user_id_val)!= str(candidate_hr):
					with closing(connection.cursor()) as cursor:
						cursor.execute("SELECT user_id FROM xl_ats_req_candidate_map where c_id= %s", [candidate_id])
						candidate_hr = cursor.fetchone()
					candidate_hr=check_hr_permission(candidate_hr,user_id_val)


			if cand_employee == 1:
				cand_or_user_type="user_id"
				resume_folder_name="e_"+str(candidate_id)
			else:
				cand_or_user_type="c_id"
				resume_folder_name="c_"+str(candidate_id)
				
			all_doc_check=Document.objects.filter(uploaded_by=candidate_id)
			
			
			all_document_name =["Other","Voter","Adhar","PAN","Passport","Driving","10th","Relieving","Employee","Exit","Employment","ID","Declaration","PF","Gratuity","Internship","Tax","Fit","Resignation","Salary","Bank","12th","Post","Graduation","Experience","Resume"]
			check_doc_name_list=[]
			if all_doc_check.count() <=26:
				for ed in all_doc_check:
					check_doc_name_list.append(ed.document_name)
				for each_doc in all_document_name:
					if each_doc in check_doc_name_list:
						continue
					else:
						Document.objects.update_or_create(uploaded_by=candidate_id, document_name=each_doc,file_uploaded='',verified='False',user_or_cid=cand_or_user_type)
			
			try:		
				resume_check=Document.objects.get(uploaded_by=candidate_id,document_name='Resume').file_uploaded
			except:
				resume_check="none"
			if resume_check == "none":
				with closing(connection.cursor()) as cursor:
					cursor.execute("select resume_filename from xl_ats_candidate_details where c_id=%s",[candidate_id])
					cand_resume_check = cursor.fetchone()

				if cand_resume_check == None or cand_resume_check[0] == None:
					cand_resume_check = ""
				else:
					cand_resume_check=cand_resume_check[0]

				if cand_resume_check == "" or cand_resume_check == "no_resume":
					# if not exist in candidate_sheet
					if "redirect" in request.GET:
						pass
					else:
						Document.objects.update_or_create(uploaded_by=candidate_id, document_name='Resume',file_uploaded='',verified='False',user_or_cid=cand_or_user_type)
				else:
					flag=1
					if "." not in cand_resume_check:
						extention_name = cand_resume_check.split("_")
						extention_name = extention_name[len(extention_name)-1]
						if extention_name.lower()== "pdf" or extention_name.lower()== "doc" or extention_name.lower()== "docx" :
							cand_resume_check_final=cand_resume_check+"."+extention_name
						else:
							flag=0
					else:
						cand_resume_check_final=cand_resume_check

					if flag==1:
						#if exist in candidate_sheet
						with closing(connection.cursor()) as cursor:
							cursor.execute("select req_id from xl_ats_req_candidate_map where c_id=%s",[candidate_id])
							copy_req_id = cursor.fetchone()
						copy_req_id=copy_req_id[0]
						server_base_path=get_server_folder_path()
						if os.path.exists('{}/filestorage/resumes/{}/{}/{}'.format(server_base_path,candidate_id,copy_req_id,cand_resume_check)) != 1: 
							Document.objects.update_or_create(uploaded_by=candidate_id, document_name='Resume',file_uploaded='',verified='False',user_or_cid=cand_or_user_type)
						else:
							if os.path.exists('media/xlplat_ats/{}/Resume'.format(resume_folder_name)) != 1: 
								if os.path.exists('media/xlplat_ats/{}'.format(resume_folder_name)) != 1:
									os.mkdir('media/xlplat_ats/{}'.format(resume_folder_name))	
								os.mkdir('media/xlplat_ats/{}/Resume'.format(resume_folder_name))
							shutil.copy('{}/filestorage/resumes/{}/{}/{}'.format(server_base_path,candidate_id,copy_req_id,cand_resume_check),'media/xlplat_ats/{}/Resume/{}'.format(resume_folder_name,cand_resume_check_final))
							Document.objects.update_or_create(uploaded_by=candidate_id, document_name='Resume',file_uploaded=cand_resume_check_final,verified='False',user_or_cid=cand_or_user_type)



			files_uploaded_data = Document.objects.filter(uploaded_by=candidate_id)
			verify_documents_files=Document.objects.filter(uploaded_by=candidate_id,verified='True')
			if cand_employee == 0:
				cand_name=candidate_name(candidate_id)
				cand_f_m_l_name=final_cand_name(cand_name)
				candidate_user_id=cid_userid_exist(candidate_id,requisition_id)
				if candidate_user_id != None and  None not in candidate_user_id:
					candidate_user_id=candidate_user_id[0]
					candidate_end_date=employee_end_date(candidate_user_id)
					only_end_date=datetime.strptime(candidate_end_date.split(" ")[1],'%Y-%m-%d')
					three_days_prior = only_end_date - timedelta(days=3)
					today = datetime.now()
					if three_days_prior <= today:
						show_exit_docs=1
					else:
						show_exit_docs=0
				else:
					show_exit_docs=0
			else:
				cand_f_m_l_name=fetch_employee_name(candidate_id)
				candidate_end_date=employee_end_date(candidate_id)
				only_end_date=datetime.strptime(candidate_end_date.split(" ")[1],'%Y-%m-%d')
				three_days_prior = only_end_date - timedelta(days=3)
				today = datetime.now()
				if three_days_prior <= today:
					show_exit_docs=1
				else:
					show_exit_docs=0
			verified_dict=[]
			for each_verify in verify_documents_files:
				verify_file_name=each_verify.document_name
				verified_dict.append(verify_file_name)
			dict=[]
			Other_Documents=[]
			document_with_file_name={}
			temp_other_file=""
			exp_letter=[]
			temp_exp_file=""
			itr_exp_docs=0
			itr_other_docs=0
			for each_uploaded_file in files_uploaded_data:
				document_name = each_uploaded_file.document_name
				# print(document_name)
				if str(document_name)!="Other" and str(document_name)!="Experience":
					filename_val=each_uploaded_file.file_uploaded
					document_with_file_name[document_name]=filename_val
				elif str(document_name)=="Other":
					temp_other_file=each_uploaded_file.file_uploaded.name
					filename_val=temp_other_file.split(",")
					itr_other_docs=0
					for file_other in filename_val:
						if file_other != "''" and file_other != "":
							Other_Documents.append("Other_{}".format(itr_other_docs))
							document_with_file_name["Other_{}".format(itr_other_docs)]=file_other
							itr_other_docs+=1
				elif str(document_name)=="Experience":
					# print("eeeeeeeeeeeee")
					temp_exp_file=each_uploaded_file.file_uploaded.name
					filename_val=temp_exp_file.split(",")
					itr_exp_docs=0
					for file_other in filename_val:
						if file_other != "''" and file_other != "":
							# print(temp_exp_file)
							exp_letter.append("Experience_{}".format(itr_exp_docs))
							document_with_file_name["Experience_{}".format(itr_exp_docs)]=file_other
							itr_exp_docs+=1
					
				dict.append(document_name)
			dict_length=len(dict)
			identity_documents = ["Resume","Adhar card", "PAN card", "Voter card", "Passport", "Driving License"]
			educational_documents = ["10th Marksheet", "12th Marksheet", "Graduation (Marksheets of all Semesters + Degree Certificate)", "Post Graduation (Marksheets of all Semesters+Degree Certificate)"]
			bank_certificates = ["Bank Details (Front Page of Passbook/Cancelled Cheque)",
								 "Salary Slips/ Banks statement(last 3 months, if applicable)",
								 "Resignation Email Received/Work Experience & Relieving Letter (If Applicable)", "Fit To Work Certificate (Issued from a Registered Medical Practitioner)",
								 "Tax Declaration Form 12 BB /Form 16/Salary Certificate"]
			joining_documents=["Declaration Form","ID Card","PF Form","Gratuity Form","Internship Agreement/ Offer Letter/ NDA","Employment Agreement/ Appointment Letter/ NDA","Other Documents"]
			exit_documents=["Exit Interview Form","Employee Clearance Form","Relieving Letters"]
			experience_letter = ["Experience letter"]
			all_docs=[]
			all_docs.append(identity_documents)
			all_docs.append(educational_documents)
			all_docs.append(bank_certificates)
			all_docs.append(joining_documents)
			all_docs.append(experience_letter)
			if exp_letter:
				all_docs.append(exp_letter)
			if Other_Documents:
				all_docs.append(Other_Documents)
			if show_exit_docs == 1:
				all_docs.append(exit_documents)
			final_val=str(final_val) 
			user_id_val=str(user_id_val)
			# if dict_length >= 26 and user_id_val != "0" and user_id_val.isdigit():
			# if (dict_length >= 26 and user_id_val != "0" and user_id_val.isdigit()) or hr_view==1 or user_id_val!=0:
			if requisition_id!="":
				with closing(connection.cursor()) as cursor:
					query_temp="SELECT organization FROM xl_ats_req_candidate_map where c_id= {} and req_id ={}".format(candidate_id,requisition_id)
					cursor.execute(query_temp)
					org_code = cursor.fetchone()
				org_code=org_code[0]
			elif user_id_val!="": 
				ttc_code=fetch_ttc_code(user_id_val)
				if "TTC-" in ttc_code or "TTC " in ttc_code:
					org_code=701
				elif "XLP" in ttc_code:
					org_code=703
				elif "TTCS" in ttc_code or "TTCs" in ttc_code:
					org_code=707
				elif "TA" in ttc_code:
					org_code=705
				elif "XLS" in ttc_code:
					org_code=709
				else:
					org_code=701
			else:
				org_code=701
			if (dict_length >= 26 and user_id_val != "0" and user_id_val.isdigit()) or hr_view==1 or user_id_val!="0":
				if candidate_hr=='1':
					return render(self.request, 'insufficient_privilages.html', {'redirect_to':'/xlplat-ats/tentative_joining.tcl'})
				else:
					return render(self.request,'uploaded_files.html',{"all_docs": all_docs,"exp_docs_val":itr_exp_docs,"other_docs_val":itr_other_docs, "final_val":final_val, "c_id":candidate_id, "req_id":requisition_id, "verified_dict":verified_dict, "document_with_file_name": document_with_file_name, "candidate_name":cand_f_m_l_name,"hr_view":hr_view,"cand_or_user_type":cand_or_user_type,"user_id_val":user_id_val,"org_code":org_code})
			else:
				# print(dict)
				# print(document_with_file_name)
				# print("gggggggggggggeeeeeeeeeeeettttttt",temp_other_file,temp_exp_file)
				document_with_file_name["Other"]=temp_other_file
				# document_with_file_name["Experience"]=temp_exp_file
				content={
					"title": "Upload Your Documents",
					"identity_docs": identity_documents,
					"educational_docs": educational_documents,
					"bank_docs": bank_certificates,
					"joining_docs": joining_documents,
					"exit_docs":exit_documents,
					"c_id": candidate_id,
					"req_id": requisition_id,
					"files_uploaded_data_db": files_uploaded_data,
					"dict": dict,
					"document_with_file_name": document_with_file_name,
					"candidate_name":cand_f_m_l_name,
					"final_val": final_val,
					"hr_view":hr_view,
					"cand_or_user_type":cand_or_user_type,
					"show_exit_documents":show_exit_docs,
					"last_document_type":"identity_documents",
					"org_code":org_code

				}
			return render(self.request,'upload_documents.html',content)
		else:
			return render(self.request, 'insufficient_privilages.html', {'redirect_to':'/xlplat-ats/tentative_joining.tcl'})

	def post(self,request):
		user_id_val = XlplatAuthMiddleware.user_id
		if user_id_val != "":
			final_val = '1'
		else: 
			final_val = 0
		if user_id_val == 0 or final_val == '1' :
			user_id_value=check_for_permission_assigned(user_id_val)
			if user_id_value == '1':
				hr_view=1
			else:
				hr_view=0
			candidate_id = request.POST['c_id']
			cand_or_user_type=request.POST['new_or_employee']
			requisition_id= request.POST['req_id']
			filename=request.POST['file_name']
			verify_documents=request.POST['verify_upload'] 
			upload_or_na_value=request.POST['upload_or_na']
			last_document_type=""
			if "last_document_type" in request.POST:
				last_document_type=request.POST['last_document_type']
			identity_documents = ["Resume","Adhar card", "PAN card", "Voter card", "Passport", "Driving License"]
			educational_documents = ["10th Marksheet", "12th Marksheet", "Graduation (Marksheets of all Semesters + Degree Certificate)", "Post Graduation (Marksheets of all Semesters+Degree Certificate)"]
			bank_certificates = ["Bank Details (Front Page of Passbook/Cancelled Cheque)",
								 "Salary Slips/ Banks statement(last 3 months, if applicable)",
								 "Resignation Email Received/Work Experience & Relieving Letter (If Applicable)", "Fit To Work Certificate (Issued from a Registered Medical Practitioner)",
								 "Tax Declaration Form 12 BB /Form 16/Salary Certificate"]
			experience_letter = ["Experience letter"]
			if request.method=="POST": 
				candidate_hr='0'
				if cand_or_user_type == "user_id":
					if str(user_id_val)!=str(candidate_id):
						candidate_hr=check_hr_permission(candidate_id,user_id_val)
					cand_f_m_l_name=fetch_employee_name(candidate_id)
					candidate_end_date=employee_end_date(candidate_id)
					only_end_date=datetime.strptime(candidate_end_date.split(" ")[1],'%Y-%m-%d')
					three_days_prior = only_end_date - timedelta(days=3)
					today = datetime.now()
					if three_days_prior <= today:
						show_exit_docs=1
					else:
						show_exit_docs=0
					add_to_folder="e_"
				else:
					if str(user_id_val)!=str(candidate_id):
						with closing(connection.cursor()) as cursor:
							cursor.execute("SELECT user_id FROM xl_ats_req_candidate_map where c_id= %s", [candidate_id])
							candidate_hr = cursor.fetchone()
						candidate_hr=check_hr_permission(candidate_hr,user_id_val)
					cand_name=candidate_name(candidate_id)
					cand_f_m_l_name=final_cand_name(cand_name)
					candidate_user_id=cid_userid_exist(candidate_id,requisition_id)
					if candidate_user_id != None and  None not in candidate_user_id:
						candidate_user_id=candidate_user_id[0]
						candidate_end_date=employee_end_date(candidate_user_id)
						only_end_date=datetime.strptime(candidate_end_date.split(" ")[1],'%Y-%m-%d')
						three_days_prior = only_end_date - timedelta(days=3)
						today = datetime.now()
						if three_days_prior <= today:
							show_exit_docs=1
						else:
							show_exit_docs=0
					else:
						show_exit_docs=0

					add_to_folder="c_"
				if "delete_file" in request.POST:
					try:
						del_doc_type=request.POST["del_doc_type"]
						# print(del_doc_type,"oooooooooooooooooo")
						file_name_database=Document.objects.get(uploaded_by=candidate_id,document_name=del_doc_type).file_uploaded.name
						if file_name_database != "''":
							file_name_database=file_name_database.replace(filename,"").replace(",,",",").strip(",")
							data_dict={}
							data_dict["file_uploaded"]=file_name_database
							if file_name_database == "":
								data_dict["verified"]='False'
							Document.objects.filter(uploaded_by=candidate_id,document_name=del_doc_type).update(**data_dict)
							if os.path.exists('media/xlplat_ats/{}{}/{}/{}'.format(add_to_folder,candidate_id,del_doc_type,filename)):
								os.remove('media/xlplat_ats/{}{}/{}/{}'.format(add_to_folder,candidate_id,del_doc_type,filename))
					except Document.DoesNotExist:
						pass
				elif upload_or_na_value != "0" or "replace_value" in request.POST:
					form=DocumentForm(self.request.POST, self.request.FILES)
					if form.is_valid():
						pdf=form.save(commit=False)
						pdf.uploaded_by=candidate_id
						pdf.user_or_cid=cand_or_user_type
						cand_id_exist=os.path.exists('media/xlplat_ats/{}{}'.format(add_to_folder,candidate_id))
						if cand_id_exist == True:
							cand_file_exist=os.path.exists('media/xlplat_ats/{}{}/{}'.format(add_to_folder,candidate_id,filename))
							if cand_file_exist == True:
								if final_val == "1" and filename!="Other": 
									shutil.rmtree('media/xlplat_ats/{}{}/{}'.format(add_to_folder,candidate_id,filename))
									os.mkdir(os.path.join('media/xlplat_ats/{}{}'.format(add_to_folder,candidate_id), filename))
								else: 
									pass
							else:
								os.mkdir(os.path.join('media/xlplat_ats/{}{}'.format(add_to_folder,candidate_id), filename))
						else:
							os.mkdir(os.path.join('media/xlplat_ats/',"{}{}".format(add_to_folder,candidate_id)))
							os.mkdir(os.path.join('media/xlplat_ats/{}{}'.format(add_to_folder,candidate_id), filename))
						
						pdf.file_uploaded.storage.location =(os.path.join('media/xlplat_ats/{}{}/{}'.format(add_to_folder,candidate_id,filename)))
						try:
						    file_already_uploaded=Document.objects.get(uploaded_by=candidate_id,document_name=filename)
						except Document.DoesNotExist:
						    file_already_uploaded = None
						data_updated_temp=""
						
						if filename=="Other":
							try:
								data_updated_temp=Document.objects.get(uploaded_by=candidate_id,document_name=filename).file_uploaded
							except Document.DoesNotExist:
								data_updated_temp=""
						
						if str(file_already_uploaded) == str(filename):
							data_update=Document.objects.get(uploaded_by=candidate_id,document_name=filename).delete()
						pdf.document_name=filename
						pdf.save()
						data= {'is_valid': True, 'name': os.path.basename(pdf.file_uploaded.name), 'url': pdf.file_uploaded.url } 
						if str(file_already_uploaded) == str(filename) and filename=="Other" and data_updated_temp!="" :
							file_name_new=Document.objects.get(uploaded_by=candidate_id,document_name=filename).file_uploaded
							Document.objects.filter(uploaded_by=candidate_id,document_name=filename).update(file_uploaded="{},{}".format(file_name_new.name,data_updated_temp.name))
						if hr_view == 0:
							notify_cand_name = fetch_employee_name(candidate_id)
							file_name_notify=Document.objects.get(uploaded_by=candidate_id,document_name=filename).file_uploaded
							if notify_cand_name == '<unknown>':
								notify_cand_name=candidate_name(candidate_id)
								notify_cand_name=final_cand_name(notify_cand_name)
								# print(notify_cand_name,file_name_notify.name)
							with closing(connection.cursor()) as cursor:
								cursor.execute("insert into xl_ats_document_notification (employee_name,employee_doc_type,employee_doc) values(%s,%s,%s)",[str(notify_cand_name),str(filename),str(file_name_notify.name)])

				else:
					data= {'is_valid': False }
					if verify_documents != "verified" and upload_or_na_value != "1":
						try:
							file_already_uploaded=Document.objects.get(uploaded_by=candidate_id,document_name=filename)
						except Document.DoesNotExist:
							file_already_uploaded = None
						if str(file_already_uploaded) == str(filename):
							data_update=Document.objects.filter(uploaded_by=candidate_id,document_name=filename).get(document_name=filename).delete()
						pdf=Document(uploaded_by=candidate_id, document_name=filename, user_or_cid=cand_or_user_type)
						pdf.save()

				files_uploaded_data = Document.objects.filter(uploaded_by=candidate_id)
				joining_documents=["Declaration Form","ID Card","PF Form","Gratuity Form","Internship Agreement/ Offer Letter/ NDA","Employment Agreement/ Appointment Letter/ NDA","Other Documents"]
				exit_documents=["Exit Interview Form","Employee Clearance Form","Relieving Letters"]
				Other_Documents=[]
				dict = []
				# print("ppppppppppppppppppppp")
				document_with_file_name={}
				exp_letter=[]
				temp_exp_file=""
				temp_other_file=""
				itr_exp_docs=0
				for each_uploaded_file in files_uploaded_data:
					document_name = each_uploaded_file.document_name
					if str(document_name)!="Other" and str(document_name)!="Experience":
						filename_val=each_uploaded_file.file_uploaded
						document_with_file_name[document_name]=filename_val
					elif str(document_name)=="Other":
						temp_other_file=each_uploaded_file.file_uploaded.name
						filename_val=temp_other_file.split(",")
						itr_other_docs=0
						for file_other in filename_val:
							if file_other != "''" and file_other != "":
								Other_Documents.append("Other_{}".format(itr_other_docs))
								document_with_file_name["Other_{}".format(itr_other_docs)]=file_other
								itr_other_docs+=1
					elif str(document_name)=="Experience":
						# print("eeeeeeeeeeeee")
						temp_exp_file=each_uploaded_file.file_uploaded.name
						filename_val=temp_exp_file.split(",")
						itr_exp_docs=0
						for file_other in filename_val:
							if file_other != "''" and file_other != "":
								# print(temp_exp_file)
								exp_letter.append("Experience_{}".format(itr_exp_docs))
								document_with_file_name["Experience_{}".format(itr_exp_docs)]=file_other
								itr_exp_docs+=1
								
					dict.append(document_name)
				dict_length = len(dict)
				# print("dddd",dict_length)
				all_docs = []
				verified_dict= []
				all_docs.append(identity_documents)
				all_docs.append(educational_documents)
				all_docs.append(bank_certificates)
				all_docs.append(joining_documents)
				all_docs.append(experience_letter)
				if exp_letter:
					all_docs.append(exp_letter)
				if Other_Documents:
					all_docs.append(Other_Documents)
				if show_exit_docs == 1:
					all_docs.append(exit_documents)
				user_id_val=str(user_id_val)
				if final_val == "1" and dict_length >= 26:
					if verify_documents == "verified":
						cand_verify_data = candidate_verify_data(candidate_id, filename, requisition_id)
						verified_files = Document.objects.filter(uploaded_by=candidate_id, verified='True')
						for each_verified in verified_files:
							verified_file_name = each_verified.document_name
							verified_dict.append(verified_file_name)
						if cand_verify_data == "updated" and user_id_val != "0" and user_id_val.isdigit() :
							if candidate_hr=='1':
								return render(self.request, 'insufficient_privilages.html', {'redirect_to':'/xlplat-ats/tentative_joining.tcl'})
							else:
								verified_value=1
								return render(self.request, 'uploaded_files.html',
										  {"all_docs": all_docs,"other_docs_val":itr_other_docs, "final_val": final_val, "c_id": candidate_id,
										   "req_id": requisition_id, "verified":verified_value, "verified_dict":verified_dict, "document_with_file_name": document_with_file_name, "candidate_name":cand_f_m_l_name,"hr_view":hr_view,"cand_or_user_type":cand_or_user_type,"user_id_val":user_id_val })


				if (dict_length >= 26 and user_id_val != "0" and user_id_val.isdigit()) or hr_view==1 or user_id_val!="0":
					if candidate_hr=='1':
						return render(self.request, 'insufficient_privilages.html', {'redirect_to':'/xlplat-ats/tentative_joining.tcl'})
					else:
						verified_files = Document.objects.filter(uploaded_by=candidate_id, verified='True')
						for each_verified in verified_files:
							verified_file_name = each_verified.document_name
							verified_dict.append(verified_file_name)
						return render(self.request, 'uploaded_files.html',{"all_docs": all_docs,"exp_docs_val":itr_exp_docs,"other_docs_val":itr_other_docs, "final_val": final_val, "c_id": candidate_id,"req_id": requisition_id,"verified_dict":verified_dict, "document_with_file_name": document_with_file_name, "candidate_name":cand_f_m_l_name,"hr_view":hr_view,"cand_or_user_type":cand_or_user_type,"user_id_val":user_id_val })
				else:
					document_with_file_name["Other"]=temp_other_file
					content = {
						"title": "Upload Your Documents",
						"identity_docs": identity_documents,
						"educational_docs": educational_documents,
						"bank_docs": bank_certificates,
						"joining_docs": joining_documents,
						"exit_docs": exit_documents,
						"c_id": candidate_id,
						"req_id": requisition_id,
						"files_uploaded_data_db": files_uploaded_data,
						"dict": dict,
						"document_with_file_name": document_with_file_name,
						"candidate_name": cand_f_m_l_name,
						"final_val": final_val,
						"hr_view":hr_view,
						"cand_or_user_type":cand_or_user_type,
						"show_exit_documents":show_exit_docs,
						"last_document_type":last_document_type,
					}
					return render(self.request,'upload_documents.html',content)
		else:
			return render(self.request, 'insufficient_privilages.html',{'redirect_to':'/intranet/'})


def fetch_ttc_code(employee_id):
	with closing(connection.cursor()) as cursor:
		cursor.execute("SELECT ttc_code from im_employees where employee_id=%s",[employee_id])
		employee_name = cursor.fetchone()
	if employee_name == None or employee_name[0] == None:
		employee_name = "<unknown>"
	else:
		employee_name=employee_name[0]
	return employee_name 


def check_hr_permission(employee_id,user_id):
	with closing(connection.cursor()) as cursor:
		cursor.execute("select permission as temp_val from im_payroll_permissions_special where user_id=%s and permission=1 limit 1",[employee_id])
		permission = cursor.fetchone()
	if permission == None or permission[0] == None:
		permission_val = "0"
	else:
		with closing(connection.cursor()) as cursor:
			cursor.execute("select permission as temp_val from im_payroll_permissions_special where user_id=%s and permission=2 limit 1",[user_id])
			permission = cursor.fetchone()
			if permission == None or permission[0] == None:
				permission_val = "1"
			else:
				permission_val="0"

	return permission_val 

class map_appraisal(View):
	def get(self,request):
		user_id_val =XlplatAuthMiddleware.user_id
		flg=0
		if user_id_val !=0:
			user_id_value=check_for_permission_assigned(user_id_val)
			if user_id_value == '1':
				hr_view=1
			else:
				hr_view=0
			if hr_view == 1:
				year_list = [datetime.now().year,datetime.now().year-1]
				content = {
				"year_list":year_list
				}
				return render(self.request,'map_appraisal.html',content)
			else:
				return render(self.request, 'insufficient_privilages.html',{'redirect_to':'/intranet/'})
		else:
			return render(self.request, 'insufficient_privilages.html',{'redirect_to':'/intranet/'})

	def post(self,request):
		user_id_val=XlplatAuthMiddleware.user_id
		flg=0
		not_successfull_div=""
		successfull_div=""
		upload_status = ""
		appr_emp_id=""
		appr_emp_name=""
		appr_selected_year = ""
		if user_id_val !=0:
			user_id_value=check_for_permission_assigned(user_id_val)
			if user_id_value == '1':
				hr_view=1
			else:
				hr_view=0
			if hr_view==1:
				if 'upload_status' in request.POST and request.method == 'POST':
					upload_status=request.POST['upload_status']
					if upload_status=="uploading":
						temp_filename = request.FILES['file']
						if os.path.exists('media/xlplat_ats/temp_appraisal') != 1:
							os.mkdir('media/xlplat_ats/temp_appraisal')
						with open(os.path.join('media/xlplat_ats/temp_appraisal',str(temp_filename)), 'wb+') as destination:
							for chunk in temp_filename.chunks():
								destination.write(chunk)
						flg=1
					else:
						flg=2
				elif 'button' in request.POST and request.method == 'POST':
					flg=3
					s_no_success=0
					s_no_error=0
					btn_value='Map appraisal to database'
					for filename in os.listdir('media/xlplat_ats/temp_appraisal'):
						lst=filename.split('.')
						fl_name=lst[0].replace(" ","").replace("_","")
						if len(lst)>1:
							fl_ext=lst[1]
						else:
							fl_ext="pdf"
						with closing(connection.cursor()) as cursor:
							cursor.execute("SELECT count(*) FROM persons WHERE REPLACE(LOWER(TRIM(first_names) || TRIM(last_name)),' ','') = %s", [fl_name.strip().lower()])
							check_for_unique_name = cursor.fetchone()
						if check_for_unique_name[0]>1:
							s_no_error+=1
							not_successfull_div+="<p style='padding:5px;font-size:15px;'>"+ str(s_no_error) +" . "+ str(filename) +"<a style='color:#ff5400' id='error_$s_no_error'><b> multiple profile exist</b></a></p>"
							os.remove(os.path.join('media/xlplat_ats/temp_appraisal',filename))
							continue

						with closing(connection.cursor()) as cursor:
							cursor.execute("SELECT person_id FROM persons WHERE REPLACE(LOWER(TRIM(first_names) || TRIM(last_name)),' ','')  = %s", [fl_name.strip().lower()])
							cand_e_id = cursor.fetchone()
						
						if cand_e_id != None:
							with closing(connection.cursor()) as cursor:
								cursor.execute("SELECT c_id FROM xl_ats_req_candidate_map where user_id= %s", [cand_e_id[0]])
								cand_c_id = cursor.fetchone()
							if cand_c_id != None:
								cand_userid = cand_c_id[0]
								appraisal_folder_name = 'c_'+str(cand_userid)
								id_type = 'c_id'
							else:
								cand_userid = cand_e_id[0]
								appraisal_folder_name = 'e_'+str(cand_userid)
								id_type = 'user_id'
							try:
								file_exist=Document.objects.get(uploaded_by=cand_userid,document_name='Other').file_uploaded
							except Document.DoesNotExist:
								file_exist="None"
							if file_exist != "None":
								appraisal_exist=re.search("appraisal_{}".format(request.POST['selected_year']),str(file_exist))
								if appraisal_exist != None:
									s_no_error+=1
									not_successfull_div+="<p style='padding:5px;font-size:15px;'>"+ str(s_no_error) +" . "+ str(filename) +"<a style='color:#ff9800' id='error_$s_no_error'><b> appraisal letter already exist</b></a></p>"
									os.remove(os.path.join('media/xlplat_ats/temp_appraisal',filename))
								else:
									s_no_success+=1
									successfull_div+="<p style='padding:5px;font-size:15px;'>"+ str(s_no_success) +" . "+ str(filename) +"<a style='color:#61C503'><b> sucessfully uploaded</b></a></p>"
									shutil.move(os.path.join('media/xlplat_ats/temp_appraisal',filename),'media/xlplat_ats/{}/Other/appraisal_{}.{}'.format(appraisal_folder_name,request.POST['selected_year'],fl_ext))
									Document.objects.filter(uploaded_by=cand_userid,document_name='Other').update(file_uploaded="appraisal_{}.{},{}".format(request.POST['selected_year'],fl_ext,file_exist.name))
							else:
								s_no_success+=1
								successfull_div+="<p style='padding:5px;font-size:15px;'>"+ str(s_no_success) +" . "+ str(filename) +"<a style='color:#61C503'><b> sucessfully uploaded</b></a></p>"
								if os.path.exists('media/xlplat_ats/{}/Other'.format(appraisal_folder_name)) != 1: 
									if os.path.exists('media/xlplat_ats/{}'.format(appraisal_folder_name)) != 1:
										os.mkdir('media/xlplat_ats/{}'.format(appraisal_folder_name))	
									os.mkdir('media/xlplat_ats/{}/Other'.format(appraisal_folder_name))
								shutil.move(os.path.join('media/xlplat_ats/temp_appraisal',filename),'media/xlplat_ats/{}/Other/appraisal_{}.{}'.format(appraisal_folder_name,request.POST['selected_year'],fl_ext))
								Document.objects.update_or_create(uploaded_by=cand_userid, document_name='Other',defaults={"file_uploaded":"appraisal_{}.{}".format(request.POST['selected_year'],fl_ext),"verified":'False',"user_or_cid":id_type})
						else:
							s_no_error+=1
							not_successfull_div+="<p style='padding:5px;font-size:15px;'>"+ str(s_no_error) +" . "+ str(filename) +"<a style='color:#ff0000' id='error_$s_no_error'><b> does not have respective employee</b></a></p>"
							os.remove(os.path.join('media/xlplat_ats/temp_appraisal',filename))
					#1:for successfull and 2 is for unccessfull
					return JsonResponse({1:successfull_div,2:not_successfull_div})

				elif 'send_appraisal' in request.POST and request.method == 'POST':
					flg=4
					appr_emp_id=request.POST['appraisal_emp_id']
					appr_emp_name=fetch_employee_name(appr_emp_id)
					if appr_emp_name == "<unknown>":
						with closing(connection.cursor()) as cursor:
							cursor.execute("SELECT user_id FROM xl_ats_req_candidate_map where c_id= %s", [appr_emp_id])
							appr_c_id = cursor.fetchone()
						appr_emp_name=fetch_employee_name(appr_c_id)
					appr_selected_year=request.POST['appraisal_year']

				year_list = [datetime.now().year,datetime.now().year-1]
				content = {
				"flag":flg,
				"status": upload_status,
				"successfull_div": successfull_div,
				"not_successfull_div": not_successfull_div,
				"year_list":year_list,
				"appr_emp_id":appr_emp_id,
				"appr_emp_name":appr_emp_name,
				"appr_selected_year":appr_selected_year
				}
				return render(self.request,'map_appraisal.html',content)
			else:
				return render(self.request, 'insufficient_privilages.html',{'redirect_to':'/intranet/'})
		else:
			return render(self.request, 'insufficient_privilages.html',{'redirect_to':'/intranet/'})

def get_all_hr_people():
	hr_team=[]
	with closing(connection.cursor()) as cursor:
		cursor.execute("select user_id from im_payroll_permissions_special where permission=1")
		all_hr_tem=cursor.fetchall()
		for each_active_employee in all_hr_tem:
			each_active = int(replace_text(str(each_active_employee),1))
			hr_team.append(each_active)
	return hr_team 

def get_all_view_people():
	hr_team=[]
	with closing(connection.cursor()) as cursor:
		cursor.execute("select user_id from im_payroll_permissions_special where permission=2")
		all_hr_tem=cursor.fetchall()
		for each_active_employee in all_hr_tem:
			each_active = str(replace_text(str(each_active_employee),1))
			hr_team.append(each_active)
	return hr_team 

class map_appraisal_mail(View):
	def get(self,request):
		user_id_val =XlplatAuthMiddleware.user_id
		flg=0
		cand_appraisal_exist = {}
		cand_appraisal_not_exist = {}
		mail_sent_list = []
		cid_list = []
		eid_list = []
		ttc_code_arr = {}
		active_cand_id_list = active_employees()
		year_list = [datetime.now().year,datetime.now().year-1]
		if user_id_val !=0:
			user_id_value=check_for_permission_assigned(user_id_val)
			if user_id_value == '1':
				hr_view=1
			else:
				hr_view=0
			if hr_view == 1:
				with closing(connection.cursor()) as cursor:
					cursor.execute("select user_id from xl_ats_appraisal_mail where appraisal_year = %s",[request.GET['selected_year']])
					m_user_id_list = cursor.fetchall()
				for m_user_id in m_user_id_list:
					with closing(connection.cursor()) as cursor:
						cursor.execute("SELECT user_id FROM xl_ats_req_candidate_map where c_id= %s and user_id is not null", [m_user_id])
						cand_u_id = cursor.fetchone()
					if cand_u_id != None:
						m_user_id=cand_u_id
					m_user_name = fetch_employee_name(m_user_id)
					mail_sent_list.append(m_user_name)
				cand_userid=0
				hr_group=get_all_hr_people()
				accounts_group=get_all_view_people()
				for emp_id in active_cand_id_list:
					if emp_id in hr_group:
						if str(user_id_val) not in accounts_group:
							continue
					emp_name = fetch_employee_name(emp_id)
					ttc_code=fetch_ttc_code(emp_id)
					if ttc_code == "<unknown>" or str(ttc_code) == "None":
						ttc_code="-"
					if emp_name == "<unknown>":
						emp_name=""

					with closing(connection.cursor()) as cursor:
						cursor.execute("SELECT c_id FROM xl_ats_req_candidate_map where user_id= %s", [emp_id])
						cand_c_id = cursor.fetchone()
					if cand_c_id != None:
						emp_id=cand_c_id[0]
						cid_list.append(emp_id)
					else:
						eid_list.append(emp_id)
					ttc_code_arr[emp_id]=ttc_code
					try:
						file_exist=Document.objects.get(uploaded_by=emp_id,document_name='Other').file_uploaded
					except Document.DoesNotExist:
						file_exist=""
					if file_exist != "":
						appraisal_exist=""
						appraisal_exist=re.search("appraisal_{}".format(request.GET['selected_year']),str(file_exist))
						if appraisal_exist != None:
							cand_appraisal_exist.update( {emp_name : emp_id} )
						else:
							cand_appraisal_not_exist.update( {emp_name : emp_id} )
					else:
						cand_appraisal_not_exist.update( {emp_name : emp_id} )
				content = {
				"app_exist" : cand_appraisal_exist,
				"app_not_exist" : cand_appraisal_not_exist,
				"selected_year":request.GET['selected_year'],
				"mail_sent_list":mail_sent_list,
				"year_list":year_list,
				"cid_list":cid_list,
				"eid_list":eid_list,
				"ttc_code_arr":ttc_code_arr
				}
				
				return render(self.request,'map_appraisal_mail.html',content)
			else:
				return render(self.request, 'insufficient_privilages.html',{'redirect_to':'/intranet/'})
		else:
			return render(self.request, 'insufficient_privilages.html',{'redirect_to':'/intranet/'})
	def post(self,request):
		user_id_val =XlplatAuthMiddleware.user_id
		flg=0
		cand_appraisal_exist = {}
		cand_appraisal_not_exist = {}
		mail_sent_list = []
		cid_list = []
		eid_list = []
		active_cand_id_list = active_employees()
		year_list = [datetime.now().year,datetime.now().year-1]
		if user_id_val !=0:
			user_id_value=check_for_permission_assigned(user_id_val)
			if user_id_value == '1':
				hr_view=1
			else:
				hr_view=0
			if hr_view == 1:
				if 'delete_flag' in request.POST:
					del_user_year = request.POST['selected_year']
					del_user = request.POST['deletion_user']
					check_href_val = request.POST['href_value']
					if 'c_' in check_href_val:
						check_href_val = check_href_val.strip('/py')
						os.remove(check_href_val)
					elif 'e_' in check_href_val:
						check_href_val = check_href_val.strip('/py')
						os.remove(check_href_val)
					try:
						file_exist=Document.objects.get(uploaded_by=del_user,document_name='Other').file_uploaded
					except Document.DoesNotExist:
						file_exist=""
					change_file_exist = str(file_exist).replace('appraisal_'+del_user_year+'.pdf','')
					change_file_exist = change_file_exist.strip(',')
					change_file_exist = change_file_exist.replace(',,',',')
					if change_file_exist == '':
						Document.objects.filter(uploaded_by = del_user,document_name='Other').delete()
					else:
						Document.objects.filter(uploaded_by = del_user,document_name='Other').update(file_uploaded=change_file_exist)
				if 'upload_single_file' in request.POST and request.method == 'POST':
					temp_filename = request.FILES['single_appraisal']
					cand_e_id=request.POST['appr_emp_id']
					with closing(connection.cursor()) as cursor:
							cursor.execute("SELECT * FROM xl_ats_req_candidate_map where c_id= %s", [cand_e_id])
							check_for_cid = cursor.fetchone()
					if check_for_cid == None:
						with closing(connection.cursor()) as cursor:
							cursor.execute("SELECT c_id FROM xl_ats_req_candidate_map where user_id= %s", [cand_e_id])
							cand_c_id = cursor.fetchone()
					else:
						cand_c_id=cand_e_id
					if cand_c_id != None:
						cand_userid = cand_c_id
						cand_userid = int(replace_text(str(cand_c_id),1))
						appraisal_folder_name = 'c_'+str(cand_userid)
						id_type = 'c_id'
					else:
						cand_userid = cand_e_id
						appraisal_folder_name = 'e_'+str(cand_userid)
						id_type = 'user_id'
					
					try:
						file_exist=Document.objects.get(uploaded_by=cand_userid,document_name='Other').file_uploaded
					except Document.DoesNotExist:
						file_exist=""
					# later we get to now that file type would be pdf no extension from filename is trimmed out 
					if file_exist != "":
						with open('media/xlplat_ats/{}/Other/appraisal_{}.{}'.format(appraisal_folder_name,request.POST['selected_year'],'pdf'), 'wb+') as destination:
							for chunk in temp_filename.chunks():
								destination.write(chunk)
						Document.objects.filter(uploaded_by=cand_userid,document_name='Other').update(file_uploaded="appraisal_{}.{},{}".format(request.POST['selected_year'],'pdf',file_exist.name))
					else:
						if os.path.exists('media/xlplat_ats/{}/Other'.format(appraisal_folder_name)) != 1:
							if os.path.exists('media/xlplat_ats/{}'.format(appraisal_folder_name)) != 1:
								os.mkdir('media/xlplat_ats/{}'.format(appraisal_folder_name))	
							os.mkdir('media/xlplat_ats/{}/Other'.format(appraisal_folder_name))
						with open('media/xlplat_ats/{}/Other/appraisal_{}.{}'.format(appraisal_folder_name,request.POST['selected_year'],'pdf'), 'wb+') as destination:
							for chunk in temp_filename.chunks():
								destination.write(chunk)
						data_document=Document(uploaded_by=cand_userid,document_name='Other',file_uploaded="appraisal_{}.{}".format(request.POST['selected_year'],'pdf'),verified='False',user_or_cid=id_type)
						data_document.save()
					# else:
					# if os.path.exists('media/xlplat_ats/temp_appraisal') != 1:
					# 	os.mkdir('media/xlplat_ats/temp_appraisal')
					# with open('media/xlplat_ats/{}/Other/appraisal_{}.{}'.format(appraisal_folder_name,request.POST['selected_year'],'.pdf'), 'wb+') as destination:
					# 	for chunk in temp_filename.chunks():
					# 		destination.write(chunk)
				if 'send_mail' in request.POST:
					cand_details_string=""
					cand_details = ""
					for i in request.POST.getlist('action[]'):
						#in candidate details we send candidate id and selected year with " _ " inbetween 
						cand_details = i +'_'+request.POST['selected_year']
						if cand_details != "":
							cand_details_string = cand_details_string + "~" + cand_details
					url="http://localhost//intranet/testing/permission-redirect"
					payload={"proc_name":"ns_appraisal_sendmail","proc_variables": cand_details_string}
					response = requests.get(url, params=payload)
				year=request.POST['selected_year']
				return HttpResponseRedirect("/py/xlplat_ats/map_appraisal_mail?selected_year={}".format(year))
			else:
				return render(self.request, 'insufficient_privilages.html',{'redirect_to':'/intranet/'})
		else:
			return render(self.request, 'insufficient_privilages.html',{'redirect_to':'/intranet/'})

class map_documents(View):
	def get(self,request):
		return render(self.request,'map_documents.html')
	def post(self,request):
		not_successfull_div=""
		successfull_div=""
		if 'upload_status' in request.POST and request.method == 'POST':
			upload_status=request.POST['upload_status']
			temp_filename = request.FILES['file']
			if os.path.exists('media/xlplat_ats/temp_map_documents') != 1:
				os.mkdir('media/xlplat_ats/temp_map_documents')
			with open(os.path.join('media/xlplat_ats/temp_map_documents',str(temp_filename)), 'wb+') as destination:
				for chunk in temp_filename.chunks():
					destination.write(chunk)
		elif 'mapping_documents' in request.POST and request.method == 'POST':
			flg=3
			s_no_success=0
			s_no_error=0
			prefix_name = request.POST['prefix_name']
			bw_name = request.POST['bw_name']
			suffix_name = request.POST['suffix_name']
			file_type = request.POST['file_type']
			upload_desired_file_name = request.POST['upload_desired_file_name']
			upload_file_location = request.POST['upload_file_location']
			fl_location = upload_file_location.split(' ')
			fl_location = fl_location[0]
			
			for filename in os.listdir('media/xlplat_ats/temp_map_documents'):
				lst=filename.split('.')
				fl_name=lst[0]
				fl_ext=lst[1]
				if bw_name == " ":
					bw_name=""
				if fl_name.startswith(prefix_name) and fl_name.endswith(suffix_name) and re.search(bw_name,fl_name) != None and file_type == '.'+fl_ext:
					# employee_name = re.sub(prefix_name, '', fl_name)
					# employee_name = re.sub(suffix_name, '', employee_name)
					# employee_name = re.sub(bw_name, '', employee_name)
					# employee_name = re.sub(' ', '', employee_name)
					fl_name=fl_name.replace(" ","")
					fl_name=fl_name.split("_")
					fl_name=fl_name[0].replace(" ","")
					
					with closing(connection.cursor()) as cursor:
						cursor.execute("SELECT count(*) FROM im_employees WHERE LOWER(TRIM(TRIM(pan_number,'	'))) = %s", [fl_name.strip().lower()])
						check_for_unique_name = cursor.fetchone()
					# with closing(connection.cursor()) as cursor:
					# 	cursor.execute("SELECT count(*) FROM persons WHERE REPLACE(LOWER(TRIM(first_names) || TRIM(last_name)),' ','') = %s", [employee_name.strip().lower()])
					# 	check_for_unique_name = cursor.fetchone()
					if check_for_unique_name[0]>1:
						s_no_error+=1
						not_successfull_div+="<p style='padding:5px;font-size:15px;'>"+ str(s_no_error) +" . "+ str(filename) +"<a style='color:#ff5400' id='error_$s_no_error'><b> multiple profile exist</b></a></p>"
						os.remove(os.path.join('media/xlplat_ats/temp_map_documents',filename))
						continue

					with closing(connection.cursor()) as cursor:
						cursor.execute("SELECT employee_id FROM im_employees WHERE LOWER(TRIM(TRIM(pan_number,'	'))) = %s", [fl_name.strip().lower()])
						cand_e_id = cursor.fetchone()
					# with closing(connection.cursor()) as cursor:
					# 	cursor.execute("SELECT person_id FROM persons WHERE REPLACE(LOWER(TRIM(first_names) || TRIM(last_name)),' ','') = %s", [employee_name.strip().lower()])
					# 	cand_e_id = cursor.fetchone()
					if cand_e_id != None:
						with closing(connection.cursor()) as cursor:
							cursor.execute("SELECT c_id FROM xl_ats_req_candidate_map where user_id= %s", [cand_e_id[0]])
							cand_c_id = cursor.fetchone()
						if cand_c_id != None:
							cand_userid = cand_c_id[0]
							emp_folder_name = 'c_'+str(cand_userid)
							id_type = 'c_id'
						else:
							cand_userid = cand_e_id[0]
							emp_folder_name = 'e_'+str(cand_userid)
							id_type = 'user_id'
						try:
							file_exist=Document.objects.get(uploaded_by=cand_userid,document_name=fl_location).file_uploaded
						except Document.DoesNotExist:
							file_exist="not_exist"
						if file_exist!="not_exist":
							if fl_location != 'Other' and file_exist != "":
								s_no_error+=1
								not_successfull_div+="<p style='padding:5px;font-size:15px;'>"+ str(s_no_error) +" . "+ str(filename) +"<a style='color:#ff9800' id='error_$s_no_error'><b> "+fl_location+" documents already exist</b></a></p>"
								os.remove(os.path.join('media/xlplat_ats/temp_map_documents',filename))
							else:
								s_no_success+=1
								#-------------------if file_exist is empty---------------------
								if os.path.exists('media/xlplat_ats/{}/{}'.format(emp_folder_name,fl_location)) != 1: 
									if os.path.exists('media/xlplat_ats/{}'.format(emp_folder_name)) != 1:
										os.mkdir('media/xlplat_ats/{}'.format(emp_folder_name))	
									os.mkdir('media/xlplat_ats/{}/{}'.format(emp_folder_name,fl_location))
								#----------------------------------------------------------------
								successfull_div+="<p style='padding:5px;font-size:15px;'>"+ str(s_no_success) +" . "+ str(filename) +"<a style='color:#61C503'><b> sucessfully uploaded</b></a></p>"
								shutil.move(os.path.join('media/xlplat_ats/temp_map_documents',filename),'media/xlplat_ats/{}/{}/{}.{}'.format(emp_folder_name,fl_location,upload_desired_file_name,fl_ext))
								if file_exist.name != "":
									Document.objects.filter(uploaded_by=cand_userid,document_name=fl_location).update(file_uploaded="{}.{},{}".format(upload_desired_file_name,fl_ext,file_exist.name))
								else:
									Document.objects.filter(uploaded_by=cand_userid,document_name=fl_location).update(file_uploaded="{}.{}".format(upload_desired_file_name,fl_ext))
						else:
							s_no_success+=1
							successfull_div+="<p style='padding:5px;font-size:15px;'>"+ str(s_no_success) +" . "+ str(filename) +"<a style='color:#61C503'><b> sucessfully uploaded</b></a></p>"
							if os.path.exists('media/xlplat_ats/{}/{}'.format(emp_folder_name,fl_location)) != 1: 
								if os.path.exists('media/xlplat_ats/{}'.format(emp_folder_name)) != 1:
									os.mkdir('media/xlplat_ats/{}'.format(emp_folder_name))	
								os.mkdir('media/xlplat_ats/{}/{}'.format(emp_folder_name,fl_location))
							shutil.move(os.path.join('media/xlplat_ats/temp_map_documents',filename),'media/xlplat_ats/{}/{}/{}.{}'.format(emp_folder_name,fl_location,upload_desired_file_name,fl_ext))
							data_document=Document(uploaded_by=cand_userid,document_name=fl_location,file_uploaded="{}.{}".format(upload_desired_file_name,fl_ext),verified='False',user_or_cid=id_type)
							data_document.save()
					else:
						s_no_error+=1
						not_successfull_div+="<p style='padding:5px;font-size:15px;'>"+ str(s_no_error) +" . "+ str(filename) +"<a style='color:#ff0000' id='error_$s_no_error'><b> does not have respective employee</b></a></p>"
						os.remove(os.path.join('media/xlplat_ats/temp_map_documents',filename))
				else:
					#incorrect format of name of document
					s_no_error+=1
					not_successfull_div+="<p style='padding:5px;font-size:15px;'>"+ str(s_no_error) +" . "+ str(filename) +"<a style='color:#ff0000' id='error_$s_no_error'><b> does not match with format</b></a></p>"
					os.remove(os.path.join('media/xlplat_ats/temp_map_documents',filename))
			#1:for successfull and 2 is for unccessfull
			return JsonResponse({1:successfull_div,2:not_successfull_div})
		return render(self.request,'map_documents.html')


class call_req(View):
	def get(self,request):
		if request.method == 'GET' and 'email' in request.GET:
			email_val = request.GET['email']
			invalid_link=0
			url="http://localhost//intranet/testing/permission-redirect"
			payload = {"proc_name": "im_transform_email2user_id", "proc_variables": email_val}
			response=requests.get(url, params=payload)
			response=response.text
			response=response.split(" ")[0]
			# print(len(response))
			if response == "{}":
				# print(response,"pppppppppppp")
				invalid_link+=1
			return render(self.request,'call_twilio_initate.html',{'email':email_val,'invalid_link':invalid_link})
		else:
			return render(self.request,'error_call.html')

def end_date_of_a_month(date):
	start_date_of_this_month=date.replace(day=1)
	month = start_date_of_this_month.month
	year = start_date_of_this_month.year
	if month == 12:
	    month = 1
	    year += 1
	else:
	    month += 1
	next_month_start_date = start_date_of_this_month.replace(month=month, year=year)

	this_month_end_date = next_month_start_date - timedelta(days=1)
	return this_month_end_date
class health_track_sys(View):
	CharField.register_lookup(Lower)
	def get(self,request):
		ins_tbl_html=""
		tbl_indx=0
		cur_month=datetime.now().month
		c_date=datetime.now()
		last_date_next=end_date_of_a_month(c_date)
		if cur_month >=1 and cur_month<=5:
			this_sheet_name="sheet_"+str((datetime.now().year)-1)+"_"+str(datetime.now().year)
		else:	
			this_sheet_name="sheet_"+str(datetime.now().year)+"_"+str((datetime.now().year)+1)
		year_list = [datetime.now().year,datetime.now().year-1]
		with closing(connection.cursor()) as cursor:
			cursor.execute("select u.user_id from registered_users u, group_distinct_member_map gm where u.user_id = gm.member_id and gm.group_id = 463 order by lower(im_name_from_user_id(u.user_id))")
			active_employees_all=cursor.fetchall()
		active_emp_list=[ele[0] for ele in active_employees_all]
		emp_list_len=len(active_emp_list)
		if "view_health_detail" in request.GET:
			table_data=health_ins_details.objects.all().values("emp_id","ins_name","relation","dob","gender").order_by('emp_id')
			with closing(connection.cursor()) as cursor:
				query="Select ie.ttc_code,hid.emp_id,hid.ins_name,hid.relation,hid.dob,hid.gender from xlplat_ats_health_ins_details hid, im_employees ie where ie.employee_id = hid.emp_id order by ie.employee_id "
				cursor.execute(query)
				table_data=cursor.fetchall()
			content = {
					"year_list":year_list,
					"this_sheet_name":this_sheet_name,
					"table_data":table_data,
					"view_health_detail":1
			}	
			return render(self.request,"health_track_sys.html",content)
		else:
			table_data=health_ins_prem_price.objects.filter(sheet_label=this_sheet_name).values('emp_id','premium',"jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec","sheet_label")
			if "search_sheet" in request.GET:
				sheet_year=request.GET["search_sheet"]
				this_sheet_name="sheet_"+sheet_year
			# table_data=health_ins_prem_price.objects.filter(sheet_label=this_sheet_name).values('emp_id','premium',"actual_premium","jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec","sheet_label")

			with closing(connection.cursor()) as cursor:
				query="Select DISTINCT xhi.emp_id,xhi.premium,im_name_from_user_id(ie.employee_id),xhi.jan,xhi.feb,xhi.mar,xhi.apr,xhi.may,xhi.jun,xhi.jul,xhi.aug,xhi.sep,xhi.oct,xhi.nov,xhi.dec,xhi.sheet_label,ie.ttc_code,to_char(irc.start_date + INTERVAL '5 hours 30 minutes','dd Mon yyyy'),policy_info_from_user_id(ie.employee_id),xhi.jan+xhi.feb+xhi.mar+xhi.apr+xhi.may+xhi.jun+xhi.jul+xhi.aug+xhi.sep+xhi.oct+xhi.nov+xhi.dec-xhi.premium as r_d_price,xhi.completion_status,to_char(irc.end_date + INTERVAL '5 hours 30 minutes','dd Mon yyyy'),xhi.exit_status,hid.ref_no,irc.start_date,rcm.trainee from xlplat_ats_health_ins_prem_price xhi, im_employees ie,im_repeating_costs irc,im_costs ic,xlplat_ats_health_ins_details hid,xl_ats_req_candidate_map rcm where hid.emp_id=xhi.emp_id and hid.relation = 'self' and xhi.sheet_label='"+this_sheet_name+"' and ie.employee_id = xhi.emp_id and irc.rep_cost_id=ic.cost_id and ic.cause_object_id=xhi.emp_id and rcm.user_id=hid.emp_id order by irc.start_date DESC"
				cursor.execute(query)
				table_data=cursor.fetchall()
			# table_data=health_ins_prem_price.objects.raw('Select xhi.emp_id,ie.ttc_codefrom xlplat_ats_health_ins_prem_price xhi, im_employees ie where xhi.sheet_label={},ie_employee_id = xhi.emp_id '.format(this_sheet_name))
			table_row_count=health_ins_prem_price.objects.filter(sheet_label=this_sheet_name).count()
			# table_row_count=5
			if table_row_count!=emp_list_len:
				for each_emp_id in active_emp_list:
					if str(each_emp_id) not in str(table_data):
						health_ins_prem_price.objects.update_or_create(emp_id=each_emp_id,sheet_label=this_sheet_name)

						em_name=fetch_employee_name(each_emp_id)
						emp_b_date=None
						with closing(connection.cursor()) as cursor:
							query="select e.birthdate,e.gender from im_employees e where e.employee_id="+str(each_emp_id)+" "
							cursor.execute(query)
							emp_b_date=cursor.fetchall()
						if len(emp_b_date)==0:
							emp_b_date=None
						if emp_b_date != None:
							emp_bdate=emp_b_date[0][0]
							emp_gen=emp_b_date[0][1]
							if emp_gen == None:
								emp_gen=""
							health_ins_details.objects.update_or_create(emp_id=each_emp_id,relation="self",defaults={"ins_name":em_name,"dob":emp_bdate,"policy_info":"self","gender":emp_gen})

				# table_data=health_ins_prem_price.objects.filter(sheet_label=this_sheet_name).values('emp_id','premium',"actual_premium","jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec","sheet_label")
				# table_data=health_ins_prem_price.objects.raw("Select xhi.emp_id,ie.ttc_code from xlplat_ats_health_ins_prem_price xhi, im_employees ie where xhi.sheet_label='{}' and ie.employee_id = xhi.emp_id ".format(this_sheet_name))
				with closing(connection.cursor()) as cursor:
					query="Select DISTINCT xhi.emp_id,xhi.premium,im_name_from_user_id(ie.employee_id),xhi.jan,xhi.feb,xhi.mar,xhi.apr,xhi.may,xhi.jun,xhi.jul,xhi.aug,xhi.sep,xhi.oct,xhi.nov,xhi.dec,xhi.sheet_label,ie.ttc_code,to_char(irc.start_date + INTERVAL '5 hours 30 minutes','dd Mon yyyy'),policy_info_from_user_id(ie.employee_id),xhi.jan+xhi.feb+xhi.mar+xhi.apr+xhi.may+xhi.jun+xhi.jul+xhi.aug+xhi.sep+xhi.oct+xhi.nov+xhi.dec-xhi.premium as r_d_price,xhi.completion_status,to_char(irc.end_date + INTERVAL '5 hours 30 minutes','dd Mon yyyy'),xhi.exit_status,hid.ref_no,irc.start_date,rcm.trainee from xlplat_ats_health_ins_prem_price xhi, im_employees ie,im_repeating_costs irc,im_costs ic,xlplat_ats_health_ins_details hid, xl_ats_req_candidate_map rcm where hid.emp_id=xhi.emp_id and hid.relation = 'self' and xhi.sheet_label='"+this_sheet_name+"' and ie.employee_id = xhi.emp_id and irc.rep_cost_id=ic.cost_id and ic.cause_object_id=xhi.emp_id and rcm.user_id=hid.emp_id order by irc.start_date DESC"
					cursor.execute(query)
					table_data=cursor.fetchall()
			
			# ------------------Query for Status update(EXIT)-------------------------
			with closing(connection.cursor()) as cursor:
				query="Update xlplat_ats_health_ins_prem_price set exit_status='exit_in_process' where sheet_label='"+this_sheet_name+"' and emp_id in (select xhi.emp_id from xlplat_ats_health_ins_prem_price xhi, im_employees ie,im_repeating_costs irc,im_costs ic where xhi.sheet_label='"+this_sheet_name+"' and ie.employee_id = xhi.emp_id and irc.rep_cost_id=ic.cost_id and ic.cause_object_id=xhi.emp_id and irc.end_date<now()) and LOWER(exit_status) not like '%exit%'"
				cursor.execute(query)
			# -------------------------------------------------------------------------

		if "ref_file" in request.GET:
			if os.path.exists('media/health_insurance') != 1:
				os.mkdir('media/health_insurance')
			wb_ref = Workbook()
			ws_ref = wb_ref.active
			if request.GET["ref_file"]=="add_prem_file":
				ws_ref.title = "Add premium"
				ws_ref.append(('emp code','name','premium'))
				wb_ref.save('media/health_insurance/add_premium_refrence.xlsx')
				return HttpResponseRedirect('/py/media/health_insurance/add_premium_refrence.xlsx')
			elif request.GET["ref_file"]=="add_month_file":
				ws_ref.title = "Add month Installment"
				ws_ref.append(('emp code','name','jan'))
				wb_ref.save('media/health_insurance/add_month_installment_refrence.xlsx')
				return HttpResponseRedirect('/py/media/health_insurance/add_month_installment_refrence.xlsx')
			elif request.GET["ref_file"]=="add_emp_det_file":
				ws_ref.title = "Add Details"
				ws_ref.append(('ref_no','name','relation','dob','gender'))
				wb_ref.save('media/health_insurance/add_emp_details_refrence.xlsx')
				return HttpResponseRedirect('/py/media/health_insurance/add_emp_details_refrence.xlsx')
		if "request_file" in request.GET:
			# response = HttpResponse(content_type='text/csv')
			# response['Content-Disposition'] = 'attachment; filename="users.csv"'
			# response=""
			# writer = csv.writer(response)
			# writer.writerow(['EmpCode',	'Name of the person','Relation to employee','Designation / Grade','VIPFlag', 'Date of Birth / Age','Gender - M/F','Sum Insured(not applicable if floater selected)','Pre Existing illness Y/N','Description of ailment','Date of Inclusion (DD.MM.YYYY)','Date of Joining','Location','Remarks'])
			a_d_status=request.GET["a_d_status"]
			# if "add" in a_d_status:	
			# 	ins_pol=request.GET["policy"]
			# 	if ins_pol == "ttc":
			# 		h_details = health_ins_details.objects.select_related('health_ins_prem_price').filter(Q(add_del_status=a_d_status) & (Q(policy_info="self") | Q(policy_info="self_spouse")) ) .values('ref_no','emp_id', 'policy_info', 'relation', 'ins_name','dob','gender').order_by('emp_id')
			# 	elif ins_pol == "ttcs":
			# 		h_details = health_ins_details.objects.filter( Q(add_del_status=a_d_status) & Q(policy_info="self_family") ).values('emp_id', 'policy_info', 'relation', 'ins_name','dob','gender').order_by('emp_id')
			# 	for each_det in h_details:
			# 		emp_id=each_det['emp_id']
			# 		emp_code=fetch_ttc_code(emp_id)
			# 		emp_rel=each_det['relation']
			# 		emp_name=each_det['ins_name']
			# 		emp_dob=each_det['dob']
			# 		emp_gender=each_det['gender']
			# 		applicant_details=employee_details(emp_id)
			# 		job_title_name="NA"
			# 		joining_date="NA"
			# 		sum_assured="NA"
			# 		if emp_rel == "self":
			# 			sum_assured="300000"
			# 			job_title_name=category_data((applicant_details[1]))
			# 			joining_date=datetime.strptime(employee_end_date(emp_id).split(" ")[0],'%Y-%m-%d').date()
			# 		writer.writerow([emp_code,emp_name,emp_rel,job_title_name,'NA',emp_dob,emp_gender,sum_assured,'N','NA',joining_date,joining_date,'',''])
			# elif "download" in a_d_status:
			lst_status=a_d_status.split("_")
			start_date="01-"+lst_status[1]+"-"+lst_status[2]+" +0530"
			start_date=datetime.strptime(start_date, "%d-%m-%Y %z")
			try:
				end_date="31-"+lst_status[1]+"-"+lst_status[2]+" +0530"
				end_date=datetime.strptime(end_date, "%d-%m-%Y %z")
			except:
				try:
					end_date="30-"+lst_status[1]+"-"+lst_status[2]+" +0530"
					end_date=datetime.strptime(end_date, "%d-%m-%Y %z")
				except:
					try:
						end_date="29-"+lst_status[1]+"-"+lst_status[2]+" +0530"
						end_date=datetime.strptime(end_date, "%d-%m-%Y %z")
					except:
						try:
							end_date="28-"+lst_status[1]+"-"+lst_status[2]+" +0530"
							end_date=datetime.strptime(end_date, "%d-%m-%Y %z")
						except:
							end_date=""
			year=int(lst_status[2])
			month=int(lst_status[1])
			sheet_str="sheet_"
			if month>=6 and month<=12:
				sheet_str=sheet_str+str(year)+"_"+str(year+1)
			elif month>=1 and month<=5:
				sheet_str=sheet_str+str(year-1)+"_"+str(year)
			wb = Workbook()
			ws = wb.active
			ws.title = "Addition TTC"
			ws1 = wb.create_sheet("Deletion TTC",1)
			ws2 = wb.create_sheet("Addition TTCS",2)
			ws3 = wb.create_sheet("Deletion TTCS",3)

			ws.append(('EmpCode',	'Name of the person','Relation to employee','Designation / Grade','VIPFlag', 'Date of Birth / Age','Gender - M/F','Sum Insured(not applicable if floater selected)','Pre Existing illness Y/N','Description of ailment','Date of Inclusion (DD.MM.YYYY)','Date of Joining','Location','Remarks'))
			
			ws1.append(('EmpCode','Date of Deletion','Name of the person','Relation to employee','Date of Birth / Age','Gender - M/F','Status'))
			
			ws2.append(('EmpCode',	'Name of the person','Relation to employee','Designation / Grade','VIPFlag', 'Date of Birth / Age','Gender - M/F','Sum Insured(not applicable if floater selected)','Pre Existing illness Y/N','Description of ailment','Date of Inclusion (DD.MM.YYYY)','Date of Joining','Location','Remarks'))
			
			ws3.append(('EmpCode','Date of Deletion','Name of the person','Relation to employee','Date of Birth / Age','Gender - M/F','Status'))
			

			# -----------------------------Addition TTC--------------------------------------
			
			with closing(connection.cursor()) as cursor:
				query="Select hid.emp_id,hid.ref_no,hid.relation,hid.ins_name,hid.dob,hid.gender,irc.end_date,xhi.exit_status from xlplat_ats_health_ins_prem_price xhi,xlplat_ats_health_ins_details hid,im_employees ie,im_repeating_costs irc,im_costs ic where xhi.sheet_label='"+sheet_str+"' and xhi.emp_id=hid.emp_id and hid.emp_id = ic.cause_object_id and ie.employee_id = ic.cause_object_id and irc.rep_cost_id=ic.cost_id and irc.start_date>='"+str(start_date)+"' and irc.start_date<='"+str(end_date)+"' and ( hid.policy_info='self' or hid.policy_info='self_spouse' ) order by hid.emp_id "
				cursor.execute(query)
				h_add_det_ttc=cursor.fetchall()
			for each_det in h_add_det_ttc:
				emp_id=each_det[0]
				emp_code=fetch_ttc_code(emp_id)
				emp_rel=each_det[2]
				emp_name=each_det[3]
				emp_dob=each_det[4]
				emp_gender=each_det[5]
				applicant_details=employee_details(emp_id)
				job_title_name="NA"
				joining_date="NA"
				sum_assured="NA"
				if emp_rel == "self":
					sum_assured="300000"
					job_title_name=category_data((applicant_details[1]))
					joining_date=datetime.strptime(employee_end_date(emp_id).split(" ")[0],'%Y-%m-%d').date()
				ws.append((emp_code,emp_name,emp_rel,job_title_name,'NA',emp_dob,emp_gender,sum_assured,'N','NA',joining_date,joining_date,'',''))
			# ----------------------------------------------------------------------------------

			# -------------------------------Deletion TTC--------------------------------------

			with closing(connection.cursor()) as cursor:
				query="Select hid.emp_id,hid.ref_no,hid.relation,hid.ins_name,hid.dob,hid.gender,irc.end_date,xhi.exit_status from xlplat_ats_health_ins_prem_price xhi,xlplat_ats_health_ins_details hid,im_employees ie,im_repeating_costs irc,im_costs ic where xhi.sheet_label='"+sheet_str+"' and xhi.emp_id=hid.emp_id and hid.emp_id = ic.cause_object_id and ie.employee_id = ic.cause_object_id and irc.rep_cost_id=ic.cost_id and irc.end_date>'"+str(start_date)+"' and irc.end_date<'"+str(end_date)+"' and ( hid.policy_info='self' or hid.policy_info='self_spouse' ) order by hid.emp_id "
				cursor.execute(query)
				emp_list=cursor.fetchall()
			for each_d in emp_list:
				e_id=each_d[0]
				e_ref=fetch_ttc_code(e_id)
				e_rel=each_d[2]
				e_ins_name=each_d[3]
				e_dob=each_d[4]
				e_gender=each_d[5]
				e_del_date=each_d[6].strftime('%Y-%m-%d')
				e_status=each_d[7]
				e_stat=""
				if e_status == "completed_exit_bonus_sheet":
					e_stat="Completed(Bonus Sheet)"
				elif e_status == "completed_exit_manually":
					e_stat="Completed(Manually)"
				else:
					e_stat="In Process"
				d_row=(e_ref,e_del_date,e_ins_name,e_rel,e_dob,e_gender,e_stat)
				ws1.append(d_row)

			# ------------------------------------------------------------------------------------

			# -----------------------------Addition TTCS--------------------------------------
			
			with closing(connection.cursor()) as cursor:
				query="Select hid.emp_id,hid.ref_no,hid.relation,hid.ins_name,hid.dob,hid.gender,irc.end_date,xhi.exit_status from xlplat_ats_health_ins_prem_price xhi,xlplat_ats_health_ins_details hid,im_employees ie,im_repeating_costs irc,im_costs ic where xhi.sheet_label='"+sheet_str+"' and xhi.emp_id=hid.emp_id and hid.emp_id = ic.cause_object_id and ie.employee_id = ic.cause_object_id and irc.rep_cost_id=ic.cost_id and irc.start_date>='"+str(start_date)+"' and irc.start_date<='"+str(end_date)+"' and hid.policy_info='self_family' order by hid.emp_id "
				cursor.execute(query)
				h_add_det_ttc=cursor.fetchall()
			for each_det in h_add_det_ttc:
				emp_id=each_det[0]
				emp_code=fetch_ttc_code(emp_id)
				emp_rel=each_det[2]
				emp_name=each_det[3]
				emp_dob=each_det[4]
				emp_gender=each_det[5]
				applicant_details=employee_details(emp_id)
				job_title_name="NA"
				joining_date="NA"
				sum_assured="NA"
				if emp_rel == "self":
					sum_assured="300000"
					job_title_name=category_data((applicant_details[1]))
					joining_date=datetime.strptime(employee_end_date(emp_id).split(" ")[0],'%Y-%m-%d').date()
				ws2.append((emp_code,emp_name,emp_rel,job_title_name,'NA',emp_dob,emp_gender,sum_assured,'N','NA',joining_date,joining_date,'',''))
			# ----------------------------------------------------------------------------------

			# -------------------------------Deletion TTCS--------------------------------------

			with closing(connection.cursor()) as cursor:
				query="Select hid.emp_id,hid.ref_no,hid.relation,hid.ins_name,hid.dob,hid.gender,irc.end_date,xhi.exit_status from xlplat_ats_health_ins_prem_price xhi,xlplat_ats_health_ins_details hid,im_employees ie,im_repeating_costs irc,im_costs ic where xhi.sheet_label='"+sheet_str+"' and xhi.emp_id=hid.emp_id and hid.emp_id = ic.cause_object_id and ie.employee_id = ic.cause_object_id and irc.rep_cost_id=ic.cost_id and irc.end_date>'"+str(start_date)+"' and irc.end_date<'"+str(end_date)+"' and hid.policy_info='self_family' order by hid.emp_id "
				cursor.execute(query)
				emp_list=cursor.fetchall()
			for each_d in emp_list:
				e_id=each_d[0]
				e_ref=fetch_ttc_code(e_id)
				e_rel=each_d[2]
				e_ins_name=each_d[3]
				e_dob=each_d[4]
				e_gender=each_d[5]
				e_del_date=each_d[6].strftime('%Y-%m-%d')
				e_status=each_d[7]
				e_stat=""
				if e_status == "completed_exit_bonus_sheet":
					e_stat="Completed(Bonus Sheet)"
				elif e_status == "completed_exit_manually":
					e_stat="Completed(Manually)"
				else:
					e_stat="In Process"
				d_row=(e_ref,e_del_date,e_ins_name,e_rel,e_dob,e_gender,e_stat)
				ws3.append(d_row)

			# ------------------------------------------------------------------------------------
			wb.save('media/health_insurance/appending.xlsx')
			return HttpResponseRedirect('/py/media/health_insurance/appending.xlsx')



		if "resp_cause" in request.GET and "resp_error" in request.GET:
			resp_error=request.GET["resp_error"]
			resp_cause=request.GET["resp_cause"]
			content = {
					"resp_error" : resp_error,
					"resp_cause" : resp_cause,
					"insurance_tbl_html" : ins_tbl_html,
					"year_list":year_list,
					"this_sheet_name":this_sheet_name,
					"table_data":table_data,
					"cur_month":cur_month,
					"c_date":c_date,
					"last_date_next":last_date_next
			}	
			return render(self.request,"health_track_sys.html",content)
		# cur_month=7
		content = {
			"insurance_tbl_html" : ins_tbl_html,
			"year_list":year_list,
			"this_sheet_name":this_sheet_name,
			"table_data":table_data,
			"cur_month":cur_month,
			"c_date":c_date,
			"last_date_next":last_date_next
		}
		return render(self.request,"health_track_sys.html",content)
	def post(self,request):
		# if "summary_send_mail" in request.POST:
		# 	# url="http://localhost//intranet/testing/permission-redirect"
		# 	# payload={"proc_name":"ns_appraisal_sendmail","proc_variables": cand_details_string}
		# 	# response = requests.get(url, params=payload)
		if "delete_profile" in request.POST:
			del_profile=request.POST["delete_profile"]
			this_sheet_name=request.POST["this_sheet_name"]
			# health_ins_prem_price.objects.filter(emp_id=del_profile,sheet_name=this_sheet_name).delete()
		if "update_profile" in request.POST:
			updt_profile=request.POST["update_profile"]
			this_sheet_name=request.POST["this_sheet_name"]
			ins_ref_no=request.POST["ref_"+updt_profile]
			ins_jun=request.POST["jun_"+updt_profile]
			ins_jul=request.POST["jul_"+updt_profile]
			ins_aug=request.POST["aug_"+updt_profile]
			ins_sep=request.POST["sep_"+updt_profile]
			ins_oct=request.POST["oct_"+updt_profile]
			ins_nov=request.POST["nov_"+updt_profile]
			ins_dec=request.POST["dec_"+updt_profile]
			ins_jan=request.POST["jan_"+updt_profile]
			ins_feb=request.POST["feb_"+updt_profile]
			ins_mar=request.POST["mar_"+updt_profile]
			ins_apr=request.POST["apr_"+updt_profile]
			ins_may=request.POST["may_"+updt_profile]
			ins_prem=request.POST["prem_"+updt_profile]
			try:
				c_comp=request.POST["check_comp_"+updt_profile]
			except:
				c_comp=""
			try:
				c_exit=request.POST["check_exit_"+updt_profile]
			except:
				c_exit=""
			health_ins_details.objects.filter(emp_id=updt_profile).update(ref_no=ins_ref_no)
			if c_comp == "completed_bonus_sheet" or c_exit == "completed_exit_bonus_sheet":
				try:
					comp_bonus_id=request.POST["bonus_id_"+updt_profile]
				except:
					comp_bonus_id=""
				bonus_return_val=int(float(ins_jun))+int(float(ins_jul))+int(float(ins_aug))+int(float(ins_sep))+int(float(ins_oct))+int(float(ins_nov))+int(float(ins_dec))+int(float(ins_jan))+int(float(ins_feb))+int(float(ins_mar))+int(float(ins_apr))+int(float(ins_may))-int(float(ins_prem))
				py_category_id=""
				with closing(connection.cursor()) as cursor:
					query="select payee_category_id from im_payroll where user_id="+str(updt_profile)+" limit 1"
					cursor.execute(query)
					py_category_id=cursor.fetchone()
				if py_category_id == None:
					py_category_id="Null"
				else:
					py_category_id=py_category_id[0]
				if comp_bonus_id == "":
					with closing(connection.cursor()) as cursor:
						query="insert into im_payroll_all_bonus_incentive(user_id,date,amount,payee_category_id,type,ex_narration,bank_narration) values("+str(updt_profile)+",'"+str(datetime.now().strftime('%Y-%m-%d'))+"',"+str(bonus_return_val)+","+str(py_category_id)+","+str(10000179)+",'Health Insurance Refund','Health Insurance Refund')"
						cursor.execute(query)
					c_comp=c_comp+"_"+str(cursor.lastrowid)
				else:
					with closing(connection.cursor()) as cursor:
						query="update im_payroll_all_bonus_incentive set user_id="+str(updt_profile)+",date='"+str(datetime.now().strftime('%Y-%m-%d'))+"',amount="+str(bonus_return_val)+",type="+str(10000179)+",ex_narration='Health Insurance Refund',bank_narration='Health Insurance Refund',payee_category_id="+str(py_category_id)+"  where oid="+str(comp_bonus_id)
						cursor.execute(query)
					c_comp=c_comp+"_"+str(comp_bonus_id)
				
			prfl_lck_status="false"
			prfl_month_count=1
			month_index_list=[]
			sheet_year=this_sheet_name.split("_")
			sheet_year1=sheet_year[1]
			sheet_year2=sheet_year[2]
			year_index=sheet_year1
			lck_status=""
			lck_month=0
			cur_month=datetime.now().month
			cur_year=datetime.now().year
			with closing(connection.cursor()) as cursor:
				query="select lock from im_payroll where user_id="+str(updt_profile)+"and year="+str(cur_year)+"and month="+str(cur_month)+""
				cursor.execute(query)
				lck_status=cursor.fetchone()
			pyrll_ins_price=5
			if lck_status == None or lck_status[0] == False:
				lck_month=cur_month
				if lck_month == 1:
					pyrll_ins_price=ins_jan
				elif lck_month == 2:
					pyrll_ins_price=ins_feb
				elif lck_month == 3:
					pyrll_ins_price=ins_mar
				elif lck_month == 4:
					pyrll_ins_price=ins_apr
				elif lck_month == 5:
					pyrll_ins_price=ins_may
				elif lck_month == 6:
					pyrll_ins_price=ins_jun
				elif lck_month == 7:
					pyrll_ins_price=ins_jul
				elif lck_month == 8:
					pyrll_ins_price=ins_aug
				elif lck_month == 9:
					pyrll_ins_price=ins_sep
				elif lck_month == 10:
					pyrll_ins_price=ins_oct
				elif lck_month == 11:
					pyrll_ins_price=ins_nov
				elif lck_month == 12:
					lck_month=0
					pyrll_ins_price=ins_dec
					
				with closing(connection.cursor()) as cursor:
					query="select ins_month from im_payroll_insurance where user_id="+updt_profile
					cursor.execute(query)
					avlbl_chk=cursor.fetchone()
				if avlbl_chk != None:
					with closing(connection.cursor()) as cursor:
						query="update im_payroll_insurance set ins_month="+str(lck_month+1)+",ins_amount=ceil("+str(pyrll_ins_price)+")  where user_id="+str(updt_profile)
						cursor.execute(query)
				else:
					with closing(connection.cursor()) as cursor:
						query="insert into im_payroll_insurance ( user_id,ins_month,ins_amount) values("+str(updt_profile)+","+str(lck_month+1)+",ceil("+str(pyrll_ins_price)+"))"
						cursor.execute(query)
			health_ins_prem_price.objects.filter(emp_id=updt_profile,sheet_label=this_sheet_name).update(jan=ins_jan,feb=ins_feb,mar=ins_mar,apr=ins_apr,may=ins_may,jun=ins_jun,jul=ins_jul,aug=ins_aug,sep=ins_sep,oct=ins_oct,nov=ins_nov,dec=ins_dec,premium=ins_prem)
			if c_exit != "":
				health_ins_prem_price.objects.filter(emp_id=updt_profile,sheet_label=this_sheet_name).update(exit_status=c_exit)
			if c_comp != "":
				health_ins_prem_price.objects.filter(emp_id=updt_profile,sheet_label=this_sheet_name).update(completion_status=c_comp)
			# return HttpResponseRedirect("/py/xlplat_ats/health_track_sys")
		
		elif "h_table_flag" in request.POST:
			emp_id=request.POST["emp_id"]
			emp_name=fetch_employee_name(emp_id)
			with closing(connection.cursor()) as cursor:
				query="select emp_rel_name, emp_relationship,emp_rel_dob from im_employees_family ief where ief.employee_id="+emp_id+"order by emp_rel_name"
				cursor.execute(query)
				all_relative=cursor.fetchall()
			count_index=1
			html_string=""
			show_dob=""
			for each_data in all_relative:
				h_ins_count=0
				h_ins_html=""
				try:
					h_ins_count=health_ins_details.objects.filter(ins_name=each_data[0]).count()
				except:
					h_ins_count=0

				if h_ins_count == 0:
					h_ins_html="<td><label>Yes</label><input type='radio' name='h_ins_inclde_"+str(count_index)+"' value='yes'><label>No</label><input type='radio' name='h_ins_inclde_"+str(count_index)+"' value='no' checked></td>"
				else:
					h_ins_html="<td><label>Yes</label><input type='radio' name='h_ins_inclde_"+str(count_index)+"' value='yes' checked><label>No</label><input type='radio' name='h_ins_inclde_"+str(count_index)+"' value='no'></td>"
				if each_data[2] != None:
					show_dob=each_data[2].strftime('%b %d, %Y')
				html_string+="<tr><td>"+str(count_index)+"</td><td><input type='hidden' name='f_emp_name_"+str(count_index)+"' value='"+str(each_data[0])+"'>"+str(each_data[0])+"</td><td><input type='hidden' name='f_emp_rel_"+str(count_index)+"' value='"+str(each_data[1])+"'>"+str(each_data[1])+"</td><td><input type='hidden' name='f_emp_dob_"+str(count_index)+"' value='"+str(each_data[2])+"'>"+str(show_dob)+" </td>"+h_ins_html+"</tr>"
				count_index+=1
			if html_string == "":
				html_string+="<tr><td colspan='5'><center style='font-size: 18px;font-weight: bold;color: grey;'>No Family Data</center></td></tr>"
			return JsonResponse({"table_string":html_string,"h_emp_name":emp_name,"count_index":count_index})
		elif "update_hlth_det_btn" in request.POST:
			c_index=request.POST["h_c_index"]
			f_id=request.POST["h_emp_det_id"]
			p_info="self"
			f_ref_no=health_ins_details.objects.filter(emp_id=f_id,relation="self").values("ref_no")
			for i in range(1,int(c_index)):
				f_name=request.POST["f_emp_name_"+str(i)]
				f_rel=request.POST["f_emp_rel_"+str(i)].lower()
				f_dob=request.POST["f_emp_dob_"+str(i)]
				f_incl=request.POST["h_ins_inclde_"+str(i)]
				f_gender=None
				if f_rel == "father" or f_rel =="son":
					f_gender="male"
				elif f_rel == "mother" or f_rel == "daughter":
					f_gender="female"
				if (f_rel== "father" or f_rel == "mother") and f_incl == "yes" :
					p_info="self_family"
				elif (f_rel == "son" or f_rel == "daughter" or f_rel == "spouse") and f_incl == "yes" :
					if p_info != "self_family":
						p_info="self_spouse"
				if f_incl == "no":
					try:
						health_ins_details.objects.filter(emp_id=f_id,ins_name=f_name).delete()
					except:
						pass
				elif f_incl == "yes":
					# t=health_ins_details.objects.filter(emp_id=f_id,ins_name=f_name)
					# if not t:
					# 	health_ins_details.objects.filter(emp_id=f_id,ins_name=f_name).update(relation=f_rel,dob=f_dob)
						
					if f_dob == "None":
						f_dob=None
					health_ins_details.objects.update_or_create(emp_id=f_id,ins_name=f_name,defaults={"dob":f_dob,"relation":f_rel,"gender":f_gender,"ref_no":f_ref_no})
						# else:
							# health_ins_details.objects.update_or_create(emp_id=f_id,ins_name=f_name,relation=f_rel)
			# s_count=health_ins_details.objects.filter(emp_id=f_id,relation="self").count()
			f_name=fetch_employee_name(f_id)
			f_rel="self"
			emp_b_date=None
			with closing(connection.cursor()) as cursor:
				query="select e.birthdate from im_employees e where e.employee_id="+f_id+" order by lower(im_name_from_user_id(e.employee_id))"
				cursor.execute(query)
				emp_b_date=cursor.fetchone()
				if emp_b_date != None:
						if emp_b_date[0] != 0:
							emp_b_date=emp_b_date[0]
			health_ins_details.objects.update_or_create(emp_id=f_id,ins_name=f_name,relation=f_rel,dob=emp_b_date)
			health_ins_details.objects.filter(emp_id=f_id).update(policy_info=p_info)
		# elif "add_prem_auto_btn" in request.POST:
		# 	uid_list=request.POST["add_hidden"]
		# 	year=request.POST["add_prem_year_auto"]
		# 	sheet_name="sheet_"+str(year)
		# 	uid_list=uid_list.split("_")
		# 	cur_month = datetime.now().strftime("%-m")
		# 	cur_month=int(cur_month)+1
		# 	for uid in uid_list:
		# 		p_price=request.POST["prem_price_"+uid]
		# 		health_ins_prem_price.objects.update_or_create(emp_id=uid,premium=p_price,sheet_label=sheet_name)
		elif "add_prem_excel_btn" in request.POST:
			temp_filename = request.FILES['add_prem_file']
			tmp_fl_extn=str(temp_filename).split(".")[1].lower()
			add_prem_year=request.POST["add_prem_year"]
			# if os.path.exists('media/health_insurance') != 1:
			# 	os.mkdir('media/health_insurance')
			# with open('media/health_insurance/add_premium_excel.{}'.format(tmp_fl_extn), 'wb+') as destination:
			# 	for chunk in temp_filename.chunks():
			# 		destination.write(chunk)
			wb = load_workbook(temp_filename)
			ws = wb.active
			m_col = ws.max_column
			m_row = ws.max_row
			if m_col != 3:
				resp_error="Header in uploaded file does not match with refrence file"
				resp_cause="refrence_file_error"
				return HttpResponseRedirect("/py/xlplat_ats/health_track_sys?resp_error="+resp_error+"&resp_cause="+resp_cause)
			hc1=ws.cell(row=1,column=1).value.lower()
			hc2=ws.cell(row=1,column=2).value.lower()
			hc3=ws.cell(row=1,column=3).value.lower()
			# wb = xlrd.open_workbook('media/health_insurance/add_premium_excel.{}'.format(tmp_fl_extn))
			# sheet = wb.sheet_by_index(0)
			# if sheet.ncols != 3:
			# 	resp_error="Header in uploaded file does not match with refrence file"
			# 	resp_cause="refrence_file_error"
			# 	return HttpResponseRedirect("/py/xlplat_ats/health_track_sys?resp_error="+resp_error+"&resp_cause="+resp_cause)
			# hc1=sheet.cell_value(0, 0).lower()
			# hc2=sheet.cell_value(0, 1).lower()
			# hc3=sheet.cell_value(0, 2).lower()
			sheet_name="sheet_"+add_prem_year
			resp_error=[]
			if hc1=="emp code" and hc2=="name" and hc3=="premium":
				for i in range(2,m_row+1):
					emp_code=ws.cell(row=i,column=1).value
					emp_name=ws.cell(row=i,column=2).value
					emp_premium=ws.cell(row=i,column=3).value
					# row_val=sheet.row_values(i)
					# emp_code=row_val[0]
					# emp_name=row_val[1]
					# emp_premium=row_val[2]
					emp_id=0
					with closing(connection.cursor()) as cursor:
						cursor.execute("SELECT employee_id FROM im_employees WHERE LOWER(ttc_code) = %s", [emp_code.strip().lower()])
						emp_id = cursor.fetchone()
					if emp_id != None:
						if emp_id[0] != 0:
							health_ins_prem_price.objects.update_or_create(emp_id=int(emp_id[0]),sheet_label=sheet_name,defaults={"premium":emp_premium})
						else:
							resp_error.append(emp_code)
							resp_cause="id_not_found"
					else:
						resp_error.append(emp_code)
						resp_cause="id_not_found"
			else:
				resp_error="Header in uploaded file does not match with refrence file"
				resp_cause="refrence_file_error"
				return HttpResponseRedirect("/py/xlplat_ats/health_track_sys?resp_error="+resp_error+"&resp_cause="+resp_cause)
			if not resp_error:
				return HttpResponseRedirect("/py/xlplat_ats/health_track_sys")
			else:
				resp_error="_".join(resp_error)
				return HttpResponseRedirect("/py/xlplat_ats/health_track_sys?resp_error="+resp_error+"&resp_cause="+resp_cause)
		elif "updt_instl_submit" in request.POST:
			temp_filename = request.FILES['updt_instl_file']
			updt_instl_year=request.POST["updt_instl_year"]
			tmp_fl_extn=str(temp_filename).split(".")[1].lower()
			# if os.path.exists('media/health_insurance') != 1:
			# 	os.mkdir('media/health_insurance')
			# with open('media/health_insurance/update_intallment.{}'.format(tmp_fl_extn), 'wb+') as destination:
			# 	for chunk in temp_filename.chunks():
			# 		destination.write(chunk)
			wb = load_workbook(temp_filename)
			ws = wb.active
			m_col = ws.max_column
			m_row = ws.max_row
			if m_col != 3:
				resp_error="Header in uploaded file does not match with refrence file"
				resp_cause="refrence_file_error"
				return HttpResponseRedirect("/py/xlplat_ats/health_track_sys?resp_error="+resp_error+"&resp_cause="+resp_cause)
			hc1=ws.cell(row=1,column=1).value.lower()
			hc2=ws.cell(row=1,column=2).value.lower()
			hc3=ws.cell(row=1,column=3).value.lower()
			# wb = xlrd.open_workbook('media/health_insurance/update_intallment.{}'.format(tmp_fl_extn))
			# sheet = wb.sheet_by_index(0)
			# if sheet.ncols != 3:
			# 	resp_error="Header in uploaded file does not match with refrence file"
			# 	resp_cause="refrence_file_error"
			# 	return HttpResponseRedirect("/py/xlplat_ats/health_track_sys?resp_error="+resp_error+"&resp_cause="+resp_cause)
			# hc1=sheet.cell_value(0, 0).lower()
			# hc2=sheet.cell_value(0, 1).lower()
			# hc3=sheet.cell_value(0, 2).lower()
			resp_error=[]
			resp_cause=""
			if hc1=="emp code" and hc2=="name" and ( hc3=="jan" or hc3=="feb" or hc3=="mar" or hc3=="apr" or hc3=="may" or hc3=="jun" or hc3=="jul" or hc3=="aug" or hc3=="sep" or hc3=="oct" or hc3=="nov" or hc3=="dec"):
				for i in range(2,m_row+1):
					emp_code=ws.cell(row=i,column=1).value
					emp_name=ws.cell(row=i,column=2).value
					emp_month=ws.cell(row=i,column=3).value
					# row_val=sheet.row_values(i)
					# emp_code=row_val[0]
					# emp_name=row_val[1]
					# emp_month=row_val[2]
					sheet_name="sheet_"+str(updt_instl_year)
					emp_id=0
					with closing(connection.cursor()) as cursor:
						cursor.execute("SELECT employee_id FROM im_employees WHERE LOWER(ttc_code) = %s", [emp_code.strip().lower()])
						emp_id = cursor.fetchone()
					if emp_id != None:
						if emp_id[0] != 0:
							health_ins_prem_price.objects.filter(emp_id=int(emp_id[0]),sheet_label=sheet_name).update(**{hc3: emp_month})
						else:
							resp_error.append(emp_code)
							resp_cause="id_not_found"
					else:
						resp_error.append(emp_code)
						resp_cause="id_not_found"
			else:
				resp_error="Header in uploaded file does not match with refrence file"
				resp_cause="refrence_file_error"
				return HttpResponseRedirect("/py/xlplat_ats/health_track_sys?resp_error="+resp_error+"&resp_cause="+resp_cause)

			if not resp_error:
				return HttpResponseRedirect("/py/xlplat_ats/health_track_sys")
			else:
				resp_error="_".join(resp_error)
				return HttpResponseRedirect("/py/xlplat_ats/health_track_sys?resp_error="+resp_error+"&resp_cause="+resp_cause)
		
		elif "emp_det_btn" in request.POST:
			temp_filename = request.FILES['add_emp_det']
			tmp_fl_extn=str(temp_filename).split(".")[1].lower()
			# if os.path.exists('media/health_insurance') != 1:
			# 	os.mkdir('media/health_insurance')
			# with open('media/health_insurance/employee_family_details.{}'.format(tmp_fl_extn), 'wb+') as destination:
			# 	for chunk in temp_filename.chunks():
			# 		destination.write(chunk)
			# wb = xlrd.open_workbook('media/health_insurance/employee_family_details.{}'.format(tmp_fl_extn))
			# sheet = wb.sheet_by_index(0)
			# For row 0 and column 0
			resp_error2=""
			resp_error=""
			resp_cause=""
			wb = load_workbook(temp_filename)
			ws = wb.active
			m_col = ws.max_column
			m_row = ws.max_row
			if m_col != 5:
				resp_error="Header in uploaded file does not match with refrence file"
				resp_cause="refrence_file_error"
				return HttpResponseRedirect("/py/xlplat_ats/health_track_sys?resp_error="+resp_error+"&resp_cause="+resp_cause)
			hc1=ws.cell(row=1,column=1).value.lower()
			hc2=ws.cell(row=1,column=2).value.lower()
			hc3=ws.cell(row=1,column=3).value.lower()
			hc4=ws.cell(row=1,column=4).value.lower()
			hc5=ws.cell(row=1,column=5).value.lower()
			# if sheet.ncols != 5:
			# 	resp_error="Header in uploaded file does not match with refrence file"
			# 	resp_cause="refrence_file_error"
			# 	return HttpResponseRedirect("/py/xlplat_ats/health_track_sys?resp_error="+resp_error+"&resp_cause="+resp_cause)
			# hc1=sheet.cell_value(0, 0).lower()
			# hc2=sheet.cell_value(0, 1).lower()
			# hc3=sheet.cell_value(0, 2).lower()
			# hc4=sheet.cell_value(0, 3).lower()
			# hc5=sheet.cell_value(0, 4).lower()
			static_ref_no=0
			static_user_id=0
			static_user_id_name=""
			if hc1=="ref_no" and hc2=="name" and hc3=="relation" and hc4=="dob" and hc5=="gender":
				for i in range(2,m_row+1):
					emp_ref=ws.cell(row=i,column=1).value
					emp_name=ws.cell(row=i,column=2).value
					emp_relation=ws.cell(row=i,column=3).value
					emp_dob=ws.cell(row=i,column=4).value
					emp_gender=ws.cell(row=i,column=5).value
					# row_val=sheet.row_values(i)
					# emp_name=row_val[1]
					# emp_relation=row_val[2]
					# emp_dob=row_val[3]
					# emp_gender=row_val[4]
					if emp_gender.lower() == "m":
						emp_gender="male"
					else:
						emp_gender="female"
					# emp_dob=datetime.strptime(str(emp_dob), "%d-%b-%y").date()
					# emp_dob=emp_dob.strftime('%b %d, %Y')
					if static_ref_no == int(emp_ref):
						with closing(connection.cursor()) as cursor:
							cursor.execute("select employee_id from im_employees_family where lower(emp_rel_name)= %s and lower(emp_relationship)= %s and employee_id=%s",[emp_name.lower(),emp_relation.lower(),static_user_id])
							check_f =cursor.fetchone()
						if check_f != None or emp_relation.lower() == "self":
							health_ins_details.objects.update_or_create(emp_id=int(static_user_id),relation=emp_relation.lower(),dob=emp_dob,gender=emp_gender,ins_name=emp_name)
						else:
							resp_error+= emp_name+"-"+emp_relation.lower()+"-"+static_user_id_name+"-"+str(static_user_id)+"_"
							resp_cause="f_profile_not_found_error"
					else:
						# it is assumed that all concerned employee data will consecutive and employee data (relation self) will be at the top of the data of respective employee)
						# Harvinder Singh----------->notify for not found candidates
						
						with closing(connection.cursor()) as cursor:
							cursor.execute("SELECT count(*) FROM persons WHERE LOWER(TRIM(first_names)|| ' ' ||TRIM(last_name)) = %s", [emp_name.strip().lower()])
							check_for_unique_name = cursor.fetchone()
						if check_for_unique_name[0]==1:
							static_ref_no=int(emp_ref)
							with closing(connection.cursor()) as cursor:
								cursor.execute("SELECT person_id FROM persons WHERE LOWER(TRIM(first_names)|| ' ' ||TRIM(last_name)) = %s", [emp_name.strip().lower()])
								unique_name_id = cursor.fetchone()
							static_user_id=unique_name_id[0]
							static_user_id_name=fetch_employee_name(static_user_id)
							with closing(connection.cursor()) as cursor:
								cursor.execute("select employee_id from im_employees_family where lower(emp_rel_name)= %s and lower(emp_relationship)= %s and employee_id=%s",[emp_name.lower(),emp_relation.lower(),unique_name_id[0]])
								check_f =cursor.fetchone()
							if check_f != None or emp_relation.lower() == "self":
								health_ins_details.objects.update_or_create(emp_id=int(unique_name_id[0]),relation=emp_relation.lower(),dob=emp_dob,gender=emp_gender,ins_name=emp_name)
							else:
								resp_error+= emp_name+"_"
								resp_cause="f_profile_not_found_error"
						else:
							resp_error2+= emp_name+"_"
							resp_cause="f_profile_not_found_error"
				sf_id_list=health_ins_details.objects.filter(Q(relation__lower="mother") | Q(relation__lower="father")).values("emp_id")
				check1=health_ins_details.objects.filter(emp_id__in=sf_id_list).update(policy_info="self_family")
				ssp_id_list=health_ins_details.objects.filter(Q(relation__lower="son") | Q(relation__lower="daughter") | Q(relation__lower="spouse")).values("emp_id")
				health_ins_details.objects.filter(Q(emp_id__in=ssp_id_list) & ~Q(emp_id__in=sf_id_list)).update(policy_info="self_spouse")
				if resp_error2 != "":
					resp_error=resp_error+"*"+resp_error2
					resp_cause="f_profile_not_found_error"
				if resp_error != "":
					return HttpResponseRedirect("/py/xlplat_ats/health_track_sys?resp_error="+resp_error+"&resp_cause="+resp_cause)
			else:
				resp_error="Header in uploaded file does not match with refrence file"
				resp_cause="refrence_file_error"
				return HttpResponseRedirect("/py/xlplat_ats/health_track_sys?resp_error="+resp_error+"&resp_cause="+resp_cause)
		return HttpResponseRedirect("/py/xlplat_ats/health_track_sys")

def get_emp_select_option():
	with closing(connection.cursor()) as cursor:
		cursor.execute("select u.user_id,im_name_from_user_id(u.user_id) from registered_users u, group_distinct_member_map gm where u.user_id = gm.member_id and gm.group_id = 463 order by lower(im_name_from_user_id(u.user_id))")
		all_candidates=cursor.fetchall()
	option_html = ""
	for tuple_cand in all_candidates:
		option_html+="<option value='"+str(tuple_cand[0])+"'><p>"+tuple_cand[1]+"</p></option>"
		# print(option_html)
	return option_html

class twillio_email_admin(View):
	def get(self,request):
		user_id_val = XlplatAuthMiddleware.user_id 
		if user_id_val != 0 :
			final_val=1
		else:
			final_val=0
		if final_val == 1:
			user_id_permission=int(check_for_permission_assigned(user_id_val))
			if user_id_permission == 1:
				hr_view=1
			else: 
				hr_view=0
		if hr_view == 1 :
			html_select=get_emp_select_option()
			twl_lst=twillio_email.objects.all().values_list('user_id')
			twl_html=""
			for ed in twl_lst:
				usr_id=ed[0]
				emp_name=fetch_employee_name(usr_id)
				twl_html+="<tr><td>"+emp_name+"</td><td><button name='emp_del' class='btn btn-danger' value='"+usr_id+"' type='submit'>Delete</button></td></tr>"
			content={
			"html_select":html_select,
			"twl_lst":twl_lst,
			"twl_html":twl_html
			}
			return render(self.request,"twillio_email_admin.html",content)
		else:
			return render(self.request,'insufficient_privilages.html',{'redirect_to':'/py/xlplat_ats/twillio_email_admin/'})
	def post(self,request):
		if "auth_twillio" in request.POST:
			sel_val=request.POST["select_twillio"]
			twillio_email.objects.update_or_create(user_id=sel_val)
		if "emp_del" in request.POST:
			usr_id=request.POST["emp_del"]
			twillio_email.objects.filter(user_id=usr_id).delete()
		return HttpResponseRedirect("/py/xlplat_ats/twillio_email_admin")
@csrf_exempt
def twl_start_call(request):
	client_no=request.GET["client_no"]
	repr_mail=request.GET["repr_mail"]
	account_sid = "AC23811c234ff98f9129a44ef7493a7256"
	auth_token = "b62d6692c2b125ae481b3cbb8f4878fa"
	client = Client(account_sid, auth_token)
	url="http://localhost/intranet/testing/permission-redirect"
	payload = {"proc_name": "im_transform_email2user_id", "proc_variables": repr_mail}
	response=requests.get(url, params=payload)
	response=response.text
	response=response.split(" ")[0]
	with closing(connection.cursor()) as cursor:
		cursor.execute("SELECT cell_phone FROM users_contact WHERE user_id = {}".format(response))
		rpr_no = cursor.fetchone()
	rpr_no=rpr_no[0]
	twl_auth_verfy=twillio_email.objects.filter(user_id=response).count()
	if twl_auth_verfy == 0:
		str_call="<Response><Say voice='man' language='en-IN'>Email id is not authorized</Say></Response>"
		return HttpResponse(str_call)
	call = client.calls.create(
	                        method='GET',
	                        url='http://localhost/py/xlplat_ats/twl_call_meta/?usr_id='+response+'&rpr_no='+rpr_no+'',
	                        to=client_no,
	                        from_='+14088728195'
	                    )
	return HttpResponse(status=200)
@csrf_exempt
def twl_call_meta(request):
	user_id=request.GET["usr_id"]
	rpr_no=request.GET["rpr_no"]
	call_sid=request.GET["CallSid"]
	call_to=request.GET["To"]
	call_from=request.GET["From"]
	call_status=request.GET["CallStatus"]
	emp_name=fetch_employee_name(user_id)
	twl_log_data.objects.create(call_sid=call_sid,user_id=user_id,call_to=call_to,call_from=call_from,call_forwarded_to=rpr_no,call_status=call_status)
	# str_call="<Response><Say voice='man' language='en-IN'>Press any key to Connect with "+emp_name+" and  Wait till the end for a voicemail </Say><Dial timeout='30' ringTone='us' method='GET' action='http://84115a772195.ngrok.io/py/xlplat_ats/twl_call_check/?usr_id="+user_id+"&amp;rpr_no="+rpr_no+"&amp;call_sid="+call_sid+"'><Number>+91"+str(rpr_no)+"</Number></Dial></Response>"
	account_sid = "AC23811c234ff98f9129a44ef7493a7256"
	auth_token = "b62d6692c2b125ae481b3cbb8f4878fa"
	client = Client(account_sid, auth_token)
	str_call=client.calls(call_sid) \
             .update(
             	twiml="<Response><Say voice='Polly.Raveena' language='en-IN'> <prosody rate='85%'>You will be connected with "+emp_name+" shortly and  Wait till the end for a voicemail </prosody></Say><Dial timeout='30' ringTone='us' method='GET' action='http://localhost//py/xlplat_ats/twl_call_check/?usr_id="+user_id+"&amp;rpr_no="+rpr_no+"&amp;call_sid="+call_sid+"'><Number>"+str(rpr_no)+"</Number></Dial></Response>")

	return HttpResponse(str_call)

@csrf_exempt
def twl_call_check(request):
	user_id=request.GET["usr_id"]
	rpr_no=request.GET["rpr_no"]
	call_sid=request.GET["call_sid"]
	call_status=request.GET["DialCallStatus"]
	if call_status == "no-answer" or call_status == "busy":
		str_call="<Response><Say voice='Polly.Raveena' language='en-IN'><prosody rate='85%'>Record your message after beep</prosody></Say><Record method='GET' timeout='15' playBeep='true' transcribe='true' action='http://localhost//py/xlplat_ats/twl_call_record/?usr_id="+user_id+"&amp;rpr_no="+rpr_no+"&amp;call_sid="+call_sid+"' /></Response>"
		return HttpResponse(str_call)
	else:
		str_call="<Response><Say voice='Polly.Raveena' language='en-IN'><prosody rate='85%'>Thank You for calling TT Consultants</prosody></Say></Response>"
		return HttpResponse(str_call)

@csrf_exempt
def twl_call_record(request):
	call_sid=request.GET["call_sid"]
	record_url=request.GET["RecordingUrl"]
	record_sid=request.GET["RecordingSid"]
	# twl_log_data.objects.filter(call_sid=call_sid).update(record_sid=record_sid,record_url=record_url)
	account_sid = "AC23811c234ff98f9129a44ef7493a7256"
	auth_token = "b62d6692c2b125ae481b3cbb8f4878fa"
	client = Client(account_sid, auth_token)
	# client.recordings('REf90fe0f321a1d3a190d4ca7eed97bee6').delete()
	# client.recordings('RE9b988e3869e6a387cf0113d3d2c81e9f').fetch()
	# transcriptions = client.transcriptions.list()
	doc = requests.get(record_url+'.mp3')
	temp_filename=str(record_sid) + ".mp3"
	if os.path.exists('media/tw_recording_data') != 1:
		os.mkdir('media/tw_recording_data')
	with open(os.path.join('media/tw_recording_data',str(temp_filename)), 'wb') as f:
	    f.write(doc.content)
	# recordings = client.recordings.list()
	# ----------------------Transcription----------------------------
	response = requests.get('https://api.twilio.com/2010-04-01/Accounts/'+account_sid+'/Recordings/'+record_sid+'/Transcriptions.json ', 
            auth = HTTPBasicAuth(account_sid,auth_token)) 
	resp_list=response.text.split("<")
	t_sid=resp_list[resp_list.index("/Sid>")-1]
	r_list=t_sid.split(">")
	t_sid=r_list[1]
	t_text = requests.get('https://api.twilio.com/2010-04-01/Accounts/'+account_sid+'/Transcriptions/'+t_sid+'.json ', 
            auth = HTTPBasicAuth(account_sid, auth_token))
	t_list=t_text.text.split("<")
	t_text_ind=""
	hit_counter=0
	while t_text_ind=="" and hit_counter<10:
		try:
			t_text_ind=t_list[t_list.index("/TranscriptionText>")-1]
		except:
			t_text = requests.get('https://api.twilio.com/2010-04-01/Accounts/'+account_sid+'/Transcriptions/'+t_sid+'.json ', 
	            auth = HTTPBasicAuth(account_sid, auth_token),timeout=5)
			t_list=t_text.text.split("<")
		hit_counter+=1
	t_list1=t_text_ind.split(">")
	t_text=t_list1[1]
	# t_text=transcription = client.transcriptions(t_sid).fetch()
	if t_text_ind=="":
		t_text=""
	twl_log_data.objects.filter(call_sid=call_sid).update(record_sid=record_sid,record_url=record_url,trans_id=t_sid,trans_text=t_text)
	# print(record_sid,t_sid,t_text,"iiiiiiiiiiiiiiii")
	# ---------------------------------------------------------------
	# ------send mail-----
	url="http://localhost//intranet/testing/permission-redirect"
	payload = {"proc_name": "twilio_send_mail", "proc_variables": record_sid}
	response=requests.get(url, params=payload)
	response=response.text
	# -----------
	# for record in recordings:
	#     print(record.sid)	
	return HttpResponse(status=200)




def candidate_language_data(self,requisition_id):
    with closing(connection.cursor()) as cursor:
        cursor.execute("SELECT trainee FROM xl_ats_req_candidate_map WHERE c_id = %s and req_id= %s", [self,requisition_id])
        row = cursor.fetchone()
    return row

def employee_end_date(candidate_id):
	url="http://localhost//intranet/testing/permission-redirect"
	payload = {"proc_name": "employee_join_terminate_date","extra_parameter": "-user_id", "proc_variables": candidate_id}
	response=requests.get(url, params=payload)
	response=response.text
	return response


def candidate_name(candidate_id):
	with closing(connection.cursor()) as cursor:
		cursor.execute("SELECT app_f_name,app_m_name,app_l_name from xl_ats_candidate_details where c_id= %s", [candidate_id])
		row = cursor.fetchone()
	return row

def candidate_verify_data(candidate_id,filename,requisition_id):
    with closing(connection.cursor()) as cursor:
        cursor.execute("UPDATE xlplat_ats_document set verified='True' WHERE uploaded_by = %s and document_name= %s", [candidate_id,filename])
        row_another = "updated"
    return row_another

def check_for_permission_assigned(user_id_val):
	url="http://localhost//intranet/testing/permission-redirect"
	payload = {"proc_name": "im_user_is_hr_p", "proc_variables": user_id_val}
	response=requests.get(url, params=payload)
	response=response.text
	if response == '0':
		payload = {"proc_name": "im_user_is_admin_p", "proc_variables": user_id_val}
		response=requests.get(url, params=payload)
		response=response.text
	return response

def check_for_permission_assigned_IT(employee_id):
    with closing(connection.cursor()) as cursor:
        cursor.execute("select u.user_id from registered_users u, group_distinct_member_map gm where u.user_id = gm.member_id and gm.group_id = 29057 and u.user_id=%s limit 1",[employee_id])
        employee_perm = cursor.fetchone()
    if employee_perm == None:
        employee_perm = "<unknown>"
    else:
        employee_perm=employee_perm[0]
    return employee_perm 

@register.filter
def get_item(dictionary, key):
    return dictionary.get(key) 

@register.filter(name='split')
def split(value, arg):
	string_var=str(value)
	if string_var != "": 
		try:
			return string_var.split('.')[1] 
		except IndexError:
			return ""
	else: 
		string_var="No file"
		return string_var

class declaration_form(View):
	def get(self,request):
		candidate_id=request.GET['c_id']
		requisition_id=request.GET['req_id']
		employee_entry_exist=Declaration_form.objects.filter(c_id=candidate_id, req_id=requisition_id)
		with closing(connection.cursor()) as cursor:
			cursor.execute("SELECT organization FROM xl_ats_req_candidate_map where c_id= %s and req_id =%s", [candidate_id,requisition_id])
			org_code = cursor.fetchone()
		org_code=org_code[0]
		if not employee_entry_exist:
			show_declaration_submission=0
			all_questions=["Have you in the last 5 years,suffered from any illness or condition which has affected your ability to work?","Do you have any medical condition or concern, that will affect your ability to perform the role applied for, that we should be aware of?","Are you currently taking any medicine?","Have you ever been refused employment, resigned, dismissed or taken early retirement on health grounds?","Have you been treated at a hospital (as an in-patient or an out-patient) in the last five years?","Will this impact on your ability to perform the role applied for?","Do you have (or have you had) any medical condition which may reoccur, e.g. epilepsy,diabetes,etc.?","Are you allergic to any Drug?"," Have you ever had your driving licence withdrawn by the DVLA for health reasons?","Do you have any problems with your hearing or eyesight?","Do you any criminal record or have been held guilty for any act by the government of India?"]
			return render(self.request,'declaration_form.html',{"all_questions": all_questions,"candidate_id":candidate_id,"requisition_id":requisition_id,"show_declaration_submission":show_declaration_submission,"org_code":org_code})
		else:
			show_declaration_submission=1
			return render(self.request,'declaration_form.html',{"show_declaration_submission":show_declaration_submission,"org_code":org_code})

	def post(self,request):
		question_name=[]
		question_id=[]
		c_id=request.POST['c_id']
		req_id=request.POST['req_id']
		question_id=request.POST.getlist('question_id')
		question_name=request.POST.getlist('question_name')
		form_data=request.POST.getlist('comment')
		employee_entry_exist=Declaration_form.objects.filter(c_id=c_id, req_id=req_id)
		if not employee_entry_exist: 
			for each_question in question_id:
				form=DeclarationForm(self.request.POST)
				pdf=form.save(commit=False)
				if form.is_valid():	
					question_id_val=int(each_question)
					question_id_val=question_id_val-1
					comment_data=form_data[question_id_val]
					if comment_data == "":
						question_response=False
					else:
						question_response=True
					pdf.question_id=each_question
					pdf.question_response=question_response
					pdf.question_details=comment_data
					pdf.save()
					show_declaration_submission=1
				else:
					show_declaration_submission=0
		else:
			show_declaration_submission=1
		return render(self.request,'declaration_form.html',{"show_declaration_submission":show_declaration_submission})

class upload_hr_documents(View):

	def get(self,request):
		listy=""
		user_id_val = XlplatAuthMiddleware.user_id 
		if user_id_val != 0 :
			final_val=1
		else:
			final_val=0
		if final_val == 1:
			user_id_permission=int(check_for_permission_assigned(user_id_val))
			if user_id_permission == 1:
				hr_view=1
			else: 
				hr_view=0
		else:
			return render(self.request,'insufficient_privilages.html',{'redirect_to':'/py/xlplat_ats/upload_hr_documents/'})
		org_name="701"
		if "doc_org_name" in request.GET:
			org_name=request.GET["doc_org_name"]
		if org_name=="701" or org_name=="":
			type_manual=1
			type_system=2
			type_pol=3
		elif org_name=="705":
			type_manual=4
			type_system=5
			type_pol=6
		elif org_name=="709":
			type_manual=7
			type_system=8
			type_pol=9
		if hr_view==1:
			hr_heading_dict={"HR Process":"HR_Process","HR System Process":"HR_System_Process","Policies":"Policies"}
			manual_process_list = hr_document_data.objects.filter(type_name=type_manual)
			system_process_list = hr_document_data.objects.filter(type_name=type_system)
			policies_list = hr_document_data.objects.filter(type_name=type_pol)

			hr_manual_process_list = []
			hr_system_process_list = []
			hr_policies_list = []

			for filename in manual_process_list:
				listy=str(filename.file_upload)
				hr_manual_process_list.append(listy.split('/')[-1])
			for filename in system_process_list:
				listy=str(filename.file_upload)
				hr_system_process_list.append(listy.split('/')[-1])
			for filename in policies_list:
				listy=str(filename.file_upload)
				hr_policies_list.append(listy.split('/')[-1])
			p_list={
				"hr_heading_dict":hr_heading_dict,
				"hr_manual_process_list":hr_manual_process_list,
				"hr_system_process_list":hr_system_process_list,
				"hr_policies_list":hr_policies_list,
				"org_name":org_name
				
			}

			return render(self.request,'upload_hr_documents.html',p_list)
		else:
			return render(self.request,'insufficient_privilages.html',{'redirect_to':'/py/xlplat_ats/upload_hr_documents/'})

	def post(self,request):
		user_id_val = XlplatAuthMiddleware.user_id
		if 'upload_flag' in request.POST:
			upload_flag=request.POST['upload_flag']
			upload_type=request.POST['upload_type']
			remove_flag="0"
			remove_upload_file=""
			org_name=request.POST['doc_org_name']
		elif 'remove_flag' in request.POST:
			remove_flag=request.POST['remove_flag']
			remove_upload_file=request.POST['remove_upload_file']
			upload_flag="0"
			upload_type=""
			org_name=""
		
		req_file_upload = request.FILES.getlist('file_upload')
		qwery = []
		listy=""
		if request.method=="POST":
			if remove_flag != "0":
				remove_upload_file = str(remove_upload_file)

				hr_document_data.objects.filter(file_upload=remove_upload_file).delete()
				
				os.remove(os.path.join(settings.BASE_DIR,'media/',remove_upload_file))
			elif upload_flag != "0":
				form=hrDocumentForm(self.request.POST, self.request.FILES)
				
				if form.is_valid():
					for f in req_file_upload: 
						folder_exist=os.path.exists(os.path.join(settings.BASE_DIR,'media/xlplat_ats/{}'.format(upload_type)))
						if folder_exist:
							pass
						else:
							os.mkdir(os.path.join(settings.BASE_DIR,'media/xlplat_ats/', upload_type))
						if upload_type =='HR Process':
							if org_name=="701" or org_name=="":
								upload_type_int=1
							elif org_name=="705":
								upload_type_int=4
							elif org_name=="709":
								upload_type_int=7
							# upload_type_int=1
						elif upload_type =='HR System Process':
							# upload_type_int=2
							if org_name=="701" or org_name=="":
								upload_type_int=2
							elif org_name=="705":
								upload_type_int=5
							elif org_name=="709":
								upload_type_int=8
						elif upload_type == 'Policies':
							# upload_type_int=3
							if org_name=="701" or org_name=="":
								upload_type_int=3
							elif org_name=="705":
								upload_type_int=6
							elif org_name=="709":
								upload_type_int=9
						file_instance=hr_document_data(file_upload=f, type_name=upload_type_int, uploaded_by=user_id_val )
						file_instance.save()
		return HttpResponseRedirect("/py/xlplat_ats/upload_hr_documents?doc_org_name={}".format(org_name))
		
					


class document_checklist(View):
	def get(self,request):
		user_id_val = XlplatAuthMiddleware.user_id 
		if user_id_val != 0 :
			final_val=1
		else:
			final_val=0

		if final_val == 1:
			user_id_permission=int(check_for_permission_assigned(user_id_val))
			# print(user_id_permission)
			if user_id_permission == 1:
				hr_view=1
			else: 
				hr_view=0
			ttc_users={}
			xlpat_users={}
			ta_users={}
			ttcs_users={}
			flag_val=0
			if request.method == 'GET' and 'flag_val' in request.GET:
				flag_val = request.GET['flag_val']
				if str(flag_val)=="1" :
					company_list_items=company_list()
					ttc_code=701
					organisation_users_ttc=candidate_from_cis(ttc_code,1)
					for each_org_ttc_cid_req in organisation_users_ttc:
						add_elements_to_specific_list(each_org_ttc_cid_req,'each_org_ttc','each_org_ttc_req',ttc_users,'c_id')
					employee_ttc_org=employee_list(701,1)
					for each_employee_org_ttc_userid in employee_ttc_org:
						add_elements_to_specific_list(each_employee_org_ttc_userid,'each_org_ttc','each_org_ttc_req',ttc_users,'e_id')
					xlpat_code=703
					organisation_users_xlpat=candidate_from_cis(xlpat_code,1)
					xlscout_code=709
					organisation_users_xlscout=candidate_from_cis(xlscout_code,1)
					for each_org_xlpat_cid_req in organisation_users_xlpat:
						add_elements_to_specific_list(each_org_xlpat_cid_req,'each_org_xlpat','each_org_xlpat_req',xlpat_users,'c_id')
					employee_xlpat_org=employee_list(703,1)
					for each_employee_org_xlpat_userid in employee_xlpat_org:
						add_elements_to_specific_list(each_employee_org_xlpat_userid,'each_org_xlpat','each_org_xlpat_req',xlpat_users,'e_id')
					ta_code=705
					organisation_users_ta=candidate_from_cis(ta_code,1)
					for each_org_ta_cid_req in organisation_users_ta:
						add_elements_to_specific_list(each_org_ta_cid_req,'each_org_ta','each_org_ta_req',ta_users,'c_id')
					employee_ta_org=employee_list(705,1)
					for each_employee_org_ta_userid in employee_ta_org:
						add_elements_to_specific_list(each_employee_org_ta_userid,'each_org_ta','each_org_ta_req',ta_users,'e_id')
					ttcs_code=707
					organisation_users_ttcs=candidate_from_cis(ttcs_code,1)
					for each_org_ttcs_cid_req in organisation_users_ttcs:
						add_elements_to_specific_list(each_org_ttcs_cid_req,'each_org_ttcs','each_org_ttcs_req',ttcs_users,'c_id')
					employee_ttcs_org=employee_list(707,1)
					for each_employee_org_ttcs_userid in employee_ttcs_org:
						add_elements_to_specific_list(each_employee_org_ttcs_userid,'each_org_ttcs','each_org_ttcs_req',ttcs_users,'e_id')

			else:
				company_list_items=company_list()
				ttc_code=701
				organisation_users_ttc=candidate_from_cis(ttc_code)
				for each_org_ttc_cid_req in organisation_users_ttc:
					add_elements_to_specific_list(each_org_ttc_cid_req,'each_org_ttc','each_org_ttc_req',ttc_users,'c_id')
				employee_ttc_org=employee_list(701)
				for each_employee_org_ttc_userid in employee_ttc_org:
					add_elements_to_specific_list(each_employee_org_ttc_userid,'each_org_ttc','each_org_ttc_req',ttc_users,'e_id')
				xlpat_code=703
				organisation_users_xlpat=candidate_from_cis(xlpat_code)
				xlscout_code=709
				organisation_users_xlscout=candidate_from_cis(xlscout_code)
				organisation_users_xlpat=organisation_users_xlpat+organisation_users_xlscout
				for each_org_xlpat_cid_req in organisation_users_xlpat:
					add_elements_to_specific_list(each_org_xlpat_cid_req,'each_org_xlpat','each_org_xlpat_req',xlpat_users,'c_id')
				employee_xlpat_org=employee_list(703)
				for each_employee_org_xlpat_userid in employee_xlpat_org:
					add_elements_to_specific_list(each_employee_org_xlpat_userid,'each_org_xlpat','each_org_xlpat_req',xlpat_users,'e_id')
				ta_code=705
				organisation_users_ta=candidate_from_cis(ta_code)
				for each_org_ta_cid_req in organisation_users_ta:
					add_elements_to_specific_list(each_org_ta_cid_req,'each_org_ta','each_org_ta_req',ta_users,'c_id')
				employee_ta_org=employee_list(705)
				for each_employee_org_ta_userid in employee_ta_org:
					add_elements_to_specific_list(each_employee_org_ta_userid,'each_org_ta','each_org_ta_req',ta_users,'e_id')
				ttcs_code=707
				organisation_users_ttcs=candidate_from_cis(ttcs_code)
				for each_org_ttcs_cid_req in organisation_users_ttcs:
					add_elements_to_specific_list(each_org_ttcs_cid_req,'each_org_ttcs','each_org_ttcs_req',ttcs_users,'c_id')
				employee_ttcs_org=employee_list(707)
				for each_employee_org_ttcs_userid in employee_ttcs_org:
					add_elements_to_specific_list(each_employee_org_ttcs_userid,'each_org_ttcs','each_org_ttcs_req',ttcs_users,'e_id')

			applicant_details=employee_details(user_id_val)
			ttc_code=applicant_details[3]
			# print("hr",hr_view,ttc_code,applicant_details[1])
			if hr_view==1:
				org_list=["TTC","XLSCOUT","TA","TTCS"]
			else:
				if "TTC-" in ttc_code or "TTC " in ttc_code:
					org_list=["TTC"]
				elif "XLP" in ttc_code:
					org_list=["XLSCOUT"]
				elif "TTCS" in ttc_code or "TTCs" in ttc_code:
					org_list=["TTCS"]
				elif "TA" in ttc_code:
					org_list=["TA"]
			
			user_id_val=int(user_id_val)
			final_list={ 
				"org_list":org_list,
				"ttc_users":ttc_users,
				"xlpat_users":xlpat_users,
				"ta_users":ta_users,
				"ttcs_users":ttcs_users,
				"hr_view":hr_view,
				"user_id_val_permission":user_id_val,
				"flag_value":str(flag_val)
			}
			return render(self.request,'document_checklist.html',final_list)
		else:
			return render(self.request,'insufficient_privilages.html',{'redirect_to':'/py/xlplat_ats/document_checklist/'})
	
	
def company_list():
	url="http://localhost//intranet/testing/permission-redirect"
	payload = {"proc_name": "im_category_get_key_value_list", "proc_variables": "Business Unit"}
	response=requests.get(url, params=payload)
	response=response.text
	return response

def candidate_from_cis(org_code,flag=0):
	# flag is to check wheather to show ex-employee list or active employee list
	date_of_today=str(datetime.now())
	if flag==0:
		with closing(connection.cursor()) as cursor:
			cursor.execute("select x.c_id,x.req_id,x.user_id,iem.ttc_code from xl_ats_req_candidate_map x,im_employees iem, registered_users u, group_distinct_member_map gm where x.organization=%s and x.user_id is not null and x.user_id=iem.employee_id and u.user_id = iem.employee_id and u.user_id = gm.member_id and gm.group_id = 463 order by lower(im_name_from_user_id(x.user_id))", [org_code])
			all_candidates=cursor.fetchall()
			return all_candidates 
	elif flag==1:
		with closing(connection.cursor()) as cursor:
			cursor.execute("select x.c_id,x.req_id,x.user_id,iem.ttc_code from xl_ats_req_candidate_map x,im_employees iem, im_repeating_costs irc, im_costs ic where x.organization=%s and x.user_id is not null and x.user_id=iem.employee_id and iem.employee_id=ic.cause_object_id and ic.cost_id=irc.rep_cost_id and irc.end_date< '{}'order by lower(im_name_from_user_id(x.user_id))".format(date_of_today), [org_code])
			all_candidates=cursor.fetchall()
			return all_candidates 

def replace_text(text,type_replace):
	if type_replace == 1:
	    chars = "(),"
	    for c in chars:
	        text = text.replace(c,"")
	    return text
	else:
		chars="[]"
		for c in chars:
			text=text.replace(c,"")
		return text

def final_cand_name(cand_name):
	cand_f_name=str(cand_name[0])
	cand_m_name=str(cand_name[1]) 
	cand_l_name=str(cand_name[2]) 
	if cand_f_name != "None" :
		cand_f_m_l_name=cand_f_name
		if cand_m_name != "None" :
			cand_f_m_l_name=cand_f_m_l_name+" "+cand_m_name
			if cand_l_name != "None" :
				cand_f_m_l_name=cand_f_m_l_name+" "+cand_l_name
		else:
			if cand_l_name != "None" :
				cand_f_m_l_name=cand_f_m_l_name+" "+cand_l_name
	return cand_f_m_l_name

def add_elements_to_specific_list(each_org_cid_req,each_org_cid,each_org_req,users_list,employee_type):
	each_org_cid=each_org_cid_req[0]
	each_org_req=str(each_org_cid_req[1])
	ttc_code_emp=each_org_cid_req[3]
	if ttc_code_emp == None:
		ttc_code_emp=""
	else:
		ttc_code_emp="("+ttc_code_emp+")"
	if employee_type == "c_id":
		each_cid_userid=each_org_cid_req[2]
		cand_name=candidate_name(each_org_cid)
		cand_f_m_l_name=final_cand_name(cand_name)
		ignore_user=0
	else:
		each_cid_userid=each_org_cid_req[0]
		cand_f_m_l_name=each_org_cid_req[2]
		check_cis_exist=cand_cid_exist(each_cid_userid)
		if check_cis_exist != None:
			ignore_user=1 
		else:
			ignore_user=0
	if cand_f_m_l_name == "":
		pass
	else:
		if ignore_user == 0:
			encode_cid=str(each_org_cid)
			encode_cid=base64.b64encode(bytes(encode_cid,"utf-8"))
			encode_cid=encode_cid.decode('utf-8')
			encode_req=base64.b64encode(bytes(each_org_req,"utf-8"))
			encode_req=encode_req.decode('utf-8')
			each_org_cid = int(replace_text(str(each_org_cid),1))
			users_list[each_org_cid]={}
			users_list[each_org_cid]['username']=cand_f_m_l_name
			users_list[each_org_cid]['candidate_id']=each_org_cid
			users_list[each_org_cid]['requisition_id']=each_org_req
			users_list[each_org_cid]['encode_cid']=encode_cid
			users_list[each_org_cid]['encode_req']=encode_req
			users_list[each_org_cid]['user_id_value']=each_cid_userid
			users_list[each_org_cid]['cid_or_userid']=employee_type 
			users_list[each_org_cid]['employee_code']=ttc_code_emp 
			
def cand_cid_exist(user_id):
	with closing(connection.cursor()) as cursor:
		cursor.execute("select c_id from xl_ats_req_candidate_map where user_id=%s",[user_id])
		cand_cid_available=cursor.fetchone()
	return cand_cid_available

def cid_userid_exist(c_id,req_id):
	with closing(connection.cursor()) as cursor:
		cursor.execute("select user_id from xl_ats_req_candidate_map where c_id=%s and req_id=%s",[c_id,req_id])
		cand_cid_available=cursor.fetchone()
	return cand_cid_available

def active_employees():
	final_active_employee=[]
	with closing(connection.cursor()) as cursor:
		# cursor.execute("select u.user_id from registered_users u, group_distinct_member_map gm where u.user_id = gm.member_id and gm.group_id = 463 order by lower(im_name_from_user_id(u.user_id))")
		cursor.execute("select distinct(ie.employee_id) from im_employees ie, im_costs ic, im_repeating_costs irc,group_distinct_member_map gm where ie.employee_id=ic.cause_object_id and ie.employee_id=gm.member_id and irc.rep_cost_id=ic.cost_id and irc.end_date>='2019-01-01'")
		active_employees_all=cursor.fetchall()
		for each_active_employee in active_employees_all:
			each_active = int(replace_text(str(each_active_employee),1))
			final_active_employee.append(each_active)
	return final_active_employee 

def employee_list(org_code,flag=0):
	# flag is to check wheather to show ex-employee list or active employee list
	date_of_today=str(datetime.now())
	if flag==0:
		with closing(connection.cursor()) as cursor:
			if org_code == 701:
				query="select e.employee_id,trim(both 'TTC- ' || 'TTC-' || 'TTC-I-' || 'TTC - ' from e.ttc_code)::integer as ttc_codes,im_name_from_user_id(e.employee_id),e.ttc_code from im_employees e, registered_users u, group_distinct_member_map gm where (e.ttc_code like 'TTC-%' or e.ttc_code like 'TTC -%')  and lower(e.ttc_code) not like '%na%' and u.user_id = e.employee_id and u.user_id = gm.member_id and gm.group_id = 463 order by lower(im_name_from_user_id(e.employee_id))"
				cursor.execute(query)
			elif org_code == 703:
				query="select e.employee_id,trim(both 'XLP- ' || 'XLP-' || 'XLP-I-' || 'XLP - '||'XLS- ' || 'XLS-' || 'XLS - ' from e.ttc_code)::integer as ttc_codes,im_name_from_user_id(e.employee_id),e.ttc_code from im_employees e, registered_users u, group_distinct_member_map gm where (e.ttc_code like 'XLP-%' or e.ttc_code like 'XLP -%' or e.ttc_code like 'XLS-%' or e.ttc_code like 'XLS -%')  and lower(e.ttc_code) not like '%na%' and u.user_id = e.employee_id and u.user_id = gm.member_id and gm.group_id = 463 order by lower(im_name_from_user_id(e.employee_id)) "
				cursor.execute(query)
			elif org_code == 705:
				query="select e.employee_id,trim(both 'TA- ' || 'TA-' || 'TA-I-' || 'TA - ' from e.ttc_code)::integer as ttc_codes,im_name_from_user_id(e.employee_id),e.ttc_code from im_employees e, registered_users u, group_distinct_member_map gm where (e.ttc_code like 'TA-%' or e.ttc_code like 'TA -%') and lower(e.ttc_code) not like '%na%' and u.user_id = e.employee_id and u.user_id = gm.member_id and gm.group_id = 463 order by lower(im_name_from_user_id(e.employee_id))"
				cursor.execute(query)
			elif org_code == 707:
				query="select e.employee_id,trim(both 'TTCS- ' || 'TTCS-' || 'TTCS-I-' || 'TTCS - ' || 'TTCs-' from e.ttc_code)::integer as ttc_codes,im_name_from_user_id(e.employee_id),e.ttc_code from im_employees e, registered_users u, group_distinct_member_map gm where (e.ttc_code like 'TTCS-%' or e.ttc_code like 'TTCS -%') and lower(e.ttc_code) not like '%na%' and u.user_id = e.employee_id and u.user_id = gm.member_id and gm.group_id = 463 order by lower(im_name_from_user_id(e.employee_id))"
				cursor.execute(query)
			all_employees=cursor.fetchall()
		return all_employees 
	elif flag==1:
		with closing(connection.cursor()) as cursor:
			if org_code == 701:
				query="select e.employee_id,trim(both 'TTC- ' || 'TTC-' || 'TTC-I-' || 'TTC - ' from e.ttc_code)::integer as ttc_codes,im_name_from_user_id(e.employee_id),e.ttc_code from im_employees e, im_repeating_costs irc, im_costs ic where (e.ttc_code like 'TTC-%' or e.ttc_code like 'TTC -%')  and lower(e.ttc_code) not like '%na%' and e.employee_id=ic.cause_object_id and ic.cost_id=irc.rep_cost_id and irc.end_date< '{}' order by lower(im_name_from_user_id(e.employee_id))".format(date_of_today)
				cursor.execute(query)
			elif org_code == 703:
				query="select e.employee_id,trim(both 'XLP- ' || 'XLP-' || 'XLP-I-' || 'XLP - ' || 'XLS- ' || 'XLS-' || 'XLS - ' from e.ttc_code)::integer as ttc_codes,im_name_from_user_id(e.employee_id),e.ttc_code from im_employees e, im_repeating_costs irc, im_costs ic where (e.ttc_code like 'XLP-%' or e.ttc_code like 'XLP -%' or e.ttc_code like 'XLS-%' or e.ttc_code like 'XLS -%')  and lower(e.ttc_code) not like '%na%' and e.employee_id=ic.cause_object_id and ic.cost_id=irc.rep_cost_id and irc.end_date< '{}' order by lower(im_name_from_user_id(e.employee_id)) ".format(date_of_today)
				cursor.execute(query)
			elif org_code == 705:
				query="select e.employee_id,trim(both 'TA- ' || 'TA-' || 'TA-I-' || 'TA - ' from e.ttc_code)::integer as ttc_codes,im_name_from_user_id(e.employee_id),e.ttc_code from im_employees e, im_repeating_costs irc, im_costs ic where (e.ttc_code like 'TA-%' or e.ttc_code like 'TA -%') and lower(e.ttc_code) not like '%na%' and e.employee_id=ic.cause_object_id and ic.cost_id=irc.rep_cost_id and irc.end_date< '{}' order by lower(im_name_from_user_id(e.employee_id))".format(date_of_today)
				cursor.execute(query)
			elif org_code == 707:
				query="select e.employee_id,trim(both 'TTCS- ' || 'TTCS-' || 'TTCS-I-' || 'TTCS - ' || 'TTCs-' from e.ttc_code)::integer as ttc_codes,im_name_from_user_id(e.employee_id),e.ttc_code from im_employees e, im_repeating_costs irc, im_costs ic where (e.ttc_code like 'TTCS-%' or e.ttc_code like 'TTCS -%') and lower(e.ttc_code) not like '%na%' and e.employee_id=ic.cause_object_id and ic.cost_id=irc.rep_cost_id and irc.end_date< '{}' order by lower(im_name_from_user_id(e.employee_id))".format(date_of_today)
				cursor.execute(query)
			all_employees=cursor.fetchall()
		return all_employees  

def fetch_employee_name(employee_id):
	with closing(connection.cursor()) as cursor:
		cursor.execute("SELECT im_name_from_user_id(%s)",[employee_id])
		employee_name = cursor.fetchone()
	if employee_name[0] == None:
		employee_name = "<unknown>"
	else:
		employee_name=employee_name[0]
	return employee_name 



def candidate_intern_employee(employee_id):
	url="http://localhost//intranet/testing/permission-redirect"
	payload = {"proc_name": "im_user_is_trainee_p", "proc_variables": employee_id}
	response=requests.get(url, params=payload)
	response=response.text
	return response  

# class greeting_card(View):
# 	def get(self,request):
# 		user_id_val = XlplatAuthMiddleware.user_id
# 		if user_id_val != "":
# 			final_val = '1'
# 		else: 
# 			final_val = 0
# 		if final_val == '1' :
# 			user_id_value=check_for_permission_assigned(user_id_val)
# 			if user_id_value == '1':
# 				hr_view=1
# 			else:
# 				hr_view=0
# 		if hr_view ==1:
# 			greetings_details=greetings_data_details.all().values_list('user_id_for','messagecharacter','user_id_by')
# 			greeting_list={}
# 			for each_msg in greetings_details:
# 				user_id_by=each_msg[0]
# 				messagecharacter=each_msg[1]
# 				user_id_for=each_msg[2]
# 				greeting_list[each_msg]['user_id_for']=user_id_for
# 				greeting_list[each_msg]['messagecharacter']=messagecharacter
# 				greeting_list[each_msg]['user_id_by']=user_id_by
# 			return render(self.request,"greeting_card.html",{"greeting_list":greeting_list})
#     def post(self,request):
#     	user_id_val = XlplatAuthMiddleware.user_id
# 		if user_id_val != "":
# 			final_val = '1'
# 		else: 
# 			final_val = 0
# 		if final_val == '1' :
# 			user_id_value=check_for_permission_assigned(user_id_val)
# 			if user_id_value == '1':
# 				hr_view=1
# 			else:
# 				hr_view=0
# 		if request.method == "POST":
# 			is_vendor_id_generated=int(request.POST['vendor_id_generated'])

class vendor_documents(View):
	def get(self,request):

		user_id_val = XlplatAuthMiddleware.user_id
		if user_id_val != "":
			final_val = '1'
		else: 
			final_val = 0
		if user_id_val == 0 or final_val == '1' :
			user_id_value=check_for_permission_assigned(user_id_val)
			if user_id_value == '1':
				hr_view=1
			else:
				hr_view=0
		if hr_view == 1:
			existing_vendors=Vendor_form.objects.all().values_list('vendor_id','first_name','last_name','date_of_collabration','created_by','picture_uploaded','id').order_by('vendor_id')
			unique_vendor_list={}
			for each_vendor in existing_vendors:
				vendor_id=each_vendor[0]
				vendor_first_name=each_vendor[1].capitalize()
				vendor_last_name=each_vendor[2].capitalize()
				vendor_dob=each_vendor[3].strftime('%b %d, %Y')
				vendor_created_by=each_vendor[4]
				vendor_picture_uploaded=each_vendor[5]
				vendor_auto_generated_id=each_vendor[6]
				vendor_created_by=fetch_employee_name(vendor_created_by)
				unique_vendor_list[vendor_id]={}
				unique_vendor_list[vendor_id]['first_name']=vendor_first_name
				unique_vendor_list[vendor_id]['last_name']=vendor_last_name
				unique_vendor_list[vendor_id]['date_of_collabration']=vendor_dob
				unique_vendor_list[vendor_id]['created_by']=vendor_created_by
				unique_vendor_list[vendor_id]['picture_uploaded']=vendor_picture_uploaded
				existing_vendors_files=Vendor_files.objects.filter(vendor_id=vendor_auto_generated_id).values_list('file_uploaded','uploaded_by','uploaded_on','id').distinct('file_uploaded')
				if not existing_vendors_files:
					unique_vendor_list[vendor_id]['files_exist']=0
				else:
					unique_vendor_list[vendor_id]['files_exist']=1
				unique_vendor_list[vendor_id]['vendor_auto_generated_id']={}
				for vendor_file in existing_vendors_files:
					vendor_filename=vendor_file[0]
					file_uploaded_by=fetch_employee_name(vendor_file[1])
					file_uploaded_on=vendor_file[2].strftime('%b %d, %Y')
					file_id=vendor_file[3]
					unique_vendor_list[vendor_id]['vendor_auto_generated_id'][file_id]={}
					unique_vendor_list[vendor_id]['vendor_auto_generated_id'][file_id]['id']=file_id
					unique_vendor_list[vendor_id]['vendor_auto_generated_id'][file_id]['filename']=vendor_filename
					unique_vendor_list[vendor_id]['vendor_auto_generated_id'][file_id]['uploaded_by']=file_uploaded_by
					unique_vendor_list[vendor_id]['vendor_auto_generated_id'][file_id]['uploaded_on']=file_uploaded_on
			return render(self.request,'vendor_documents.html',{"vendor_list":unique_vendor_list})
		else:
			return render(self.request,'insufficient_privilages.html',{"redirect_to":'/py/xlplat_ats/vendor_documents/'})

	def post(self,request):
		user_id_val = XlplatAuthMiddleware.user_id
		if user_id_val != "":
			final_val = '1'
		else: 
			final_val = 0
		if final_val == '1' :
			user_id_value=check_for_permission_assigned(user_id_val)
			if user_id_value == '1':
				hr_view=1
			else:
				hr_view=0
		if request.method == "POST":
			is_vendor_id_generated=int(request.POST['vendor_id_generated'])
			if is_vendor_id_generated == 0:
				form=VendorForm(self.request.POST, self.request.FILES)
				if form.is_valid():
					data_saved=form.save(commit=False)
					generate_vendor_id=new_vendor_id()
					data_saved.vendor_id=generate_vendor_id
					data_saved.created_by=user_id_val
					file_exist_or_upload(vendor_id=generate_vendor_id,filename='portrait',actual_filename='')
					data_saved.picture_uploaded.storage.location =(os.path.join('media/xlplat_ats/vendor_documents/'+generate_vendor_id+'/'+'portrait'))
					data_saved.save()
					return HttpResponseRedirect("/py/xlplat_ats/vendor_documents/")
			else:
				vendor_id=request.POST['vendor_id']
				delete_file_name=int(request.POST['delete_uploaded_file'])

				if delete_file_name != 1:
					filename=request.FILES['file_uploaded']
					file_id_value=request.POST['file_id_value']
					if file_id_value != "":
						form=VendorUploadDocumentForm(request.FILES)
						if form.is_valid():
							vendor_updated_file=Vendor_files.objects.get(id=file_id_value)
							vendor_object = Vendor_files._meta.get_field('file_uploaded')
							vendor_object_value = str(vendor_object.value_from_object(vendor_updated_file))
							vendor_save=VendorUploadDocumentForm(instance=vendor_updated_file)
							vendor_files=vendor_save.save(commit=False)
							vendor_files.file_uploaded=filename
							vendor_files.uploaded_by=user_id_val
							file_exist_or_upload(vendor_id,'Other_Documents',vendor_object_value)
							vendor_files.file_uploaded.storage.location =(os.path.join('media/xlplat_ats/vendor_documents/'+vendor_id+'/'+'Other_Documents'))
							vendor_files.save()
					else:
						vendor_details=Vendor_form.objects.get(vendor_id=vendor_id)
						form=VendorUploadDocumentForm(request.FILES)
						if form.is_valid():
							vendor_files=form.save(commit=False)
							vendor_files.vendor_id=vendor_details
							vendor_files.file_uploaded=filename
							vendor_files.uploaded_by=user_id_val
							file_exist_or_upload(vendor_id=vendor_id,filename='Other_Documents',actual_filename='')
							vendor_files.file_uploaded.storage.location =(os.path.join('media/xlplat_ats/vendor_documents/'+vendor_id+'/'+'Other_Documents'))
							vendor_files.save()
				else: 
					delete_filename=request.POST['delete_file']
					os.remove("media/xlplat_ats/vendor_documents/"+vendor_id+"/Other_Documents/"+delete_filename)
					file_id_db=request.POST['file_id']
					Vendor_files.objects.get(id=file_id_db,file_uploaded=delete_filename).delete()
				return HttpResponseRedirect("/py/xlplat_ats/vendor_documents/")
				

def new_vendor_id():
	url="http://localhost//intranet/testing/permission-redirect"
	payload = {"proc_name": "get_vendor_id", "proc_variables": "" }
	response=requests.get(url, params=payload)
	response=response.text
	return response 


def file_exist_or_upload(vendor_id,filename,actual_filename):
	vendor_id_exist=os.path.exists('media/xlplat_ats/vendor_documents/'+vendor_id)
	if vendor_id_exist == True:
		vendor_file_exist=os.path.exists('media/xlplat_ats/vendor_documents/'+vendor_id+'/'+filename)
		if vendor_file_exist == True:	
			vendor_actual_file_exist=os.path.exists('media/xlplat_ats/vendor_documents/'+vendor_id+'/'+filename+'/'+actual_filename)
			if vendor_actual_file_exist == True and actual_filename != "" :
				os.remove('media/xlplat_ats/vendor_documents/'+vendor_id+'/'+filename+'/'+actual_filename)	
		else:
			# shutil.rmtree('media/xlplat_ats/vendor_documents/'+vendor_id+'/'+filename)
			os.mkdir(os.path.join('media/xlplat_ats/vendor_documents/'+vendor_id, filename))
	else:
		os.mkdir(os.path.join('media/xlplat_ats/vendor_documents',vendor_id))
		os.mkdir(os.path.join('media/xlplat_ats/vendor_documents/'+vendor_id, filename))


class apply_resign(View):
	def get(self,request):
		user_id_val = XlplatAuthMiddleware.user_id
		if user_id_val != 0:
			final_val = 1
		else: 
			final_val = 0
		if final_val == 1:
			users_list={}
			check_user_is_hr_p=int(check_for_permission_assigned(user_id_val))
			it_team=0
			if check_user_is_hr_p!=1:
				it_team=check_for_permission_assigned_IT(user_id_val)
				if str(it_team)==user_id_val:
					it_team=1
				else:
					it_team=0
			# user_resign_data_available=Resignation_data.objects.filter(resignation_of=user_id_val,deleted=False)
			last_working_date,application_status_db,manager_approved,it_approved,admin_approved,hr_approved,accounts_approved,hr_clearance_approved_final,exit_approved=("",)*9
			try:
				resignation_data_dict=Resignation_data.objects.get(resignation_of=user_id_val,deleted=False)
				last_working_date=resignation_data_dict.relieve_date
				application_status_db=resignation_data_dict.application_status
				manager_approved=resignation_data_dict.manager_clearance_approved
				it_approved=resignation_data_dict.it_clearance_approved
				admin_approved=resignation_data_dict.admin_clearance_approved
				hr_approved=resignation_data_dict.hr_clearance_approved
				accounts_approved=resignation_data_dict.accounts_it_clearance_approved
				hr_clearance_approved_final=resignation_data_dict.hr_clearance_approved_final
				exit_approved=resignation_data_dict.exit_approved
			except Resignation_data.DoesNotExist:
				resignation_data_dict=""
			user_supervisor=""
			# if check_user_is_hr_p != 1:
			user_supervisor=check_supervisor(user_id_val,1)
			if check_user_is_hr_p == 1 or user_supervisor or it_team==1:
				check_user_is_hr_admin_p=1
				if not resignation_data_dict:
					user_already_resigned=0
				else:
					user_already_resigned=1
				user_resign_data_available=list(Resignation_data.objects.filter(deleted=False,hr_clearance_approved_final=False).exclude(resignation_of=user_id_val).values_list('resignation_of', flat=True).order_by('-relieve_date'))
				if check_user_is_hr_p != 1 and it_team !=1 :
					user_resign_data_available=check_supervisor(user_id_val,2,",".join(map(str, user_resign_data_available)))
				for each_user_resign in user_resign_data_available:
					users_list[each_user_resign]={}
					cand_name=fetch_employee_name(each_user_resign)
					users_list[each_user_resign]['Name']=cand_name	
					users_list[each_user_resign]['color']="#fda9a9"	
				user_resign_data_available=list(Resignation_data.objects.filter(deleted=False,hr_clearance_approved_final=True).exclude(resignation_of=user_id_val).values_list('resignation_of', flat=True).order_by('-relieve_date'))
				if check_user_is_hr_p != 1 and it_team !=1 :
					user_resign_data_available=check_supervisor(user_id_val,2,",".join(map(str, user_resign_data_available)))
				for each_user_resign in user_resign_data_available:
					users_list[each_user_resign]={}
					cand_name=fetch_employee_name(each_user_resign)
					users_list[each_user_resign]['Name']=cand_name	
					users_list[each_user_resign]['color']="#a2f1b6"	
					
			else:
				check_user_is_hr_admin_p=0
				if not resignation_data_dict:
					user_already_resigned=0
				else:
					user_already_resigned=1
			send_dictionary={"user_val_id":user_id_val,"user_already_resigned":user_already_resigned,"user_hr_p":check_user_is_hr_admin_p,"users_list":users_list}
			if user_already_resigned == 1:
				relieve_approved_str="active_1"
				manager_approved_str,it_approved_str,admin_approved_str,hr_approved_str,accounts_approved_str,exit_approved_str,hr_clearance_approved_final_str,all_approved_str=("inactive_1",)*8
				if application_status_db == "approved":
					relieve_approved_str="done"
					exit_approved_str="active_1"
					manager_approved_str,all_approved_str=("active_1",)*2
					if exit_approved == True:
						exit_approved_str="done"
						all_approved="active_1"
						if manager_approved == True:
							manager_approved_str="done"
							it_approved_str="active_1"
							admin_approved_str="active_1"
							hr_approved_str="active_1"
							accounts_approved_str="active_1"
							if it_approved == True and admin_approved == True and hr_approved == True and accounts_approved == True:
								all_approved_str,it_approved_str,admin_approved_str,hr_approved_str,accounts_approved_str=("done",)*5
								hr_clearance_approved_final_str="active_1"
								if hr_clearance_approved_final == True:
									hr_clearance_approved_final_str="done"
							else:
								if it_approved == True:
									it_approved_str="done"
								if admin_approved == True :
									admin_approved_str="done"
								if hr_approved == True :
									hr_approved_str="done"
								if accounts_approved == True :
									accounts_approved_str="done"
				send_2={"last_working_date":last_working_date,"application_status_db":application_status_db,"relieve_approved":relieve_approved_str,"manager_approved":manager_approved_str,"it_approved":it_approved_str,"accounts_approved":accounts_approved_str,"admin_approved":admin_approved_str,"hr_approved":hr_approved_str,"exit_approved":exit_approved_str,"hr_clearance_approved_final":hr_clearance_approved_final_str,"all_approved":all_approved_str}
				send_dictionary.update(send_2)
			return render(self.request,'apply_resign_index.html',send_dictionary) 
		else:
			return render(self.request, 'insufficient_privilages.html', {'redirect_to':'/py/xlplat_ats/apply_resign/'})

	def post(self,request):
		user_id_val = XlplatAuthMiddleware.user_id
		if user_id_val != 0 :
			final_val=1
		else:
			final_val=0	
		if final_val == 1:
			if request.method == "POST":
				relieve_date_user=request.POST['relieve_date']
				resign_reason=request.POST['reason_of_leaving'] 
				if relieve_date_user != "" and resign_reason != "":
					date_changed_by_fake=0
					exit_closed_by_fake=0
					cand_app_status='requested'
					new_resignation=Resignation_data.objects.get_or_create(resignation_of=user_id_val,application_status=cand_app_status,relieve_date=relieve_date_user,user_relieve_date=relieve_date_user,reason_of_leaving=resign_reason,date_changed_by=date_changed_by_fake,exit_closed_by=exit_closed_by_fake,deleted=False)

					send_mail("HR_MANAGER_IT_MANAGER1",user_id_val,"request_resignation")
					tomorrow=date.today() + timedelta(days=1)
					send_mail_reminder(user_id_val,"01:00:00",tomorrow,1)
					return HttpResponseRedirect("/py/xlplat_ats/apply_resign/")
		else: 
			return render(self.request, 'insufficient_privilages.html', {'redirect_to':'/py/xlplat_ats/apply_resign/'})
class pf_gratuity(View):
	def get(self,request):
		candidate_id=request.GET['c_id']
		requisition_id=request.GET['req_id']
		with closing(connection.cursor()) as cursor:
			cursor.execute("SELECT organization FROM xl_ats_req_candidate_map where c_id= %s and req_id =%s", [candidate_id,requisition_id])
			org_code = cursor.fetchone()
		org_code=org_code[0]
		employee_entry_exist=pf_gratuity_data.objects.filter(c_id=candidate_id, req_id=requisition_id)
		user_id_val = XlplatAuthMiddleware.user_id
		hr_view=0
		if user_id_val != 0:
			final_val = 1
		else: 
			final_val = 0
		if final_val == 1:
			hr_view=check_for_permission_assigned(user_id_val)
			# print("========>",hr_view,"|",user_id_val)
		if not employee_entry_exist:
			check_submission=0
			send_dict={"candidate_id":candidate_id,"requisition_id":requisition_id,"check_submission":check_submission,"hr_view":hr_view,"org_code":org_code}
			return render(self.request,'pf_gratuity.html',send_dict)
		else:
			p_g_status = pf_gratuity_data.objects.filter(c_id=candidate_id, req_id=requisition_id).values('pf_gr_status')[0]
			# p_g_status=employee_entry_exist.pf_gr_status
			check_gratuity=0
			check_pf=0
			p_stat=p_g_status["pf_gr_status"]
			# print(p_g_status,p_stat)
			if p_stat == 1:
				check_gratuity=1
			elif p_stat == 2:
				check_pf=1
			check_submission=1
			# print("ppppppppppppp",p_g_status,check_pf)
			send_dict={"candidate_id":candidate_id,"requisition_id":requisition_id,"check_submission":check_submission,"hr_view":hr_view,"check_gratuity":check_gratuity,"check_pf":check_pf,"org_code":org_code}
			return render(self.request,'pf_gratuity.html',send_dict)

	def post(self,request):
		# pf_gr_status is all about what is the status of form
		# 0 means niether gratuity form has been filled nor pf form
		# 1 means gratuity form has been filled but not pf form
		# 2 means gratuity and pf has been filled
		user_id_val = XlplatAuthMiddleware.user_id
		hr_view=0
		if user_id_val != 0:
			final_val = 1
		else: 
			final_val = 0
		if final_val == 1:
			hr_view=check_for_permission_assigned(user_id_val)
		if "gratuity_flag" in request.POST:
			gr_val_1 = request.POST["gratuity_val_1"]
			gr_val_2 = request.POST["gratuity_val_2"]
			nomni_count = request.POST["nom_count"]
			st_name=request.POST["statement_name"]
			st_sex=request.POST["statement_sex"]
			st_religion=request.POST["statement_religion"]
			st_maritial=request.POST["statement_maritial"]
			st_village=request.POST["statement_village"]
			st_thana=request.POST["statement_thana"]
			st_sub_div=request.POST["statement_sub_div"]
			st_po_off=request.POST["statement_po_office"]
			st_district=request.POST["statement_district"]
			st_state=request.POST["statement_state"]
			st_place=request.POST["statement_place"]
			st_date=request.POST["statement_date"]
			insert_dict={
				"gratuity_val_1":gr_val_1,
				"gratuity_val_2":gr_val_2,
				"statement_name":st_name,
				"statement_sex":st_sex,
				"statement_religion":st_religion,
				"statement_maritial":st_maritial,
				"statement_village":st_village,
				"statement_thana":st_thana,
				"statement_sub_div":st_sub_div,
				"statement_po_office":st_po_off,
				"statement_district":st_district,
				"statement_state":st_state,
				"statement_place":st_place,
				"statement_date":st_date
			}
			if "statement_dept" in request.POST and "statement_emp_id" in request.POST and "statement_appt_date" in request.POST:
				st_dept=request.POST["statement_dept"]
				st_emp_id=request.POST["statement_emp_id"]
				st_appt_date=request.POST["statement_appt_date"]
				wt_1=request.POST["witness_1"]
				wt_2=request.POST["witness_2"]
				wt_1_name=request.POST["witness_1_name"]
				wt_2_name=request.POST["witness_2_name"]
				wt_place=request.POST["witness_place"]
				wt_date=request.POST["witness_date"]
				insert_dict["statement_dept"]=st_dept
				insert_dict["statement_emp_id"]=st_emp_id
				insert_dict["statement_appt_date"]=st_appt_date
				insert_dict["witness_1"]=wt_1
				insert_dict["witness_2"]=wt_2
				insert_dict["witness_place"]=wt_place
				insert_dict["witness_date"]=wt_date
				insert_dict["witness_1_name"]=wt_1_name
				insert_dict["witness_2_name"]=wt_2_name
			n_count=int(nomni_count)
			for nc in range(n_count):
				indx=nc+1
				indx=str(indx)
				insert_dict["nom_name_"+indx]=request.POST["nom_name_"+indx]
				insert_dict["nom_rel_"+indx]=request.POST["nom_rel_"+indx]
				insert_dict["nom_age_"+indx]=request.POST["nom_age_"+indx]
				insert_dict["nom_percent_"+indx]=request.POST["nom_percent_"+indx]
			c_id=request.POST["c_id"]
			req_id=request.POST["req_id"]
			with closing(connection.cursor()) as cursor:
				cursor.execute("SELECT organization FROM xl_ats_req_candidate_map where c_id= %s and req_id =%s", [c_id,req_id])
				org_code = cursor.fetchone()
			org_code=org_code[0]

			for ans_name,ans_responce in insert_dict.items():
				# print(ans_name,ans_responce)
				pf_gratuity_data.objects.create(c_id=c_id,req_id=req_id,ans_name=ans_name,ans_responce=ans_responce,pf_gr_status=1,form_name="gratuity")
			send_dict={"candidate_id":c_id,"requisition_id":req_id,"check_submission":1,"hr_view":hr_view,"check_gratuity":1,"check_pf":0,"org_code":org_code}
			return render(self.request,'pf_gratuity.html',send_dict)
		elif "decide_flag" in request.POST:
			# print("entered")
			c_id=request.POST["c_id"]
			req_id=request.POST["req_id"]
			with closing(connection.cursor()) as cursor:
				cursor.execute("SELECT organization FROM xl_ats_req_candidate_map where c_id= %s and req_id =%s", [c_id,req_id])
				org_code = cursor.fetchone()
			org_code=org_code[0]

			form_no=request.POST["decide_form"]
			employee_det=pf_gratuity_data.objects.filter(c_id=c_id, req_id=req_id)
			send_dict={"candidate_id":c_id,"requisition_id":req_id,"check_submission":1,"check_gratuity":1,"check_pf":0,"form_no":form_no,"org_code":org_code}
			for ed in employee_det:
				send_dict[ed.ans_name]=ed.ans_responce
			# print(send_dict)
			return render(self.request,'pf_gratuity.html',send_dict)
		elif "pf_form_2_flag" in request.POST:
			pf_father_name = request.POST["pf_father_name"]
			pf_birth_date = request.POST["pf_birth_date"]
			pf_acc_no = request.POST["pf_acc_no"]
			pf_per_addr=request.POST["pf_per_addr"]
			pf_tem_addr=request.POST["pf_tem_addr"]
			pf_epf_count=request.POST["pf_epf_count"]
			pf_eps_count=request.POST["pf_eps_count"]
			pf_eps2_count=request.POST["pf_eps2_count"]
			pf_place=request.POST["pf_place"]
			pf_date=request.POST["pf_date"]
			pf_eps_date=request.POST["pf_eps_date"]
			pf_epf_date=request.POST["pf_epf_date"]

			insert_dict={
				"pf_father_name":pf_father_name,
				"pf_birth_date":pf_birth_date,
				"pf_acc_no":pf_acc_no,
				"pf_per_addr":pf_per_addr,
				"pf_tem_addr":pf_tem_addr,
				"pf_place":pf_place,
				"pf_date":pf_date,
				"pf_eps_date":pf_eps_date,
				"pf_epf_date":pf_epf_date
			}
			n_count=int(pf_epf_count)
			for nc in range(n_count):
				indx=nc+1
				indx=str(indx)
				insert_dict["pf_epf_name_"+indx]=request.POST["pf_epf_name_"+indx]
				insert_dict["pf_epf_rel_"+indx]=request.POST["pf_epf_rel_"+indx]
				insert_dict["pf_epf_birth_date_"+indx]=request.POST["pf_epf_birth_date_"+indx]
				insert_dict["pf_epf_percent_"+indx]=request.POST["pf_epf_percent_"+indx]
				insert_dict["pf_epf_condition_"+indx]=request.POST["pf_epf_condition_"+indx]
			n_count=int(pf_eps_count)
			for nc in range(n_count):
				indx=nc+1
				indx=str(indx)
				insert_dict["pf_eps_name_"+indx]=request.POST["pf_eps_name_"+indx]
				insert_dict["pf_eps_addr_"+indx]=request.POST["pf_eps_addr_"+indx]
				insert_dict["pf_eps_birth_date_"+indx]=request.POST["pf_eps_birth_date_"+indx]
				insert_dict["pf_eps_rel_"+indx]=request.POST["pf_eps_rel_"+indx]
			n_count=int(pf_eps2_count)
			for nc in range(n_count):
				indx=nc+1
				indx=str(indx)
				insert_dict["pf_eps2_name_"+indx]=request.POST["pf_eps2_name_"+indx]
				insert_dict["pf_eps2_birth_date_"+indx]=request.POST["pf_eps2_birth_date_"+indx]
				insert_dict["pf_eps2_rel_"+indx]=request.POST["pf_eps2_rel_"+indx]
			c_id=request.POST["c_id"]
			req_id=request.POST["req_id"]
			with closing(connection.cursor()) as cursor:
				cursor.execute("SELECT organization FROM xl_ats_req_candidate_map where c_id= %s and req_id =%s", [c_id,req_id])
				org_code = cursor.fetchone()
			org_code=org_code[0]

			for ans_name,ans_responce in insert_dict.items():
				# print(ans_name,ans_responce)
				pf_gratuity_data.objects.create(c_id=c_id,req_id=req_id,ans_name=ans_name,ans_responce=ans_responce,pf_gr_status=2,form_name="pf")
			pf_gratuity_data.objects.filter(c_id=c_id,req_id=req_id).update(avail_pf=2,pf_gr_status=2)
			send_dict={"candidate_id":c_id,"requisition_id":req_id,"check_submission":1,"hr_view":hr_view,"check_gratuity":0,"check_pf":1,"org_code":org_code}
			return render(self.request,'pf_gratuity.html',send_dict)
		elif "pf_form_11_flag" in request.POST:
			pf_father_spouse = request.POST["pf_father_spouse"]
			pf_birth_date = request.POST["pf_birth_date"]
			pf_email_id = request.POST["pf_email_id"]
			pf_mobile_no=request.POST["pf_mobile_no"]
			pf_condition1=request.POST["pf_condition1"]
			pf_condition2=request.POST["pf_condition2"]
			pf_in_worker=request.POST["pf_in_worker"]
			pf_bank_acc_no=request.POST["pf_bank_acc_no"]
			pf_ifsc=request.POST["pf_ifsc"]
			pf_place=request.POST["pf_place"]
			pf_date=request.POST["pf_date"]
			is_father_spouse=request.POST["is_father_spouse"]
			insert_dict={
				"pf_father_spouse":pf_father_spouse,
				"pf_birth_date":pf_birth_date,
				"pf_email_id":pf_email_id,
				"pf_mobile_no":pf_mobile_no,
				"pf_condition1":pf_condition1,
				"pf_condition2":pf_condition2,
				"pf_in_worker":pf_in_worker,
				"pf_bank_acc_no":pf_bank_acc_no,
				"pf_ifsc":pf_ifsc,
				"pf_place":pf_place,
				"pf_date":pf_date,
				"is_father_spouse":is_father_spouse
			}
			if "pf_aadhar" in request.POST:
				pf_aadhar=request.POST["pf_aadhar"]
				insert_dict["pf_aadhar"]=pf_aadhar
			if "pf_pan" in request.POST:
				pf_pan=request.POST["pf_pan"]
				insert_dict["pf_pan"]=pf_pan


			if "pf_uan_no" in request.POST:
				pf_uan_no=request.POST["pf_uan_no"]
				insert_dict["pf_uan_no"]=pf_uan_no
			
			# ########## PF previous account NO ############
			if "pf_pre_acc_no_AP" in request.POST:
				pf_pre_acc_no_AP=request.POST["pf_pre_acc_no_AP"]
				insert_dict["pf_pre_acc_no_AP"]=pf_pre_acc_no_AP

			if "pf_pre_acc_no_HYD" in request.POST:
				pf_pre_acc_no_HYD=request.POST["pf_pre_acc_no_HYD"]
				insert_dict["pf_pre_acc_no_HYD"]=pf_pre_acc_no_HYD
			
			if "pf_pre_acc_no_EST_CODE" in request.POST:
				pf_pre_acc_no_EST_CODE=request.POST["pf_pre_acc_no_EST_CODE"]
				insert_dict["pf_pre_acc_no_EST_CODE"]=pf_pre_acc_no_EST_CODE
			
			if "pf_pre_acc_no_EXTN" in request.POST:
				pf_pre_acc_no_EXTN=request.POST["pf_pre_acc_no_EXTN"]
				insert_dict["pf_pre_acc_no_EXTN"]=pf_pre_acc_no_EXTN
			
			if "pf_pre_acc_no_PF_NO" in request.POST:
				pf_pre_acc_no_PF_NO=request.POST["pf_pre_acc_no_PF_NO"]
				insert_dict["pf_pre_acc_no_PF_NO"]=pf_pre_acc_no_PF_NO
			# ###################################################

			if "pf_date_exit" in request.POST:
				pf_date_exit=request.POST["pf_date_exit"]
				insert_dict["pf_date_exit"]=pf_date_exit
			if "pf_scheme_cert_no" in request.POST:
				pf_scheme_cert_no=request.POST["pf_scheme_cert_no"]
				insert_dict["pf_scheme_cert_no"]=pf_scheme_cert_no
			if "pf_ppo_no" in request.POST:
				pf_ppo_no=request.POST["pf_ppo_no"]
				insert_dict["pf_ppo_no"]=pf_ppo_no


			if "pf_country_origin" in request.POST:
				pf_country_origin=request.POST["pf_country_origin"]
				insert_dict["pf_country_origin"]=pf_country_origin
			if "pf_passport_no" in request.POST:
				pf_passport_no=request.POST["pf_passport_no"]
				insert_dict["pf_passport_no"]=pf_passport_no
			if "pf_passport_valid_from" in request.POST:
				pf_passport_valid_from=request.POST["pf_passport_valid_from"]
				insert_dict["pf_passport_valid_from"]=pf_passport_valid_from
			if "pf_passport_valid_to" in request.POST:
				pf_passport_valid_to=request.POST["pf_passport_valid_to"]
				insert_dict["pf_passport_valid_to"]=pf_passport_valid_to

			c_id=request.POST["c_id"]
			req_id=request.POST["req_id"]
			with closing(connection.cursor()) as cursor:
				cursor.execute("SELECT organization FROM xl_ats_req_candidate_map where c_id= %s and req_id =%s", [c_id,req_id])
				org_code = cursor.fetchone()
			org_code=org_code[0]
			for ans_name,ans_responce in insert_dict.items():
				# print(ans_name,ans_responce)
				pf_gratuity_data.objects.create(c_id=c_id,req_id=req_id,ans_name=ans_name,ans_responce=ans_responce,pf_gr_status=2,form_name="pf")
			pf_gratuity_data.objects.filter(c_id=c_id,req_id=req_id).update(avail_pf=11,pf_gr_status=2)
			
			
			send_dict={"candidate_id":c_id,"requisition_id":req_id,"check_submission":1,"hr_view":hr_view,"check_gratuity":0,"check_pf":1,"org_code":org_code}
			return render(self.request,'pf_gratuity.html',send_dict)





def modify_user_application(request,ap_id):
	user_id_val = XlplatAuthMiddleware.user_id
	if user_id_val != 0:
		final_val = 1
	else: 
		final_val = 0
	if final_val == 1 and user_id_val != ap_id:
		hr_permission=str(check_for_permission_assigned(user_id_val))
		manager_id=get_supervisor(ap_id)
		imm_manager_id=get_imm_supervisor(ap_id)
		vp_id=get_vp_id_db(ap_id)
		it_team=0
		if hr_permission!="1":
			it_team=check_for_permission_assigned_IT(user_id_val)
			if str(it_team)==user_id_val:
				it_team=1
			else:
				it_team=0
		if hr_permission == "1" or str(manager_id) == str(user_id_val) or str(vp_id) == str(user_id_val) or str(imm_manager_id) == str(user_id_val):
			it_team=0
		if hr_permission == "1" or str(manager_id) == str(user_id_val) or str(vp_id) == str(user_id_val) or str(imm_manager_id) == str(user_id_val) or it_team == 1:
			if request.method == "POST" and (hr_permission == "1" or str(manager_id) == str(user_id_val) or str(vp_id) == str(user_id_val) or str(imm_manager_id) == str(user_id_val)):
				revert_address="/py/xlplat_ats/modify_user_application/{}".format(ap_id)
				resignation_approved=""
				if 'approve_resignation' in request.POST:
					resignation_approved=request.POST['approve_resignation']
				if resignation_approved == "approved":
					today_date = date.today().strftime("%Y-%m-%d")
					send_mail("EMPLOYEE_HR_MANAGER",ap_id,"date_approved")
					try:
						resignation_data_dict=Resignation_data.objects.get(resignation_of=ap_id,deleted=False)	
					except Resignation_data.DoesNotExist:
						resignation_data_dict=""
					if resignation_data_dict:
						relieve_date=resignation_data_dict.relieve_date
						relieve_date=datetime.strptime(str(relieve_date),'%Y-%m-%d')
						mail_date=relieve_date - timedelta(days=3)
						send_mail_reminder(ap_id,"02:00:00",mail_date,1)
					obj, created = Resignation_data.objects.update_or_create(resignation_of=ap_id,deleted=False,defaults={'application_status':resignation_approved,'resignation_approved_on':today_date})
				elif 'new_relieve_date' in request.POST:
					changed_relieve_date=request.POST['new_relieve_date']
					if changed_relieve_date != "":
						updated_by=user_id_val
						ap_date_history=str(Resignation_data.objects.get(resignation_of=ap_id,deleted=False).user_relieve_date)
						ap_date_history="{},{}".format(ap_date_history,changed_relieve_date)
						obj, created = Resignation_data.objects.update_or_create(resignation_of=ap_id,deleted=False,defaults={'relieve_date':changed_relieve_date,"date_changed_by":updated_by,'user_relieve_date':ap_date_history})
				elif 'new_resign_date' in request.POST:
					changed_resign_date=request.POST['new_resign_date']
					if changed_resign_date != "":
						obj, created = Resignation_data.objects.update_or_create(resignation_of=ap_id,deleted=False,defaults={'user_resign_date':changed_resign_date})
				elif 'delete_confirm' in request.POST:
					Resignation_data.objects.filter(resignation_of=ap_id,deleted=False).update(deleted=True)
					send_mail_reminder(ap_id,"","",0)
					revert_address="/py/xlplat_ats/apply_resign/"
				return HttpResponseRedirect(revert_address)

			ap_application_status,manager_approved,it_approved,admin_approved,hr_approved,accounts_approved,hr_clearance_approved_final,exit_approved=("",)*8
			manager_name=""
			if manager_id != None and manager_id:
				manager_name=fetch_employee_name(manager_id)
			try:
				resignation_data_dict=Resignation_data.objects.get(resignation_of=ap_id,deleted=False)
			except Resignation_data.DoesNotExist:
				resignation_data_dict=""
			ap_name=fetch_employee_name(ap_id)
			image_src="/py/media/images/profile.jpg"
			send_dict={}
			if not resignation_data_dict:
				ap_relieve_date=""
				ap_resign_reason=""
				approval_val=""
			else:
				server_base_path=get_server_folder_path()
				image_exists=glob.glob(server_base_path+"/filestorage/users/"+ap_id+"/portrait*")
				if len(image_exists) != 0:
					image_name=image_exists[0].split("/")[-1]
					image_src="/intranet/download/user/"+ap_id+"/"+image_name
				else:
					try:
						cand_id=get_candidate_id(ap_id)
						if cand_id != None:
							image_exists=glob.glob(server_base_path+"/filestorage/atsfiles/"+cand_id+"/portrait")
							if len(image_exists) != 0:
								image_name=image_exists[0].split("/")[-1]
								image_src="/intranet/download/atsfiles/"+cand_id+"/"+image_name
					except:
						image_src="/py/media/images/profile.jpg"
				ap_relieve_date=resignation_data_dict.relieve_date
				resign_date=resignation_data_dict.user_resign_date
				ap_resign_reason=resignation_data_dict.reason_of_leaving
				ap_application_status=resignation_data_dict.application_status
				manager_approved=resignation_data_dict.manager_clearance_approved
				it_approved=resignation_data_dict.it_clearance_approved
				admin_approved=resignation_data_dict.admin_clearance_approved
				hr_approved=resignation_data_dict.hr_clearance_approved
				accounts_approved=resignation_data_dict.accounts_it_clearance_approved
				hr_clearance_approved_final=resignation_data_dict.hr_clearance_approved_final
				exit_approved=resignation_data_dict.exit_approved
				if ap_application_status == "approved":
					approval_val=1
				else:
					approval_val=0
				relieve_approved_str="active_1"
				manager_approved_str,it_approved_str,admin_approved_str,hr_approved_str,accounts_approved_str,exit_approved_str,hr_clearance_approved_final_str,all_approved_str=("inactive_1",)*8
				if ap_application_status == "approved":
					relieve_approved_str="done"
					exit_approved_str="active_1"
					manager_approved_str,all_approved_str=("active_1",)*2
					if exit_approved == True:
						exit_approved_str="done"
						all_approved="active_1"
						if manager_approved == True:
							manager_approved_str="done"
							it_approved_str,admin_approved_str,hr_approved_str,accounts_approved_str=("active_1",)*4
							if it_approved == True and admin_approved == True and hr_approved == True and accounts_approved == True:
								all_approved_str,it_approved_str,admin_approved_str,hr_approved_str,accounts_approved_str=("done",)*5
								hr_clearance_approved_final_str="active_1"
								if hr_clearance_approved_final == True:
									hr_clearance_approved_final_str="done"
							else:
								if it_approved == True:
									it_approved_str="done"
								if admin_approved == True :
									admin_approved_str="done"
								if hr_approved == True :
									hr_approved_str="done"
								if accounts_approved == True :
									accounts_approved_str="done"
				send_dict={"resign_date":resign_date,"relieve_approved":relieve_approved_str,"manager_approved":manager_approved_str,"it_approved":it_approved_str,"accounts_approved":accounts_approved_str,"admin_approved":admin_approved_str,"hr_approved":hr_approved_str,"exit_approved":exit_approved_str,"hr_clearance_approved_final":hr_clearance_approved_final_str,"all_approved":all_approved_str}
			send_dict.update({"manager_name":manager_name,"applicant_id":ap_id,"applicant_name":ap_name,"applicant_relieve_date":ap_relieve_date,"applicant_resign_reason":ap_resign_reason,"approval_val":approval_val,"image_src":image_src,"hr_permission":str(hr_permission),"it_team":it_team})
			return render(request,'modify_user_application.html',send_dict)
		else:
			return render(request, 'insufficient_privilages.html', {'redirect_to':'/py/xlplat_ats/modify_user_application/{}'.format(ap_id)})
	else:
		return render(request, 'insufficient_privilages.html', {'redirect_to':'/py/xlplat_ats/modify_user_application/{}'.format(ap_id)}) 


def get_vp_id_db(user_id):
	query="select get_hod_from_user_id({})".format(user_id)
	with closing(connection.cursor()) as cursor:
		cursor.execute(query)
		row=cursor.fetchone()
		if not row or not len(row):
			return row
	return row[0]

def greeting_card(request,ap_id):
	user_id_val = XlplatAuthMiddleware.user_id
	if user_id_val != 0:
		final_val = 1
	else: 
		final_val = 0
	if final_val == 1:
		hr_view=check_for_permission_assigned(user_id_val)
		if str(hr_view) == "1" and str(user_id_val) == str(ap_id):
			hr_view = "0"
		employee_name=fetch_employee_name(ap_id)
		applicant_details=employee_details(ap_id)
		department_name=category_data(applicant_details[0])
		job_title_name=category_data((applicant_details[1]))
		joining_date=datetime.strptime(employee_end_date(ap_id).split(" ")[0],'%Y-%m-%d').date()
		manager=fetch_employee_name(applicant_details[2])
		ttc_code=applicant_details[3]
		VP_id=get_vp_id_db(ap_id)
		greetings_msg=get_greeting_card_msg(ap_id)
		greetings_by=get_greeting_card_by(ap_id)

		if str(user_id_val) == str(VP_id):
			hr_view="1"
		# allowed only for vp
		if hr_view == "1" or str(user_id_val) == str(ap_id):
			image_header="/py/media/images/TT-Consultants-patent-Logo.png"
			header_name="header_ttc"
			company_text="TTC"
			if "TTC-" in ttc_code or "TTC " in ttc_code:
				image_header="/py/media/images/TT-Consultants-patent-Logo.png"
				extra_image="width: 22vh"
				header_name="header_ttc"
				company_text="TTC"
			elif "XLP" in ttc_code:
				image_header="/py/media/images/XLPAT.png"
				extra_image="width: 16vh"
				header_name="header_xl"
				company_text="XLSCOUT"
			elif "TTCS" in ttc_code or "TTCs" in ttc_code:
				# image_header="/py/media/images/ttcs_img.png"
				image_header="/py/media/images/TT-Consultants-patent-Logo.png"
				extra_image="width: 22vh"
				header_name="header_ttcs"
				company_text="TTC"
			elif "TA" in ttc_code:
				image_header="/py/media/images/ta_final_logo.jpg"
				extra_image="width: 15vh"
				header_name="header_ta"
				company_text="TA"
			elif "XLS" in ttc_code:
				image_header="/py/media/images/xlscout.png"
				extra_image="width: 22vh"
				header_name="header_xls"
				company_text="XLSCOUT"
			greetings_mail=""
			if 'send_greetings_mail' in request.POST:
				greetings_mail=request.POST['send_greetings_mail']
			if greetings_mail == "send_greetings":
				send_data="{}~{}~{}".format(user_id_for,permission,user_id_by)
				url="http://localhost//intranet/testing/permission-redirect"
				payload = {"proc_name": "send_greetings_mail", "proc_variables": send_data}
				response=requests.get(url, params=payload)
			
			try:
				resignation_data_dict=Resignation_data.objects.get(resignation_of=ap_id,deleted=False)
				last_working_date=resignation_data_dict.relieve_date
				application_status_db=resignation_data_dict.application_status
				message="Resignation has not been Approved!"
				only_relieve_date=datetime.strptime(str(last_working_date),'%Y-%m-%d')
			except Resignation_data.DoesNotExist:
				application_status_db=""
				last_working_date=""
				message="Not applied for Resignation!"
				only_relieve_date=datetime.strptime(str("2099-12-31"),'%Y-%m-%d')
			three_days_relieve = only_relieve_date - timedelta(days=3)
			today = datetime.now()
			data_dict_greeting={"applicant_id":ap_id,"employee_name":employee_name,'department_name':department_name,'job_title_name':job_title_name,'manager':manager,'joining_date':joining_date,'last_date':last_working_date,"greetings_msg":greetings_msg,"greetings_by":greetings_by,"hr_permission":hr_view,"image_header":image_header,"extra_image":extra_image,"company_text":company_text}
			data_dict_extend={}

			
			return render(request,'greeting_card.html',data_dict_greeting)
			
		
		
def exit_employee_form(request,ap_id):
	user_id_val = XlplatAuthMiddleware.user_id
	if user_id_val != 0:
		final_val = 1
	else: 
		final_val = 0
	if final_val == 1:
		hr_view=check_for_permission_assigned(user_id_val)
		if str(hr_view) == "1" and str(user_id_val) == str(ap_id):
			hr_view = "0"
		employee_name=fetch_employee_name(ap_id)
		applicant_details=employee_details(ap_id)
		department_name=category_data(applicant_details[0])
		job_title_name=category_data((applicant_details[1]))
		joining_date=datetime.strptime(employee_end_date(ap_id).split(" ")[0],'%Y-%m-%d').date()
		manager=fetch_employee_name(applicant_details[2])
		ttc_code=applicant_details[3]
		VP_id=get_vp_id_db(ap_id)
		if str(user_id_val) == str(VP_id):
			hr_view="1"
		# allowed only for vp
		if hr_view == "1" or str(user_id_val) == str(ap_id):
			image_header="/py/media/images/TT-Consultants-patent-Logo.png"
			header_name="header_ttc"
			company_text="TTC"
			if "TTC-" in ttc_code or "TTC " in ttc_code:
				image_header="/py/media/images/TT-Consultants-patent-Logo.png"
				extra_image="width: 22vh"
				header_name="header_ttc"
				company_text="TTC"
			elif "XLP" in ttc_code:
				image_header="/py/media/images/XLPAT.png"
				extra_image="width: 16vh"
				header_name="header_xl"
				company_text="XLSCOUT"
			elif "TTCS" in ttc_code or "TTCs" in ttc_code:
				# image_header="/py/media/images/ttcs_img.png"
				image_header="/py/media/images/TT-Consultants-patent-Logo.png"
				extra_image="width: 22vh"
				header_name="header_ttcs"
				company_text="TTC"
			elif "TA" in ttc_code:
				image_header="/py/media/images/ta_final_logo.jpg"
				extra_image="width: 15vh"
				header_name="header_ta"
				company_text="TA"
			elif "XLS" in ttc_code:
				image_header="/py/media/images/xlscout.png"
				extra_image="width: 22vh"
				header_name="header_xls"
				company_text="XLSCOUT"
			text_questions={1:"Why are you leaving "+company_text+"?",2:"What circumstances would have prevented your departure?",3:"What did you like most about your job?",4:"What did you like least about your job?"}
			manager_questions={1:"Was consistently fair",2:"Provided recognition",3:"Resolved complaints",4:"Was sensitive to employees' needs",5:"Provided feedback on performance",6:"Was receptive to open communication"}
			manager_options={1:"Almost Always",2:"Usually",3:"Sometimes",4:"Never"}
			company_questions={1:"Cooperation within your division/program",2:"Cooperation with other department",3:"Equipment provided (materials, resources, facilities)",4:"Company's performance review system",5:"Company's new employee orientation program",6:"Rate of pay for your job",7:"Career development/Advancement opportunities",8:"Physical working conditions"}
			company_options={1:"Excellent",2:"Good",3:"Fair",4:"Poor"}
			holidays_questions={1:"Paid holidays",2:"Medical plan",3:"Sick leave"}
			holidays_options={1:"Excellent",2:"Good",3:"Fair",4:"Poor",5:"No Opinion"}
			form_filled=0
			try:
				resignation_data_dict=Resignation_data.objects.get(resignation_of=ap_id,deleted=False)
				last_working_date=resignation_data_dict.relieve_date
				application_status_db=resignation_data_dict.application_status
				message="Resignation has not been Approved!"
				only_relieve_date=datetime.strptime(str(last_working_date),'%Y-%m-%d')
			except Resignation_data.DoesNotExist:
				application_status_db=""
				last_working_date=""
				message="Not applied for Resignation!"
				only_relieve_date=datetime.strptime(str("2099-12-31"),'%Y-%m-%d')
			three_days_relieve = only_relieve_date - timedelta(days=3)
			today = datetime.now()
			data_dict_exit={"applicant_id":ap_id,"employee_name":employee_name,'department_name':department_name,'job_title_name':job_title_name,'manager':manager,'joining_date':joining_date,'last_date':last_working_date,"hr_permission":hr_view,"image_header":image_header,"extra_image":extra_image,"company_text":company_text}
			data_dict_extend={}
			print_check=0
			if request.method == "POST":
				if "save_exit" in request.POST:
					print_check=1
			if request.method == "GET" or print_check==1:
				
				if application_status_db == "approved":
					

					exit_form_filled=resignation_data_dict.exit_form_json

					if exit_form_filled != "" :
						if hr_view == "1":
							exit_approved=resignation_data_dict.exit_approved
							show_interview_form_submission=0
							exit_form_filled=json.loads(exit_form_filled)
							text_questions_dict=exit_form_filled['text_questions']
							manager_questions_dict=exit_form_filled['manager_questions']
							company_questions_dict=exit_form_filled['company_questions']
							holidays_questions_dict_val=exit_form_filled['holidays_questions']
							yes_or_no_val=exit_form_filled['yes_or_no']
							yes_or_no_commentval=exit_form_filled['yes_or_no_comments']
							type_workload_val=exit_form_filled['type_workload']
							org_recommend_val=exit_form_filled['org_recommend']
							ttc_better_work_val=exit_form_filled['ttc_better_work']
							hr_remark=""
							if "hr_remark" in exit_form_filled:
								hr_remark=exit_form_filled["hr_remark"]
							data_dict_extend={"show_interview_form_submission":show_interview_form_submission,"text_questions_dict":text_questions_dict,"manager_questions_dict":manager_questions_dict,"company_questions_dict":company_questions_dict,"holidays_questions_dict_val":holidays_questions_dict_val,"yes_or_no_val":yes_or_no_val,"yes_or_no_commentval":yes_or_no_commentval,"type_workload_val":type_workload_val,"org_recommend_val":org_recommend_val,"ttc_better_work_val":ttc_better_work_val,"hr_remark":hr_remark,"exit_approved":exit_approved,"message":"","form_filled":form_filled}
						elif three_days_relieve <= today:
							data_dict_extend={"show_interview_form_submission":1,"message":"","form_filled":form_filled}
						else:
							data_dict_extend={"text_based_question":"","show_interview_form_submission":0,"message":"Not Allowed!!","form_filled":form_filled}
						
					elif hr_view == "0":
						if three_days_relieve <= today:
							data_dict_extend={"text_based_question":text_questions,"manager_questions":manager_questions,"company_questions":company_questions,"holidays_questions":holidays_questions,"show_interview_form_submission":0,"manager_options":manager_options,"company_options":company_options,"holidays_options":holidays_options,"form_filled":form_filled}
						else:
							data_dict_extend={"text_based_question":"","show_interview_form_submission":0,"form_filled":form_filled,"message":"Not Allowed!!"}
					else:
						data_dict_extend={"text_based_question":"","show_interview_form_submission":0,"form_filled":form_filled,"message":"User has not filled the Exit Interview Form"}
				else:
					data_dict_extend={"text_based_question":"","show_interview_form_submission":0,"form_filled":form_filled,"message":message}
				data_dict_exit.update(data_dict_extend)
				if print_check==1:
					# media/xlplat_ats
					candidate_id=get_candidate_id(ap_id)
					if candidate_id != None and candidate_id != "":
						add_to_folder="c_"
						
						user_or_cid="c_id"
					else:
						add_to_folder="e_"
						candidate_id=ap_id
						user_or_cid="user_id"
					filename="Exit"	
					cand_id_exist=os.path.exists('/web/projop/xlplatpy/media/xlplat_ats/{}{}'.format(add_to_folder,candidate_id))
					if cand_id_exist == True:
						cand_file_exist=os.path.exists('/web/projop/xlplatpy/media/xlplat_ats/{}{}/{}'.format(add_to_folder,candidate_id,filename))
						if cand_file_exist == True:
							shutil.rmtree('/web/projop/xlplatpy/media/xlplat_ats/{}{}/{}'.format(add_to_folder,candidate_id,filename))
							os.mkdir(os.path.join('/web/projop/xlplatpy/media/xlplat_ats/{}{}'.format(add_to_folder,candidate_id), filename))
						else:
							os.mkdir(os.path.join('/web/projop/xlplatpy/media/xlplat_ats/{}{}'.format(add_to_folder,candidate_id), filename))
					else:
						os.mkdir(os.path.join('/web/projop/xlplatpy/media/xlplat_ats/',"{}{}".format(add_to_folder,candidate_id)))
						os.mkdir(os.path.join('/web/projop/xlplatpy/media/xlplat_ats/{}{}'.format(add_to_folder,candidate_id), filename))
					
					Document.objects.filter(uploaded_by=candidate_id,document_name=filename).delete()
					obj , crt = Document.objects.get_or_create(uploaded_by=candidate_id,document_name=filename,file_uploaded="exit_interview_form.pdf",user_or_cid=user_or_cid)
					t1 = threading.Thread(target=pdffromhtml, args=('http://localhost//py/xlplat_ats/exit_employee_form_internal/{}'.format(ap_id),"{}/exit_interview_form.pdf".format(os.path.join('/web/projop/xlplatpy/media/xlplat_ats/{}{}/{}'.format(add_to_folder,candidate_id,filename))),"{}/xlplat_ats/templates/{}.html".format(settings.BASE_DIR,header_name),))
					t1.start()
					return HttpResponseRedirect("/py/xlplat_ats/exit_employee_form/{}".format(ap_id))
				return render(request,'exit_employee_form.html',data_dict_exit)


			else:
				if "approve" in request.POST:
					application_data=Resignation_data.objects.get(resignation_of=ap_id,deleted=False)
					exit_form_json=json.loads(application_data.exit_form_json)
					exit_form_json['hr_remark']=request.POST['hr_remark']
					exit_form_json_str=json.dumps(exit_form_json)
					Resignation_data.objects.filter(resignation_of=ap_id,deleted=False).update(exit_approved=True,exit_form_json=exit_form_json_str)
					return HttpResponseRedirect("/py/xlplat_ats/exit_employee_form/{}".format(ap_id))
				elif "save_exit" in request.POST:
					return HttpResponseRedirect("/py/xlplat_ats/exit_employee_form/{}".format(ap_id))
				else:
					text_based_questions=[]
					manager_question_val=[]
					rating_val=[]
					holiday_question_val=[]
					exit_form_json={}
					text_based_questions=request.POST.getlist('text_based_questions')
					# exit_form_json['certify_check']=request.POST['certify_check']
					i=1
					exit_form_json['text_questions']={}
					for value in text_based_questions:
						exit_form_json['text_questions'][text_questions[i]]=value
						i+=1
					
					exit_form_json['manager_questions']={}
					for key,value in manager_questions.items():
						exit_form_json['manager_questions'][value]=request.POST["manager_question_val.{}".format(key)]
					
					exit_form_json['company_questions']={}
					for key,value in company_questions.items():
						exit_form_json['company_questions'][value]=request.POST["rating_val.{}".format(key)]
					exit_form_json['yes_or_no']=request.POST['yes_or_no']
					exit_form_json['yes_or_no_comments']=request.POST['yes_or_no_comments']
					exit_form_json['type_workload']=request.POST['type_workload']
					exit_form_json['holidays_questions']={}
					for key,value in holidays_questions.items():
						exit_form_json['holidays_questions'][value]=request.POST["holiday_question_val.{}".format(key)]
					exit_form_json['org_recommend']=request.POST['org_recommend']
					exit_form_json['ttc_better_work']=request.POST['ttc_better_work']
					exit_form_json_str=json.dumps(exit_form_json)
					Resignation_data.objects.filter(resignation_of=ap_id,deleted=False).update(exit_form_json=exit_form_json_str)
					form_filled=1
					show_interview_form_submission=1
					send_mail("HR_VP",ap_id,"exit_approval")
					data_dict_extend={"show_interview_form_submission":show_interview_form_submission,"message":"","form_filled":form_filled}
					return HttpResponseRedirect("/py/xlplat_ats/exit_employee_form/{}".format(ap_id))

			
		else:
			return render(request, 'insufficient_privilages.html', {'redirect_to':'/py/xlplat_ats/exit_employee_form/{}'.format(ap_id)})
	else:
		return render(request, 'insufficient_privilages.html', {'redirect_to':'/py/xlplat_ats/exit_employee_form/{}'.format(ap_id)})


def exit_employee_form_internal(request,ap_id):
		hr_view="1"
		employee_name=fetch_employee_name(ap_id)
		applicant_details=employee_details(ap_id)
		department_name=category_data(applicant_details[0])
		job_title_name=category_data((applicant_details[1]))
		joining_date=datetime.strptime(employee_end_date(ap_id).split(" ")[0],'%Y-%m-%d').date()
		manager=fetch_employee_name(applicant_details[2])
		ttc_code=applicant_details[3]
		if hr_view == "1":
			if "TTC-" in ttc_code or "TTC " in ttc_code:
				image_header="/py/media/images/TT-Consultants-patent-Logo.png"
				extra_image="width: 190px"
				company_text="TTC"
			elif "XLP" in ttc_code:
				image_header="/py/media/images/XLPAT.png"
				extra_image="width: 160px"
				company_text="XLSCOUT"
			elif "TTCS" in ttc_code or "TTCs" in ttc_code:
				# image_header="/py/media/images/ttcs_img.png"
				image_header="/py/media/images/TT-Consultants-patent-Logo.png"
				extra_image="width: 190px"
				company_text="TTC"
			elif "TA" in ttc_code:
				image_header="/py/media/images/ta_final_logo.jpg"
				extra_image="width: 155px"
				company_text="TA"
			elif "XLS" in ttc_code:
				image_header="/py/media/images/xlscout.png"
				extra_image="width: 175px"
				company_text="XLSCOUT"
			text_questions={1:"Why are you leaving "+company_text+"?",2:"What circumstances would have prevented your departure?",3:"What did you like most about your job?",4:"What did you like least about your job?"}
			manager_questions={1:"Was consistently fair",2:"Provided recognition",3:"Resolved complaints",4:"Was sensitive to employees' needs",5:"Provided feedback on performance",6:"Was receptive to open communication"}
			manager_options={1:"Almost Always",2:"Usually",3:"Sometimes",4:"Never"}
			company_questions={1:"Cooperation within your division/program",2:"Cooperation with other department",3:"Equipment provided (materials, resources, facilities)",4:"Company's performance review system",5:"Company's new employee orientation program",6:"Rate of pay for your job",7:"Career development/Advancement opportunities",8:"Physical working conditions"}
			company_options={1:"Excellent",2:"Good",3:"Fair",4:"Poor"}
			holidays_questions={1:"Paid holidays",2:"Medical plan",3:"Sick leave"}
			holidays_options={1:"Excellent",2:"Good",3:"Fair",4:"Poor",5:"No Opinion"}
			form_filled=0
			try:
				resignation_data_dict=Resignation_data.objects.get(resignation_of=ap_id,deleted=False)
				last_working_date=resignation_data_dict.relieve_date
				application_status_db=resignation_data_dict.application_status
				message="Resignation has not been Approved!"
				only_relieve_date=datetime.strptime(str(last_working_date),'%Y-%m-%d')
			except Resignation_data.DoesNotExist:
				application_status_db=""
				last_working_date=""
				message="Not applied for Resignation!"
				only_relieve_date=datetime.strptime(str("2099-12-31"),'%Y-%m-%d')
			three_days_relieve = only_relieve_date - timedelta(days=3)
			today = datetime.now()
			image_header="/py/media/images/TT-Consultants-patent-Logo.png"
			data_dict_exit={"applicant_id":ap_id,"employee_name":employee_name,'department_name':department_name,'job_title_name':job_title_name,'manager':manager,'joining_date':joining_date,'last_date':last_working_date,"hr_permission":hr_view,"image_header":image_header,"extra_image":extra_image}
			data_dict_extend={}
			if request.method == "GET":
				

				if application_status_db == "approved":
					

					exit_form_filled=resignation_data_dict.exit_form_json

					if exit_form_filled != "" :
						if hr_view == "1":
							exit_approved=resignation_data_dict.exit_approved
							show_interview_form_submission=0
							exit_form_filled=json.loads(exit_form_filled)
							text_questions_dict=exit_form_filled['text_questions']
							manager_questions_dict=exit_form_filled['manager_questions']
							company_questions_dict=exit_form_filled['company_questions']
							holidays_questions_dict_val=exit_form_filled['holidays_questions']
							yes_or_no_val=exit_form_filled['yes_or_no']
							yes_or_no_commentval=exit_form_filled['yes_or_no_comments']
							type_workload_val=exit_form_filled['type_workload']
							org_recommend_val=exit_form_filled['org_recommend']
							ttc_better_work_val=exit_form_filled['ttc_better_work']
							hr_remark=""
							if "hr_remark" in exit_form_filled: 
								hr_remark=exit_form_filled['hr_remark']
							data_dict_extend={"show_interview_form_submission":show_interview_form_submission,"text_questions_dict":text_questions_dict,"manager_questions_dict":manager_questions_dict,"company_questions_dict":company_questions_dict,"holidays_questions_dict_val":holidays_questions_dict_val,"yes_or_no_val":yes_or_no_val,"yes_or_no_commentval":yes_or_no_commentval,"type_workload_val":type_workload_val,"org_recommend_val":org_recommend_val,"ttc_better_work_val":ttc_better_work_val,"hr_remark":hr_remark,"exit_approved":exit_approved,"message":"","form_filled":form_filled}
						elif three_days_relieve <= today:
							data_dict_extend={"show_interview_form_submission":1,"message":"","form_filled":form_filled}
						else:
							data_dict_extend={"text_based_question":"","show_interview_form_submission":0,"message":"Not Allowed!!","form_filled":form_filled}
						
					elif hr_view == "0":
						if three_days_relieve <= today:
							data_dict_extend={"text_based_question":text_questions,"manager_questions":manager_questions,"company_questions":company_questions,"holidays_questions":holidays_questions,"show_interview_form_submission":0,"manager_options":manager_options,"company_options":company_options,"holidays_options":holidays_options,"form_filled":form_filled}
						else:
							data_dict_extend={"text_based_question":"","show_interview_form_submission":0,"form_filled":form_filled,"message":"Not Allowed!!"}
					else:
						data_dict_extend={"text_based_question":"","show_interview_form_submission":0,"form_filled":form_filled,"message":"User has not filled the Exit Interview Form"}
				else:
					data_dict_extend={"text_based_question":"","show_interview_form_submission":0,"form_filled":form_filled,"message":message}
				data_dict_exit.update(data_dict_extend)
				data_dict_exit["check_print"]=1
				return render(request,'exit_employee_form.html',data_dict_exit)




def employee_clearance_form(request,ap_id,type_val):
	user_id_val = XlplatAuthMiddleware.user_id
	if user_id_val != 0:
		final_val = 1
	else: 
		final_val = 0
	if final_val == 1 and type_val != "":
		hr_view=check_for_permission_assigned(user_id_val)
		if str(hr_view) == "1" and user_id_val == ap_id:
			hr_view = "0"
		employee_name=fetch_employee_name(ap_id)
		applicant_details=employee_details(ap_id)
		department_name=category_data(applicant_details[0])
		job_title_name=category_data((applicant_details[1]))
		joining_date=datetime.strptime(employee_end_date(ap_id).split(" ")[0],'%Y-%m-%d').date()
		manager=fetch_employee_name(applicant_details[2])
		ttc_code=applicant_details[3]
		vp_id=get_vp_id_db(ap_id)
		vp_name=fetch_employee_name(vp_id)
		loan_amount_paid=get_loan_paid_amt(ap_id)
		amount_taken= get_loan_approved_amt(ap_id)
	    
		amount_paid_statement="Total amount of loan taken {} \nTotal amount paid {}".format(amount_taken,loan_amount_paid)
		# check of it, finance,admin
		#check for vp for manager clearance
		vp_view=0
		if str(vp_id)==str(user_id_val):
			vp_view=1
		manager_view=0
		if str(applicant_details[2])==user_id_val or str(applicant_details[4]) == user_id_val:
			manager_view=1
		if manager_view==1 or hr_view == "1" or vp_view == 1:
			view_or_edit=0
		else:
			view_or_edit=1
		employee_questions={"Handing over done to:":"","Data Backup and Mails Backup/Re-direct to:":""}
		it_questions={"Desktop/ Dock station/ Laptop":["",""],"Head Phones":["",""],"Other Assets handled by employee":["",""],"Cell Phone/Sim/ Data Card":["",""],"User Account/Mail Access":["",""],"VPN/Phone Access":["",""],"Other Applications":["",""],"Removal of Biometric ID":["",""]}
		admin_questions={"Pedestal Keys":"","Car/ Bike stickers":"","Others (Misc. Dues)":""}
		accounts_questions={"Loan Taken":amount_paid_statement,"Travel Advance":"","TAX form filled on XLPLAT":""}
		hr_questions={"Exit Interview Form":"","Employee Identity Card/Access Card":""}
		print_check=0
		each_leave=""
		if 	request.method == "POST":
			if "save_exit" in request.POST:
				print_check=1
		if request.method == "GET" or print_check == 1:
			form_filled=0
			try:
				resignation_data_dict=Resignation_data.objects.get(resignation_of=ap_id,deleted=False)
				last_working_date=resignation_data_dict.relieve_date
				only_relieve_date=datetime.strptime(str(last_working_date),'%Y-%m-%d')
				application_status_db=resignation_data_dict.application_status
				resign_date=resignation_data_dict.user_resign_date
				manager_clearance=resignation_data_dict.manager_clearance_json
				if manager_clearance != "":
					manager_clearance=json.loads(manager_clearance)
				accounts_clearance_json=resignation_data_dict.accounts_clearance_json
				if accounts_clearance_json != "":
					accounts_questions=json.loads(accounts_clearance_json)
				admin_clearance_json=resignation_data_dict.admin_clearance_json
				if admin_clearance_json != "":
					admin_questions=json.loads(admin_clearance_json)
				it_clearance_json=resignation_data_dict.it_clearance_json
				if it_clearance_json != "":
					it_questions=json.loads(it_clearance_json)
				hr_clearance_json=resignation_data_dict.hr_clearance_json
				if hr_clearance_json != "":
					hr_questions=json.loads(hr_clearance_json)
				
				activities_list={}
				if "activities_list" in manager_clearance:
					activities_list=manager_clearance["activities_list"]
					manager_clearance.pop('activities_list')
					employee_questions=manager_clearance
				manager_approved=resignation_data_dict.manager_clearance_approved
				it_approved=resignation_data_dict.it_clearance_approved
				admin_approved=resignation_data_dict.admin_clearance_approved
				hr_approved=resignation_data_dict.hr_clearance_approved
				accounts_approved=resignation_data_dict.accounts_it_clearance_approved
				employee_clearance_filled=resignation_data_dict.employee_clearance_filled
				manager_clearance_date=resignation_data_dict.manager_clearance_date
				it_clearance_date=resignation_data_dict.it_clearance_date
				admin_clearance_date=resignation_data_dict.admin_clearance_date
				hr_clearance_date=resignation_data_dict.hr_clearance_date
				accounts_clearance_date=resignation_data_dict.accounts_clearance_date
				processing_date=resignation_data_dict.processing_date
				address_correspondence=resignation_data_dict.correspondence_address
				hr_clearance_approved_final=resignation_data_dict.hr_clearance_approved_final
				hr_clearance_date_final=resignation_data_dict.hr_clearance_date_final
				send_to_vp_val=resignation_data_dict.send_to_vp

				if address_correspondence==None:
					address_correspondence=""
			except Resignation_data.DoesNotExist:
				application_status_db,last_working_date,resign_date,manager_approved,it_approved,hr_approved,admin_approved,accounts_approved,manager_clearance_date,it_clearance_date,hr_clearance_date,admin_clearance_date,accounts_clearance_date,processing_date,hr_clearance_date_final,send_to_vp_val=("",)*16
				manager_clearance={}
				accounts_clearance_json={}
				admin_clearance_json={}
				it_clearance_json={}
				activities_list={}
				employee_clearance_filled,hr_clearance_approved_final=(False,)*2
				only_relieve_date=datetime.strptime(str("2099-12-31"),'%Y-%m-%d')
			three_days_relieve = only_relieve_date - timedelta(days=3)
			today = datetime.now()
			if processing_date=="":
				processing_date=date.today()
			if employee_clearance_filled == False and view_or_edit==0:
				message_emp="{} has not filled the clearance form!!".format(employee_name)
			elif employee_clearance_filled == True and user_id_val==ap_id:
				message_emp="Form Filled Successfully!!"
			else:
				message_emp=""
			if vp_view == 1 and send_to_vp_val == False and manager_view != 1:
				message_emp="Manager has not approved the clearance form"

			if manager_view == 1 and send_to_vp_val == True and vp_view != 1:
				message_emp="Sent to {} for approval".format(vp_name)
			if hr_view == "1":
				vp_view = 1
			message_it=""
			message_admin=""
			message_finance=""
			message_hr=""
			message_all=""
			if employee_clearance_filled == False:
				message_it= "{} has not filled the clearance form!!".format(employee_name)
				message_admin="{} has not filled the clearance form!!".format(employee_name)
				message_finance="{} has not filled the clearance form!!".format(employee_name)
			elif manager_approved == False: 
				message_it= "Manager has not approved the clearance form"
				message_admin="Manager has not approved the clearance form"
				message_finance="Manager has not approved the clearance form"
			# check if it department or hr
			
			it_view=0
			admin_view=0
			finance_view=0
			if type_val == "IT":
				it_view=permission_check(user_id_val,"Helpdesk")
				if str(it_view) != "1" and str(hr_view) != "1":
					message_it="You don't have enough permissions!!!"
			elif type_val == "ADMIN":
				admin_view=permission_check(user_id_val,"Office Admin")
				if str(admin_view) != "1" and str(hr_view) != "1":
					message_admin="You don't have enough permissions!!!"
			elif type_val == "ACCOUNTS":
				finance_view=permission_check(user_id_val,"Accounting")
				if str(finance_view) != "1" and str(hr_view) != "1":
					message_finance="You don't have enough permissions!!!"

			if type_val=="MANAGER":
				if view_or_edit!=0 and user_id_val!=ap_id:
					message_emp="Access Denied"
			if str(hr_view) != "1":
				message_hr="You don't have enough permissions!!!"
			else:
				it_view=1
				admin_view=1
				finance_view=1
			image_header="/py/media/images/TT-Consultants-patent-Logo.png"
			header_name="header_ttc"
			org_name="TT Consultants"
			if "TTC-" in ttc_code or "TTC " in ttc_code:
				image_header="/py/media/images/TT-Consultants-patent-Logo.png"
				image_style="width:24vh;"
				header_name="header_ttc"
			elif "XLP" in ttc_code:
				image_header="/py/media/images/XLPAT.png"
				image_style="width:17vh;"
				header_name="header_xl"
				org_name="XLSCOUT"
			elif "TTCS" in ttc_code or "TTCs" in ttc_code:
				# image_header="/py/media/images/ttcs_img.png"
				image_header="/py/media/images/TT-Consultants-patent-Logo.png"
				image_style="width:24vh;"
				header_name="header_ttcs"
			elif "TA" in ttc_code:
				image_header="/py/media/images/ta_final_logo.jpg"
				image_style="width:16vh;"
				header_name="header_ta"
				org_name="Talwar Advocates"
			elif "XLS" in ttc_code:
				image_header="/py/media/images/xlscout.png"
				extra_image="width: 22vh"
				header_name="header_xls"
				org_name="XLSCOUT"
			send_dict={"applicant_id":ap_id,"type_val":type_val,"employee_name":employee_name,'department_name':department_name,'ttc_code':ttc_code,"c_date":processing_date,'resign_date':resign_date,'job_title_name':job_title_name,'manager':manager,'joining_date':joining_date,'last_date':last_working_date,"image_header":image_header,"image_style":image_style,'org_name':org_name}
			

			if three_days_relieve > today:
				send_dict.update({"message_date":"Form Closed!!!"})
			else:

				if type_val == "ALL" and str(hr_view) == "1":
					view_all=1
				else:
					view_all=0
				if type_val == "ALL" and (admin_approved != True or accounts_approved != True or it_approved != True or hr_approved != True or view_all== 0):
					view_all=0
					message_all="Not Permitted!!"
					send_dict.update({"message_all":message_all})
				else:
					if type_val=="MANAGER" or type_val=="ALL":
						send_dict.update({"employee_questions":employee_questions,"manager_approved":manager_approved,"message_emp":message_emp,"activities_count":len(activities_list),"activities_list":activities_list,"view_or_edit":view_or_edit,"manager_clearance_date":manager_clearance_date,"address_corr":address_correspondence,"vp_view":vp_view,"send_to_vp_val":send_to_vp_val})
					if str(it_view) == "1" or type_val=="ALL":
						send_dict.update({"it_questions":it_questions,"it_approved":it_approved,"message_it":message_it,"it_clearance_date":it_clearance_date})
					if str(finance_view) == "1" or type_val=="ALL":
						send_dict.update({"accounts_questions":accounts_questions,"accounts_approved":accounts_approved,"message_accounts":message_finance,"accounts_clearance_date":accounts_clearance_date})
					if str(admin_view) or type_val=="ALL":
						send_dict.update({"admin_questions":admin_questions,"admin_approved":admin_approved,"message_admin":message_admin,"admin_clearance_date":admin_clearance_date})
					if (type_val=="HR" and str(hr_view) == "1") or type_val=="ALL":
						sick_leaves_left=0
						paid_leaves_left=0
						casual_leaves_left=0
						get_leaves_count_all=get_leaves_left(ap_id,last_working_date)
						if get_leaves_count_all != "":
							get_leaves_count_list = get_leaves_count_all.split("~")
							paid_leaves_left=get_leaves_count_list[0]
							sick_leaves_left=get_leaves_count_list[1]
							casual_leaves_left=get_leaves_count_list[2]
						if "XLP" in ttc_code:
							expected_last_date = resign_date +relativedelta(months=+2)
						else:
							expected_last_date = resign_date + relativedelta(months=+1)
						notice_period_fault=expected_last_date-last_working_date
						
						send_dict.update({"sick_leaves":sick_leaves_left,"casual_leaves":casual_leaves_left,"paid_leaves":paid_leaves_left,"expected_last_date":expected_last_date,"final_short":notice_period_fault.days,"hr_questions":hr_questions,"address_corr":address_correspondence,"message_hr":message_hr,"hr_approved":hr_approved,"hr_clearance_date":hr_clearance_date})
					if type_val=="ALL":
						send_dict.update({"message_all":message_all,"hr_clearance_approved_final":hr_clearance_approved_final,"hr_clearance_date_final":hr_clearance_date_final})
				send_dict.update({"view_all":view_all,"message_date":""})
			if print_check==1:
				# media/xlplat_ats
				candidate_id=get_candidate_id(ap_id)
				if candidate_id != None and candidate_id != "":
					add_to_folder="c_"
					
					user_or_cid="c_id"
				else:
					add_to_folder="e_"
					candidate_id=ap_id
					user_or_cid="user_id"
				filename="Employee"	
				cand_id_exist=os.path.exists('/web/projop/xlplatpy/media/xlplat_ats/{}{}'.format(add_to_folder,candidate_id))
				if cand_id_exist == True:
					cand_file_exist=os.path.exists('/web/projop/xlplatpy/media/xlplat_ats/{}{}/{}'.format(add_to_folder,candidate_id,filename))
					if cand_file_exist == True:
						shutil.rmtree('/web/projop/xlplatpy/media/xlplat_ats/{}{}/{}'.format(add_to_folder,candidate_id,filename))
						os.mkdir(os.path.join('/web/projop/xlplatpy/media/xlplat_ats/{}{}'.format(add_to_folder,candidate_id), filename))
					else:
						os.mkdir(os.path.join('/web/projop/xlplatpy/media/xlplat_ats/{}{}'.format(add_to_folder,candidate_id), filename))
				else:
					os.mkdir(os.path.join('/web/projop/xlplatpy/media/xlplat_ats/',"{}{}".format(add_to_folder,candidate_id)))
					os.mkdir(os.path.join('/web/projop/xlplatpy/media/xlplat_ats/{}{}'.format(add_to_folder,candidate_id), filename))
				
				Document.objects.filter(uploaded_by=candidate_id,document_name=filename).delete()
				obj , crt = Document.objects.get_or_create(uploaded_by=candidate_id,document_name=filename,file_uploaded="employee_clearance_form.pdf",user_or_cid=user_or_cid)
				t1 = threading.Thread(target=pdffromhtml, args=('http://localhost//py/xlplat_ats/employee_clearance_form_internal/{}/ALL'.format(ap_id),"{}/employee_clearance_form.pdf".format(os.path.join('/web/projop/xlplatpy/media/xlplat_ats/{}{}/{}'.format(add_to_folder,candidate_id,filename))),"{}/xlplat_ats/templates/{}.html".format(settings.BASE_DIR,header_name),))
				t1.start()
				return HttpResponseRedirect("/py/xlplat_ats/employee_clearance_form/{}/{}".format(ap_id,"ALL"))
			return render(request,'employee_clearance_form.html',send_dict)
			
		else:
			current_date=date.today().strftime("%Y-%m-%d")
			if "emp_clear_sub" in request.POST:
				type_val="MANAGER"
				itr = 1
				manager_clearance_final={}
				manager_q_1={}
				manager_q_2={}
				while itr:
					if "activity_{}".format(itr) in request.POST and "status_val_{}".format(itr) in request.POST:
						manager_q_1[request.POST["activity_{}".format(itr)].replace("\"","")]=request.POST["status_val_{}".format(itr)].replace("\"","")
					else:
						break
					itr+=1
				itr = 1
				manager_clearance_final['activities_list']=manager_q_1
				while itr <= len(employee_questions):
					if "employee_a_{}".format(itr) in request.POST and "employee_q_{}".format(itr) in request.POST:
						manager_clearance_final[request.POST["employee_q_{}".format(itr)].replace("\"","")]=request.POST["employee_a_{}".format(itr)].replace("\"","")
					itr+=1
				
				correspond_addr=""
				if "correspond_addr" in request.POST:
					correspond_addr=request.POST["correspond_addr"].replace("\"","")
				send_mail("MANAGER1_MANAGER",ap_id,"manager_clearance_approve")
				tomorrow=date.today() + timedelta(days=1)
				send_mail_reminder(ap_id,"05:00:00",tomorrow,1)
				final_dict={"manager_clearance_json":json.dumps(manager_clearance_final),"employee_clearance_filled":True,"processing_date":current_date,"correspondence_address":correspond_addr}
			elif "emp_clear_app" in request.POST:
				send_mail("HR_IT_ADMIN_ACCOUNTS",ap_id,"manager_clearance_mail")
				tomorrow=date.today() + timedelta(days=1)
				send_mail_reminder(ap_id,"03:00:00",tomorrow,1)
				type_val="MANAGER"
				final_dict={"manager_clearance_approved":True,"manager_clearance_date":current_date}
			elif "emp_clear_rej" in request.POST:
				type_val="MANAGER"
				final_dict={"employee_clearance_filled":False}
			elif "emp_clear_send" in request.POST:
				send_mail("VP",ap_id,"manager_clearance_approve")
				tomorrow=date.today() + timedelta(days=1)
				send_mail_reminder(ap_id,"06:00:00",tomorrow,1)
				type_val="MANAGER"
				final_dict={"send_to_vp":True}
			elif "emp_clear_vp_rej" in request.POST:
				type_val="MANAGER"
				final_dict={"send_to_vp":False,"employee_clearance_filled":False}
			elif "it_approve" in request.POST:
				type_val="IT"
				itr = 1
				it_clearance_final={}
				while itr <= len(it_questions):
					if "it_stat_{}".format(itr) in request.POST and "it_ques_{}".format(itr) in request.POST:
						it_clearance_final[request.POST["it_ques_{}".format(itr)].replace("\"","")]=[request.POST["it_stat_{}".format(itr)].replace("\"",""),request.POST["it_remark_{}".format(itr)].replace("\"","")]
					itr+=1
				final_dict={"it_clearance_json":json.dumps(it_clearance_final),"it_clearance_approved":True,"it_clearance_date":current_date}
			elif "admin_approve" in request.POST:
				type_val="ADMIN"
				itr = 1
				admin_clearance_final={}
				while itr <= len(it_questions):
					if "admin_stat_{}".format(itr) in request.POST and "admin_ques_{}".format(itr) in request.POST:
						admin_clearance_final[request.POST["admin_ques_{}".format(itr)].replace("\"","")]=request.POST["admin_stat_{}".format(itr)].replace("\"","")
					itr+=1
				final_dict={"admin_clearance_json":json.dumps(admin_clearance_final),"admin_clearance_approved":True,"admin_clearance_date":current_date}
			elif "finance_approve" in request.POST or "finance_save" in request.POST:
				type_val="ACCOUNTS"
				itr = 1
				accounts_clearance_final={}
				while itr <= len(accounts_questions):
					if "acc_ques_{}".format(itr) in request.POST and "acc_ans_{}".format(itr) in request.POST:
						accounts_clearance_final[request.POST["acc_ques_{}".format(itr)].replace("\"","")]=request.POST["acc_ans_{}".format(itr)].replace("\"","")
					itr+=1
				if "finance_save" in request.POST:
					final_dict={"accounts_clearance_json":json.dumps(accounts_clearance_final)}
				else:
					final_dict={"accounts_clearance_json":json.dumps(accounts_clearance_final),"accounts_it_clearance_approved":True,"accounts_clearance_date":current_date}
			elif "hr_approve" in request.POST or "hr_save" in request.POST:
				type_val="HR"
				itr = 1
				hr_clearance_final={}
				while itr <= len(accounts_questions):
					if "hr_ques_{}".format(itr) in request.POST and "hr_stat_{}".format(itr) in request.POST:
						hr_clearance_final[request.POST["hr_ques_{}".format(itr)].replace("\"","")]=request.POST["hr_stat_{}".format(itr)].replace("\"","")
					itr+=1
				
				if "hr_save" in request.POST:
					final_dict={"hr_clearance_json":json.dumps(hr_clearance_final),"correspondence_address":request.POST["correspond_addr"]}
				else:
					final_dict={"hr_clearance_json":json.dumps(hr_clearance_final),"hr_clearance_approved":True,"correspondence_address":request.POST["correspond_addr"],"hr_clearance_date":current_date}

			elif "hr_final_approve" in request.POST:
				type_val="ALL"
				final_dict={"hr_clearance_approved_final":True,"hr_clearance_date_final":current_date}
			Resignation_data.objects.filter(resignation_of=ap_id,deleted=False).update(**final_dict)


			try:
				resignation_data_dict=Resignation_data.objects.get(resignation_of=ap_id,deleted=False)
				manager_approved=resignation_data_dict.manager_clearance_approved
				it_approved=resignation_data_dict.it_clearance_approved
				admin_approved=resignation_data_dict.admin_clearance_approved
				hr_approved=resignation_data_dict.hr_clearance_approved
				accounts_approved=resignation_data_dict.accounts_it_clearance_approved
				hr_clearance_approved_final=resignation_data_dict.hr_clearance_approved_final
			except Resignation_data.DoesNotExist:
				manager_approved,it_approved,admin_approved,hr_approved,accounts_approved,hr_clearance_approved_final=(False,)*6
			if manager_approved and it_approved and admin_approved and hr_approved and accounts_approved and not hr_clearance_approved_final :
				send_mail_reminder(ap_id,"04:00:00",date.today(),1)
			return HttpResponseRedirect("/py/xlplat_ats/employee_clearance_form/{}/{}".format(ap_id,type_val))	

	else:
		return render(request, 'insufficient_privilages.html', {'redirect_to':"/py/xlplat_ats/employee_clearance_form/{}/{}".format(ap_id,type_val)})

def pdffromhtml(url,path,header):
	pdfkit.from_url(url,path,options = {
				'page-size': 'Letter',
				'margin-top': '1in',
				'margin-right': '0.7in',
				'margin-bottom': '0.9in',
				'margin-left': '0.7in',
				'encoding': "UTF-8",
				'header-html': header,
				'no-outline':None
				})

def employee_clearance_form_internal(request,ap_id,type_val):
	hr_view="1"
	employee_name=fetch_employee_name(ap_id)
	applicant_details=employee_details(ap_id)
	department_name=category_data(applicant_details[0])
	job_title_name=category_data((applicant_details[1]))
	joining_date=datetime.strptime(employee_end_date(ap_id).split(" ")[0],'%Y-%m-%d').date()
	manager=fetch_employee_name(applicant_details[2])
	ttc_code=applicant_details[3]
	# check of it, finance,admin
	if hr_view == "1":
		view_or_edit=0
	else:
		view_or_edit=1
	employee_questions={"Handing over done to:":"","Data Backup and Mails Backup/Re-direct to:":""}
	it_questions={"Desktop/ Dock station/ Laptop":["",""],"Head Phones":["",""],"Other Assets handled by employee":["",""],"Cell Phone/Sim/ Data Card":["",""],"User Account/Mail Access":["",""],"VPN/Phone Access":["",""],"Other Applications":["",""],"Removal of Biometric ID":["",""]}
	admin_questions={"Pedestal Keys":"","Car/ Bike stickers":"","Others (Misc. Dues)":""}
	accounts_questions={"Loans Advance":"","Travel Advance":"","TAX form filled on XLPLAT":""}
	hr_questions={"Exit Interview Form":"","Employee Identity Card/Access Card":""}
	if request.method == "GET":
		form_filled=0
		try:
			resignation_data_dict=Resignation_data.objects.get(resignation_of=ap_id,deleted=False)
			last_working_date=resignation_data_dict.relieve_date
			only_relieve_date=datetime.strptime(str(last_working_date),'%Y-%m-%d')
			application_status_db=resignation_data_dict.application_status
			resign_date=resignation_data_dict.user_resign_date
			manager_clearance=resignation_data_dict.manager_clearance_json
			if manager_clearance != "":
				manager_clearance=json.loads(manager_clearance)
			accounts_clearance_json=resignation_data_dict.accounts_clearance_json
			if accounts_clearance_json != "":
				accounts_questions=json.loads(accounts_clearance_json)
			admin_clearance_json=resignation_data_dict.admin_clearance_json
			if admin_clearance_json != "":
				admin_questions=json.loads(admin_clearance_json)
			it_clearance_json=resignation_data_dict.it_clearance_json
			if it_clearance_json != "":
				it_questions=json.loads(it_clearance_json)
			hr_clearance_json=resignation_data_dict.hr_clearance_json
			if hr_clearance_json != "":
				hr_questions=json.loads(hr_clearance_json)
			
			activities_list={}
			if "activities_list" in manager_clearance:
				activities_list=manager_clearance["activities_list"]
				manager_clearance.pop('activities_list')
				employee_questions=manager_clearance
			manager_approved=resignation_data_dict.manager_clearance_approved
			it_approved=resignation_data_dict.it_clearance_approved
			admin_approved=resignation_data_dict.admin_clearance_approved
			hr_approved=resignation_data_dict.hr_clearance_approved
			accounts_approved=resignation_data_dict.accounts_it_clearance_approved
			employee_clearance_filled=resignation_data_dict.employee_clearance_filled
			manager_clearance_date=resignation_data_dict.manager_clearance_date
			it_clearance_date=resignation_data_dict.it_clearance_date
			admin_clearance_date=resignation_data_dict.admin_clearance_date
			hr_clearance_date=resignation_data_dict.hr_clearance_date
			accounts_clearance_date=resignation_data_dict.accounts_clearance_date
			processing_date=resignation_data_dict.processing_date
			address_correspondence=resignation_data_dict.correspondence_address
			hr_clearance_approved_final=resignation_data_dict.hr_clearance_approved_final
			hr_clearance_date_final=resignation_data_dict.hr_clearance_date_final
			if address_correspondence==None:
				address_correspondence=""
		except Resignation_data.DoesNotExist:
			application_status_db,last_working_date,resign_date,manager_approved,it_approved,hr_approved,admin_approved,accounts_approved,manager_clearance_date,it_clearance_date,hr_clearance_date,admin_clearance_date,accounts_clearance_date,processing_date,hr_clearance_date_final=("",)*15
			manager_clearance={}
			accounts_clearance_json={}
			admin_clearance_json={}
			it_clearance_json={}
			activities_list={}
			employee_clearance_filled,hr_clearance_approved_final=(False,)*2
			only_relieve_date=datetime.strptime(str("2099-12-31"),'%Y-%m-%d')
		three_days_relieve = only_relieve_date - timedelta(days=3)
		today = datetime.now()
		if processing_date=="":
			processing_date=date.today()
		if employee_clearance_filled == False and view_or_edit==0:
			message_emp="{} has not filled the clearance form!!".format(employee_name)
		else:
			message_emp=""
		message_it=""
		message_admin=""
		message_finance=""
		message_hr=""
		message_all=""
		if employee_clearance_filled == False:
			message_it= "{} has not filled the clearance form!!".format(employee_name)
			message_admin="{} has not filled the clearance form!!".format(employee_name)
			message_finance="{} has not filled the clearance form!!".format(employee_name)
		elif manager_approved == False: 
			message_it= "Manager has not approved the clearance form"
			message_admin="Manager has not approved the clearance form"
			message_finance="Manager has not approved the clearance form"
		# check if it department or hr
		
		it_view=0
		admin_view=0
		finance_view=0
		if type_val == "IT":
			it_view=permission_check(user_id_val,"Helpdesk")
			if str(it_view) != "1" and str(hr_view) != "1":
				message_it="You don't have enough permissions!!!"
		elif type_val == "ADMIN":
			admin_view=permission_check(user_id_val,"Office Admin")
			if str(admin_view) != "1" and str(hr_view) != "1":
				message_admin="You don't have enough permissions!!!"
		elif type_val == "ACCOUNTS":
			finance_view=permission_check(user_id_val,"Accounting")
			if str(finance_view) != "1" and str(hr_view) != "1":
				message_finance="You don't have enough permissions!!!"

		if str(hr_view) != "1":
			message_hr="You don't have enough permissions!!!"
		else:
			it_view=1
			admin_view=1
			finance_view=1
		image_header="/py/media/images/TT-Consultants-patent-Logo.png"
		org_name="TT Consultants"
		if "TTC-" in ttc_code or "TTC " in ttc_code:
			image_header="/py/media/images/TT-Consultants-patent-Logo.png"
			image_style="width:24vh;"
		elif "XLP" in ttc_code:
			image_header="/py/media/images/XLPAT.png"
			image_style="width:17vh;"
			org_name="XLSCOUT"
		elif "TTCS" in ttc_code or "TTCs" in ttc_code:
			# image_header="/py/media/images/ttcs_img.png"
			image_header="/py/media/images/TT-Consultants-patent-Logo.png"
			image_style="width:24vh;"
		elif "TA" in ttc_code:
			image_header="/py/media/images/ta_final_logo.jpg"
			image_style="width:16vh;"
			org_name="Talwar Advocates"
		elif "XLS" in ttc_code:
			image_header="/py/media/images/xlscout.png"
			extra_image="width: 22vh"
			org_name="XLSCOUT"
		send_dict={"applicant_id":ap_id,"type_val":type_val,"employee_name":employee_name,'department_name':department_name,'ttc_code':ttc_code,"c_date":processing_date,'resign_date':resign_date,'job_title_name':job_title_name,'manager':manager,'joining_date':joining_date,'last_date':last_working_date,"image_header":image_header,"image_style":image_style,"org_name":org_name}


		if three_days_relieve > today:
			send_dict.update({"message_date":"Form Closed!!!"})
		else:

			if type_val == "ALL" and str(hr_view) == "1":
				view_all=1
			else:
				view_all=0
			if type_val == "ALL" and (admin_approved != True or accounts_approved != True or it_approved != True or hr_approved != True or view_all== 0):
				view_all=0
				message_all="Not Permitted!!"
				send_dict.update({"message_all":message_all})
			else:
				if type_val=="MANAGER" or type_val=="ALL":
					send_dict.update({"employee_questions":employee_questions,"manager_approved":manager_approved,"message_emp":message_emp,"activities_count":len(activities_list),"activities_list":activities_list,"view_or_edit":view_or_edit,"manager_clearance_date":manager_clearance_date,"address_corr":address_correspondence})
				if str(it_view) == "1" or type_val=="ALL":
					send_dict.update({"it_questions":it_questions,"it_approved":it_approved,"message_it":message_it,"it_clearance_date":it_clearance_date})
				if str(finance_view) == "1" or type_val=="ALL":
					send_dict.update({"accounts_questions":accounts_questions,"accounts_approved":accounts_approved,"message_accounts":message_finance,"accounts_clearance_date":accounts_clearance_date})
				if str(admin_view) or type_val=="ALL":
					send_dict.update({"admin_questions":admin_questions,"admin_approved":admin_approved,"message_admin":message_admin,"admin_clearance_date":admin_clearance_date})
				if (type_val=="HR" and str(hr_view) == "1") or type_val=="ALL":
					sick_leaves_left=0
					paid_leaves_left=0
					casual_leaves_left=0
					get_leaves_count_all=get_leaves_left(ap_id,last_working_date)
					if get_leaves_count_all != "":
						get_leaves_count_list = get_leaves_count_all.split("~")
						paid_leaves_left=get_leaves_count_list[0]
						sick_leaves_left=get_leaves_count_list[1]
						casual_leaves_left=get_leaves_count_list[2]
					if "XLP" in ttc_code:
						expected_last_date = resign_date +relativedelta(months=+2)
					else:
						expected_last_date = resign_date + relativedelta(months=+1)
					notice_period_fault=expected_last_date-last_working_date
					
					send_dict.update({"sick_leaves":sick_leaves_left,"casual_leaves":casual_leaves_left,"paid_leaves":paid_leaves_left,"expected_last_date":expected_last_date,"final_short":notice_period_fault.days,"hr_questions":hr_questions,"address_corr":address_correspondence,"message_hr":message_hr,"hr_approved":hr_approved,"hr_clearance_date":hr_clearance_date})
				if type_val=="ALL":
					send_dict.update({"message_all":message_all,"hr_clearance_approved_final":hr_clearance_approved_final,"hr_clearance_date_final":hr_clearance_date_final})
			send_dict.update({"view_all":view_all,"message_date":""})
		send_dict["print_check"]=1
		return render(request,'employee_clearance_form.html',send_dict)

def relieving_letter_emp(request,ap_id):
	user_id_val = XlplatAuthMiddleware.user_id
	if user_id_val != 0:
		final_val = 1
	else: 
		final_val = 0
	if final_val == 1 and user_id_val != ap_id:
		hr_view=str(check_for_permission_assigned(user_id_val))
		if hr_view == "1":
			applicant_details=employee_details(ap_id)
			ttc_code=applicant_details[3]
			company_name=""
			organiztion_unit=""
			document_type=""
			if "TTC-" in ttc_code or "TTC " in ttc_code:
				company_name="Talwar & Talwar Consultants Pvt. Ltd."
				organiztion_unit="TTC"
				document_type=701
			elif "XLP" in ttc_code:
				company_name="XLPAT TT CONSULTANTS PVT. LTD."
				organiztion_unit="XLP"
				document_type=703
			elif "TTCS" in ttc_code or "TTCs" in ttc_code:
				company_name="Talwar & Talwar Consultants and Services Pvt. Ltd."
				organiztion_unit="TTCS"
				document_type=707
			elif "TA" in ttc_code:
				company_name="Talwar Advocates"
				organiztion_unit="TA"
				document_type=705
			elif "XLS" in ttc_code:
				company_name="XLSCOUT XLPAT Pvt. Ltd"
				organiztion_unit="XLS"
				document_type=709
			if request.method == "POST":
				revert_address="/py/xlplat_ats/relieving_letter/{}".format(ap_id)
				resignation_approved=""
				try:
					ref_bool=Resignation_data.objects.get(resignation_of=ap_id,deleted=False).relieving_reference_no
				except Resignation_data.DoesNotExist:
					ref_bool=""
				if 'save_template' in request.POST and ref_bool != "":
					# assign document_numbers to docs if relieving resignation false
					mail_body=request.POST["template_body"]
					if ref_bool == False:
						now = datetime.now()
						doc_ref_year= now.year
						month= now.month
						if( month < 4 ):
							doc_ref_year=doc_ref_year-1
						f_year="{}-{}".format(doc_ref_year,doc_ref_year+1)
						document_ref_number_already_exist=get_doc_number(document_type,doc_ref_year)
						if document_ref_number_already_exist != None:
							document_number=document_ref_number_already_exist+1
						else:
							document_number=1
						document_number_2=document_number+1
						insert_doc_number(document_number,document_type,ap_id,user_id_val,now.strftime("%Y-%m-%d"),"Relieving Letter",False,doc_ref_year)
						insert_doc_number(document_number_2,document_type,ap_id,user_id_val,now.strftime("%Y-%m-%d"),"Relieving Letter 2",False,doc_ref_year)

						reference_number="HRD/REC/{}/{}/{}".format(organiztion_unit,document_number,f_year)
						reference_number_2="HRD/REC/{}/{}/{}".format(organiztion_unit,document_number_2,f_year)
						mail_body_f=mail_body.replace("Ref. No.: <b></b>","Ref. No: {}".format(reference_number))
						mail_body=mail_body_f.replace("Ref. No:      <b></b>","Ref. No: {}".format(reference_number_2))
					Resignation_data.objects.filter(resignation_of=ap_id,deleted=False).update(relieving_template=mail_body,relieving_reference_no=True)
				return HttpResponseRedirect(revert_address)
			else:
				ap_application_status,relieve_temp,message=("",)*3
				ap_name=fetch_employee_name(ap_id)
				
				department_name=category_data(applicant_details[0])
				job_title_name=category_data((applicant_details[1]))
				joining_date=datetime.strptime(employee_end_date(ap_id).split(" ")[0],'%Y-%m-%d').date()
				
				try:
					resignation_data_dict=Resignation_data.objects.get(resignation_of=ap_id,deleted=False)
				except Resignation_data.DoesNotExist:
					resignation_data_dict=""
				send_dict={}
				if not resignation_data_dict:
					ap_relieve_date=""
					ap_resign_reason=""
					approval_val=""
					message="Not Permitted!!!"
				else:
					ap_relieve_date=resignation_data_dict.relieve_date
					resign_date=resignation_data_dict.user_resign_date
					ap_application_status=resignation_data_dict.application_status
					relieve_temp=resignation_data_dict.relieving_template
					if relieve_temp == None:
						relieve_temp=""
					if ap_application_status != "approved":
						message="Not Permitted!!!"
					

			return render(request,'relieving_letter.html',{"applicant_id":ap_id,"relieve_temp":relieve_temp,"c_date":date.today().strftime('%d %b, %Y'),"employee_name":ap_name,"ttc_code":ttc_code,"company_name":company_name,"designation":job_title_name,"start_date":joining_date.strftime('%d %b, %Y'),"relieve_date":ap_relieve_date.strftime('%d %b, %Y'),"resign_date":resign_date.strftime('%d %b, %Y'),"message":message})
		else:
			return render(request, 'insufficient_privilages.html', {'redirect_to':'/py/xlplat_ats/relieving_letter/{}/'.format(ap_id)})	
	else:
		return render(request, 'insufficient_privilages.html', {'redirect_to':'/py/xlplat_ats/relieving_letter/{}'.format(ap_id)})

def get_doc_number(doc_type,doc_year):
	if str(doc_type) in ["709","703"]:
		query="select document_number::integer from xl_ats_document_reference_number where document_type in ('{}','{}') and doc_ref_year={} order by document_number desc limit 1".format(703,709,doc_year)
		with closing(connection.cursor()) as cursor:
			cursor.execute(query)
			row=cursor.fetchone()
			if not row or not len(row):
				return row
	else:
		with closing(connection.cursor()) as cursor:
			cursor.execute("select document_number::integer from xl_ats_document_reference_number where document_type=%s and doc_ref_year=%s order by document_number desc limit 1",[doc_type,doc_year])
			row=cursor.fetchone()
			if not row or not len(row):
				return row
	return row[0]

def insert_doc_number(document_number,document_type,issued_to,issued_by,issue_date,issue_purpose,manual_entry,doc_ref_year):
	with closing(connection.cursor()) as cursor:
		cursor.execute("insert into xl_ats_document_reference_number(document_number,document_type,issued_to,issued_by,issue_date,issue_purpose,manual_entry,doc_ref_year) values(%s,%s,%s,%s,%s,%s,%s,%s)",[document_number,document_type,issued_to,issued_by,issue_date,issue_purpose,manual_entry,doc_ref_year])

def print_relieving_letter(request,ap_id):
	user_id_val = XlplatAuthMiddleware.user_id
	if user_id_val != 0:
		final_val = 1
	else: 
		final_val = 0
	if final_val == 1 and user_id_val != ap_id:
		hr_view=str(check_for_permission_assigned(user_id_val))
		if hr_view == "1":
			if request.method == "POST":
				revert_address="/py/xlplat_ats/relieving_letter/{}".format(ap_id)
				resignation_approved=""
				if 'print_with_header' in request.POST:
					header_flag=1
				else:
					header_flag=0
				ap_name=fetch_employee_name(ap_id)
				applicant_details=employee_details(ap_id)
				ttc_code=applicant_details[3]
				if "TTC-" in ttc_code or "TTC " in ttc_code:
					organisation_name="TALWAR & TALWAR CONSULTANTS Pvt. Ltd." 
					organisation_name_another="<font style='font-size:12px'>Talwar & Talwar Consultants Pvt. Ltd.</font>" 
					image_name="TT-Consultants-patent-Logo.png" 
					registered_address="<font style='font-size:12px'>SCO-304,Sector 38D,Chandigarh,India -160037</font>" 
					extra_style=""
					extra_emails="<b>hr@ttconsultants.com</b>" 
					organisation_unit="TTC"
					organisation_category_id=10000105 
					pan_no="AACCT5126D" 
					gst_no="03AACCT5126D1ZW" 
					iec_no="2206003481"
				elif "XLP" in ttc_code:
					# organisation_name="XLPAT TT CONSULTANTS PRIVATE LIMITED"
					# organisation_name_another="<font style='font-size:12px'>XLPAT TT CONSULTANTS PVT LTD<font style='font-size:12px'>" 
					# image_name="XLPAT.png" 
					# extra_emails="<b>hr@ttconsultants.com</b>"
					# registered_address="<font style='font-size:12px'>SCO-304,Sector 38D,Chandigarh,India -160037</font>"  
					# extra_style=""
					# organisation_unit="XLP"
					# organisation_category_id=10000106 
					# pan_no="AAACX1235B" 
					# gst_no="03AAACX1235B1Z4" 
					# iec_no=""
					organisation_name="XLSCOUT XLPAT PRIVATE LIMITED"
					organisation_name_another="<font style='font-size:12px'>XLSCOUT XLPAT Pvt. Ltd<font style='font-size:12px'>" 
					image_name="xlscout.png" 
					extra_emails="<b>hr@xlscout.ai</b><br>(formerly known as XLPAT TT Consultants Pvt. Ltd.)"
					registered_address="<font style='font-size:12px'>HOUSE NO 580, SECTOR 18B, CHANDIGARH -160018</font>"  
					extra_style=""
					organisation_unit="XLS"
					organisation_category_id=10000106
					pan_no="AAACX1235B" 
					gst_no="03AAACX1235B1Z4" 
					iec_no="2216904571"

				elif "TTCS" in ttc_code or "TTCs" in ttc_code:
					organisation_name="TALWAR & TALWAR CONSULTANTS AND SERVICES Pvt. Ltd." 
					organisation_name_another="<font style='font-size:12px'>Talwar & Talwar Consultants and Services Pvt. Ltd.</font>" 
					# image_name="ttcs_img.png" 
					image_name="TT-Consultants-patent-Logo.png" 
					extra_emails="<b>hr@ttconsultants.com</b>"
					registered_address="<font style='font-size:12px'>SCO-304,Sector 38D,Chandigarh,India-160037</font>"  
					extra_style=""
					organisation_unit="TTCS"
					organisation_category_id=10000107 
					pan_no="AADCT5236H" 
					gst_no="" 
					iec_no="2212000146"
				elif "TA" in ttc_code:
					organisation_name_another="Talwar Advocates" 
					image_name="ta_final_logo.jpg" 
					registered_address="413-P,Sector 6,Panchkula,Haryana-134109."
					extra_emails="<b>prosecution@talwaradvocates.com</b>" 
					extra_style=""
					organisation_unit="TA"
					organisation_category_id=10000106 
					pan_no="ADWPT8331Q" 
					gst_no="" 
					iec_no=""
				elif "XLS" in ttc_code:
					organisation_name="XLSCOUT XLPAT PRIVATE LIMITED"
					organisation_name_another="<font style='font-size:12px'>XLSCOUT XLPAT Pvt. Ltd<font style='font-size:12px'>" 
					image_name="xlscout.png" 
					extra_emails="<b>hr@xlscout.ai</b><br>(formally known as XLPAT TT Consultants Pvt. Ltd.)"
					registered_address="<font style='font-size:12px'>HOUSE NO 580, SECTOR 18B, CHANDIGARH -160018</font>"  
					extra_style=""
					organisation_unit="XLS"
					organisation_category_id=10000106
					pan_no="AAACX1235B" 
					gst_no="03AAACX1235B1Z4" 
					iec_no="2216904571"

				name=""  
				address="" 
				city="" 
				pincode="" 
				phone="" 
				mobile="" 
				website="" 
				cin_no="" 
				pan_no_employer=""
				footer_actual=""
				if header_flag == 1:
					payee_details=get_payee_details(organisation_category_id)
					if payee_details != None and None not in payee_details:
						name=payee_details[0]  
						address=payee_details[1] 
						city=payee_details[2] 
						pincode=payee_details[3] 
						phone=payee_details[4] 
						mobile=payee_details[5] 
						website=payee_details[6] 
						cin_no=payee_details[7] 
						pan_no_employer=payee_details[8] 

					if "TTC-" in ttc_code or "TTC " in ttc_code:
						footer_actual="<font style='font-size:12px'><b class='change_color'>CIN:</b>&nbsp;&nbsp;{}&nbsp;&nbsp;|&nbsp;&nbsp;<b class='change_color'>IEC:</b>&nbsp;&nbsp;{}&nbsp;&nbsp;|&nbsp;&nbsp;<b class='change_color'>PAN:</b>&nbsp;&nbsp;{}&nbsp;&nbsp;|&nbsp;&nbsp;<b class='change_color'>GST:</b>{}</font>".format(cin_no,iec_no,pan_no_employer,gst_no)
						extra_style="width: 50vh"
					elif "XLP" in ttc_code:
						footer_actual="<b class='change_color'>PAN:</b>&nbsp;&nbsp;{}&nbsp;&nbsp;|&nbsp;&nbsp;<b class='change_color'>GST:</b>{}".format(pan_no_employer,gst_no)
						extra_style="width: 25vh"
					elif "TTCS" in ttc_code or "TTCs" in ttc_code:
						footer_actual="<font style='font-size:12px'><b class='change_color'>CIN:</b>&nbsp;&nbsp;{}&nbsp;&nbsp;|&nbsp;&nbsp;<b class='change_color'>IEC:</b>&nbsp;&nbsp;{}&nbsp;&nbsp;|&nbsp;&nbsp;<b class='change_color'>PAN:</b>&nbsp;&nbsp;{}</font>".format(cin_no,iec_no,pan_no,gst_no) 
						extra_style="width: 43vh"
					elif "TA" in ttc_code:
						footer_actual="<font style='font-size:12px'><b class='change_color'>PAN:</b>&nbsp;&nbsp;{}</font>".format(pan_no_employer) 
						extra_style="width: 33vh"
					elif "XLS" in ttc_code:
						footer_actual="<font style='font-size:12px'><b class='change_color'>GSTIN:</b>&nbsp;&nbsp;{}&nbsp;&nbsp;|&nbsp;&nbsp;<b class='change_color'>PAN:</b>&nbsp;&nbsp;{}&nbsp;&nbsp;|&nbsp;&nbsp;<b class='change_color'>CIN:</b>&nbsp;&nbsp;{}&nbsp;&nbsp;|&nbsp;&nbsp;<b class='change_color'>IEC:</b>{}</font>".format(gst_no,pan_no,cin_no,iec_no)
						extra_style="width: 25vh"
				try:
					relieve_temp=Resignation_data.objects.get(resignation_of=ap_id,deleted=False).relieving_template
				except Resignation_data.DoesNotExist:
					relieve_temp=""
				return render(request,'print_relieving_letter.html',{"applicant_id":ap_id,"template_avail":relieve_temp,"address":address,"city":city,"pincode":pincode,"phone":phone,"website":website,"extra_emails":extra_emails,"organisation_name_another":organisation_name_another,"footer_val":footer_actual,"header_flag":header_flag,"image_name":image_name,"registered_address":registered_address,"extra_style":extra_style})
			else:
				return render(request, 'insufficient_privilages.html', {'redirect_to':'/py/xlplat_ats/relieving_letter/{}/'.format(ap_id)})	
		else:
			return render(request, 'insufficient_privilages.html', {'redirect_to':'/py/xlplat_ats/relieving_letter/{}/'.format(ap_id)})	
	else:
		return render(request, 'insufficient_privilages.html', {'redirect_to':'/py/xlplat_ats/relieving_letter/{}'.format(ap_id)})


def send_mail_reminder(user_id,type_val,date,flag):
	if flag == 1:
		with closing(connection.cursor()) as cursor:
			cursor.execute("insert into im_notification_list(object_id,type,sent,deleted,date,time) values(%s,'exit_apply',false,false,%s,%s)",[user_id,date,type_val])
	elif flag==0:
		with closing(connection.cursor()) as cursor:
			cursor.execute("delete from im_notification_list where object_id=%s and type='exit_apply'",[user_id])

def get_payee_details(dept_id):
	with closing(connection.cursor()) as cursor:
		cursor.execute("select name,address,city,pincode,phone,mobile,website,cin_no,pan_no_employer from im_payroll_payee_firm where category_id='%s'",[dept_id])
		row=cursor.fetchone()
	return row
	

def send_mail(users,user_id,permission):
	send_data="{}~{}~{}".format(users,user_id,permission)
	url="http://localhost//intranet/testing/permission-redirect"
	payload = {"proc_name": "send_mail_exit", "proc_variables": send_data}
	response=requests.get(url, params=payload)

def employee_details(ap_id):
	with closing(connection.cursor()) as cursor:
		cursor.execute("select acs_object__name(department_id) as department_id,job_title,immediate_supervisor_id,ttc_code,supervisor_id from im_employees where employee_id=%s",[ap_id])
		row=cursor.fetchone()
	return row

def category_data(category_id):
	url="http://localhost//intranet/testing/permission-redirect"
	payload = {"proc_name": "im_category_from_id", "proc_variables": category_id}
	response=requests.get(url, params=payload)
	response=response.text
	return response  

def check_supervisor(user_id,query,user_id_list=""):
	if query == 1:
		with closing(connection.cursor()) as cursor:
			cursor.execute("select employee_id from im_employees where supervisor_id=%s limit 1",[user_id])
			row=cursor.fetchone()
			if not row or not len(row):
				return row
		return row[0]
	elif query == 2:
		query="select employee_id from im_employees where supervisor_id={} and employee_id in ({})".format(user_id,user_id_list)
		with closing(connection.cursor()) as cursor:
			cursor.execute(query)
			row=[item[0] for item in cursor.fetchall()]
		return row

def get_candidate_id(user_id):
	with closing(connection.cursor()) as cursor:
		cursor.execute("select c_id from xl_ats_req_candidate_map where user_id=%s limit 1",[user_id])
		row=cursor.fetchone()
		if not row or not len(row):
			return row
	return row[0]

def get_supervisor(user_id):
	with closing(connection.cursor()) as cursor:
		cursor.execute("select supervisor_id from im_employees where employee_id=%s limit 1",[user_id])
		row=cursor.fetchone()
		if not row or not len(row):
			return row
	return row[0]
def get_imm_supervisor(user_id):
	with closing(connection.cursor()) as cursor:
		cursor.execute("select immediate_supervisor_id from im_employees where employee_id=%s limit 1",[user_id])
		row=cursor.fetchone()
		if not row or not len(row):
			return row
	return row[0]
def get_loan_approved_amt(user_id):
	with closing(connection.cursor()) as cursor:
		cursor.execute("select sum(approved_amount) from im_payroll_loan_reimb where user_id={}".format(user_id))
		row=cursor.fetchone()
		if not row or not len(row):
			return row
	return row[0]
def get_loan_paid_amt(user_id):
	with closing(connection.cursor()) as cursor:
		cursor.execute("select sum(amount) from im_payroll_allowance_deductions where narration like 'Loan::%' and user_id={}".format(user_id))
		row=cursor.fetchone()
	if not row or not len(row):
		return row
	return row[0]
def get_greeting_card_msg(user_id_for):
	with closing(connection.cursor()) as cursor:
		cursor.execute("select messagecharacter from resign_messages where user_id_for={}".format(user_id_for))
		row=cursor.fetchone()
		if not row or not len(row):
			return row
	return row[0]
def get_greeting_card_by(user_id_for):
	with closing(connection.cursor()) as cursor:
		cursor.execute("select im_name_from_user_id({})".format(user_id_for))
		row=cursor.fetchone()
		if not row or not len(row):
			return row
	return row[0]
def get_server_folder_path():
	url="http://localhost//intranet/testing/permission-redirect"
	payload = {"web_folder": "get folder"}
	response=requests.get(url, params=payload)
	response=response.text
	return response

def get_leaves_left(user_id,last_date):
	url="http://localhost//intranet/testing/permission-redirect"
	payload = {"proc_name": "get_current_leave_balance_exit","extra_parameter":"-params","proc_variables": "{}_{}".format(user_id,last_date) }
	response=requests.get(url, params=payload)
	response=response.text
	return response 

def permission_check(user_id,position):
	url="http://localhost//intranet/testing/permission-redirect"
	payload = {"permission_check": "{}_{}".format(user_id,position)}
	response=requests.get(url, params=payload)
	response=response.text
	return response

@register.filter
def get_range(value):
	return range(value)

@register.filter
def get_item_list(list, key):
    return list[key] 

@register.filter
def split_string(str1,seprator):
	r_str=str1.split(seprator)
	return r_str

@register.filter
def reduce_date(date,days):
	if type(date)==str:
		date=datetime.strptime(date,'%d %b %Y').date()
	date=date - timedelta(days=days)
	return date

@register.filter
def date_checker(date1,date2,sub_days=0):
	# print(date1,date2,"------")
	if type(date1)==str:
		# print(date1,"oo")
		date1=datetime.strptime(date1,'%d %b %Y').date()
	if type(date2)==str:
		# print(date2,"||")
		date2=datetime.strptime(date2,'%d %b %Y').date()
	if sub_days!=0:
		date1=date1 - timedelta(days=sub_days)
	if date1>date2:
		# print("1")
		return 1
	elif date2>date1:
		return 2
	elif date2 == date1:
		return 3
	else:
		return 0
